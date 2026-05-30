"""
导入师资表数据
- 从 Excel 导入 295 条老师记录
- 追加现有 3 位 tutor（如不在表中）
- 回填已有排课的 teacher_id
"""
import sqlite3
import os
import sys

# 添加 backend 目录到路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Excel 路径
EXCEL_PATH = os.path.expanduser(
    "~/Library/Containers/com.tencent.xinWeChat/Data/Documents/"
    "xwechat_files/zhenyahate_efc8/temp/drag/"
    "鱼之跃-师资背景介绍（含等级）.xlsx"
)

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "tms.db")


def import_excel():
    try:
        import openpyxl
    except ImportError:
        print("❌ 需要 openpyxl: pip install openpyxl")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["现有师资明细表"]

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    # 清空重导（仅首次）
    conn.execute("DELETE FROM teachers")

    count = 0
    for r in range(2, ws.max_row + 1):
        name = (ws.cell(r, 2).value or "").strip()
        if not name:
            continue

        conn.execute(
            """INSERT INTO teachers
               (name, academic_background, highest_degree, subjects,
                teaching_direction, tools, teaching_style, level,
                pay_rate, payment_method, notes)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (
                name,
                str(ws.cell(r, 3).value or ""),
                str(ws.cell(r, 4).value or ""),
                str(ws.cell(r, 5).value or ""),
                str(ws.cell(r, 6).value or ""),
                str(ws.cell(r, 7).value or ""),
                str(ws.cell(r, 8).value or ""),
                str(ws.cell(r, 9).value or ""),
                str(ws.cell(r, 10).value or ""),
                str(ws.cell(r, 11).value or ""),
                str(ws.cell(r, 12).value or ""),
            ),
        )
        count += 1

    conn.commit()
    print(f"✅ 从 Excel 导入 {count} 条老师记录")

    # 追加现有 tutor（如表中没有）
    tutors = conn.execute(
        "SELECT id, display_name, phone FROM users WHERE role='tutor'"
    ).fetchall()

    added = 0
    for t in tutors:
        exists = conn.execute(
            "SELECT id FROM teachers WHERE name=?", (t["display_name"],)
        ).fetchone()
        if not exists:
            conn.execute(
                "INSERT INTO teachers (name, phone, notes) VALUES (?,?,?)",
                (t["display_name"], t["phone"] or "", "源自旧系统"),
            )
            print(f"  ➕ 追加: {t['display_name']}")
            added += 1

    conn.commit()
    print(f"✅ 追加 {added} 位现有老师")

    # 回填已有排课的 teacher_id
    schedules = conn.execute(
        """SELECT s.id, s.tutor_id, u.display_name as tutor_name
           FROM schedules s
           LEFT JOIN users u ON s.tutor_id = u.id
           WHERE s.teacher_id IS NULL AND s.tutor_id IS NOT NULL"""
    ).fetchall()

    filled = 0
    for s in schedules:
        teacher = conn.execute(
            "SELECT id FROM teachers WHERE name=? LIMIT 1",
            (s["tutor_name"],),
        ).fetchone()
        if teacher:
            conn.execute(
                "UPDATE schedules SET teacher_id=? WHERE id=?",
                (teacher["id"], s["id"]),
            )
            filled += 1

    conn.commit()
    print(f"✅ 回填 {filled} 条排课的 teacher_id")

    # 验证
    total = conn.execute("SELECT COUNT(*) as c FROM teachers").fetchone()["c"]
    enrolled_active = conn.execute(
        "SELECT COUNT(*) as c FROM teachers WHERE active=1"
    ).fetchone()["c"]
    print(f"\n📊 师资库总计: {total} 人（活跃: {enrolled_active} 人）")

    conn.close()
    print("🎉 导入完成")


if __name__ == "__main__":
    import_excel()
