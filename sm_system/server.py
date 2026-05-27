#!/usr/bin/env python3
"""
TMS — Tutoring Management System (was SM System)
Python stdlib only (http.server + sqlite3)
"""
import json, os, sys, time, hashlib, secrets, sqlite3, uuid, re, hmac
import urllib.parse
from http.server import HTTPServer, BaseHTTPRequestHandler
from http import HTTPStatus
from datetime import datetime, date, timedelta

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'sm.db')
STATIC_DIR = os.path.join(BASE_DIR, 'static')

# AI report generation (optional — requires study-abroad-ai/.env with DEEPSEEK_API_KEY)
try:
    from sm_system.ai_reports import generate_report, REPORT_TYPES
    _AI_AVAILABLE = True
except Exception as e:
    _AI_AVAILABLE = False
    print(f"[TMS] AI report module not loaded: {e}")

SCHEMA_VERSION = 8

# ─── Database ───────────────────────────────────────────────────────────────

def get_db():
    db = sqlite3.connect(DB_PATH, timeout=10)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys=ON")
    return db


def _run_migrations(db, old_version):
    """Run schema migrations sequentially."""
    if old_version < 2:
        db.executescript("""
            CREATE TABLE IF NOT EXISTS sessions (
                token TEXT PRIMARY KEY,
                user_id TEXT NOT NULL REFERENCES users(id),
                expiry INTEGER NOT NULL,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_sessions_expiry ON sessions(expiry);

            ALTER TABLE users ADD COLUMN classin_account TEXT DEFAULT '';
            ALTER TABLE users ADD COLUMN last_login_at TEXT;
            ALTER TABLE users ADD COLUMN is_tutor INTEGER DEFAULT 0;
            ALTER TABLE users ADD COLUMN subjects TEXT DEFAULT '[]';
            ALTER TABLE users ADD COLUMN timezone TEXT DEFAULT 'Asia/Shanghai';
            ALTER TABLE users ADD COLUMN rate REAL DEFAULT 0;

            ALTER TABLE leads ADD COLUMN classin_account TEXT DEFAULT '';
            ALTER TABLE leads ADD COLUMN referral_source_id TEXT REFERENCES leads(id);
            ALTER TABLE leads ADD COLUMN pool_status TEXT DEFAULT 'active';
            ALTER TABLE leads ADD COLUMN pool_return_count INTEGER DEFAULT 0;

            CREATE TABLE IF NOT EXISTS course_packages (
                id TEXT PRIMARY KEY,
                lead_id TEXT NOT NULL REFERENCES leads(id),
                package_name TEXT DEFAULT '',
                total_hours REAL NOT NULL DEFAULT 0,
                used_hours REAL DEFAULT 0,
                price REAL DEFAULT 0,
                unit_price REAL DEFAULT 0,
                valid_from TEXT,
                valid_until TEXT,
                status TEXT DEFAULT 'active',
                classin_course_id TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_packages_lead ON course_packages(lead_id);

            CREATE TABLE IF NOT EXISTS tutor_availability (
                id TEXT PRIMARY KEY,
                tutor_id TEXT NOT NULL REFERENCES users(id),
                weekday INTEGER NOT NULL,
                start_slot TEXT NOT NULL,
                end_slot TEXT NOT NULL,
                timezone TEXT DEFAULT 'Asia/Shanghai',
                effective_from TEXT,
                effective_until TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_tutor_avail ON tutor_availability(tutor_id, weekday);

            CREATE TABLE IF NOT EXISTS schedules (
                id TEXT PRIMARY KEY,
                lead_id TEXT NOT NULL REFERENCES leads(id),
                tutor_id TEXT REFERENCES users(id),
                package_id TEXT REFERENCES course_packages(id),
                coordinator_id TEXT REFERENCES users(id),
                type TEXT DEFAULT 'single',
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                timezone TEXT DEFAULT 'Asia/Shanghai',
                duration_minutes INTEGER DEFAULT 60,
                status TEXT DEFAULT 'pending',
                topic TEXT DEFAULT '',
                notes TEXT DEFAULT '',
                classin_lesson_id TEXT DEFAULT '',
                classin_room_url TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime')),
                updated_at TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_schedules_lead ON schedules(lead_id);
            CREATE INDEX IF NOT EXISTS idx_schedules_tutor ON schedules(tutor_id);
            CREATE INDEX IF NOT EXISTS idx_schedules_time ON schedules(start_time);
            CREATE INDEX IF NOT EXISTS idx_schedules_status ON schedules(status);

            CREATE TABLE IF NOT EXISTS lesson_series (
                id TEXT PRIMARY KEY,
                lead_id TEXT NOT NULL REFERENCES leads(id),
                tutor_id TEXT REFERENCES users(id),
                coordinator_id TEXT REFERENCES users(id),
                package_id TEXT REFERENCES course_packages(id),
                subject TEXT DEFAULT '',
                recurrence TEXT DEFAULT 'weekly',
                day_of_week TEXT DEFAULT '',
                time_of_day TEXT DEFAULT '',
                timezone TEXT DEFAULT 'Asia/Shanghai',
                start_date TEXT,
                end_date TEXT,
                total_lessons INTEGER DEFAULT 0,
                status TEXT DEFAULT 'active',
                notes TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS consumption_log (
                id TEXT PRIMARY KEY,
                package_id TEXT NOT NULL REFERENCES course_packages(id),
                schedule_id TEXT REFERENCES schedules(id),
                hours_scheduled REAL DEFAULT 0,
                hours_actual REAL DEFAULT 0,
                hours_consumed REAL DEFAULT 0,
                classin_raw_data TEXT DEFAULT '',
                confirmed_by TEXT REFERENCES users(id),
                confirmed_at TEXT,
                status TEXT DEFAULT 'pending_confirm',
                notes TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_consumption_pkg ON consumption_log(package_id);
            CREATE INDEX IF NOT EXISTS idx_consumption_status ON consumption_log(status);

            CREATE TABLE IF NOT EXISTS referrals (
                id TEXT PRIMARY KEY,
                referrer_lead_id TEXT NOT NULL REFERENCES leads(id),
                referred_lead_id TEXT REFERENCES leads(id),
                status TEXT DEFAULT 'pending',
                reward_type TEXT DEFAULT '',
                reward_amount REAL DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                enrolled_at TEXT
            );

            CREATE TABLE IF NOT EXISTS pool_return_log (
                id TEXT PRIMARY KEY,
                lead_id TEXT NOT NULL REFERENCES leads(id),
                reason TEXT DEFAULT '',
                returned_by TEXT REFERENCES users(id),
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS audit_log (
                id TEXT PRIMARY KEY,
                user_id TEXT REFERENCES users(id),
                action TEXT NOT NULL,
                resource_type TEXT DEFAULT '',
                resource_id TEXT DEFAULT '',
                detail TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_audit_time ON audit_log(created_at);
        """)
    if old_version < 3:
        db.executescript("""
            CREATE TABLE IF NOT EXISTS ai_reports (
                id TEXT PRIMARY KEY,
                lead_id TEXT NOT NULL REFERENCES leads(id),
                report_type TEXT NOT NULL,
                schedule_id TEXT REFERENCES schedules(id),
                title TEXT DEFAULT '',
                content TEXT DEFAULT '',
                raw_data TEXT DEFAULT '',
                status TEXT DEFAULT 'draft',
                created_by TEXT REFERENCES users(id),
                created_at TEXT DEFAULT (datetime('now','localtime')),
                updated_at TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_ai_reports_lead ON ai_reports(lead_id, report_type);

            ALTER TABLE schedules ADD COLUMN classin_summary TEXT DEFAULT '';
            ALTER TABLE schedules ADD COLUMN teacher_feedback TEXT DEFAULT '';
        """)
    if old_version < 4:
        db.executescript("""
            CREATE TABLE IF NOT EXISTS pool_config (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            INSERT OR IGNORE INTO pool_config (key, value) VALUES ('timeout_days', '7');
            INSERT OR IGNORE INTO pool_config (key, value) VALUES ('max_returns', '3');
            INSERT OR IGNORE INTO pool_config (key, value) VALUES ('silent_pool_timeout', '30');
            INSERT OR IGNORE INTO pool_config (key, value) VALUES ('source_timeouts', '{}');
            INSERT OR IGNORE INTO pool_config (key, value) VALUES ('assignment_strategy', 'manual');
            INSERT OR IGNORE INTO pool_config (key, value) VALUES ('daily_claim_limit', '10');
            INSERT OR IGNORE INTO pool_config (key, value) VALUES ('active_lead_limit', '30');
            db.execute("CREATE TABLE IF NOT EXISTS assignment_rules (
                id TEXT PRIMARY KEY, priority INTEGER DEFAULT 0,
                condition_field TEXT, condition_value TEXT,
                assignee_id TEXT REFERENCES users(id),
                is_active INTEGER DEFAULT 1,
                created_at TEXT DEFAULT '')");

            CREATE TABLE IF NOT EXISTS contracts (
                id TEXT PRIMARY KEY,
                lead_id TEXT NOT NULL REFERENCES leads(id),
                contract_no TEXT NOT NULL,
                total_amount REAL DEFAULT 0,
                paid_amount REAL DEFAULT 0,
                status TEXT DEFAULT 'active',
                signed_at TEXT,
                valid_from TEXT,
                valid_until TEXT,
                notes TEXT DEFAULT '',
                created_by TEXT REFERENCES users(id),
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_contracts_lead ON contracts(lead_id);

            CREATE TABLE IF NOT EXISTS payments (
                id TEXT PRIMARY KEY,
                contract_id TEXT NOT NULL REFERENCES contracts(id),
                amount REAL NOT NULL,
                pay_method TEXT DEFAULT '',
                pay_at TEXT,
                notes TEXT DEFAULT '',
                created_by TEXT REFERENCES users(id),
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_payments_contract ON payments(contract_id);

            ALTER TABLE leads ADD COLUMN merge_target_id TEXT REFERENCES leads(id);
            ALTER TABLE leads ADD COLUMN merged_at TEXT;

            ALTER TABLE schedules ADD COLUMN consume_status TEXT DEFAULT 'pending_confirm';
        """)
    if old_version < 5:
        db.executescript("""
            CREATE TABLE IF NOT EXISTS tutor_settlements (
                id TEXT PRIMARY KEY,
                tutor_id TEXT NOT NULL REFERENCES users(id),
                period_start TEXT NOT NULL,
                period_end TEXT NOT NULL,
                total_lessons INTEGER DEFAULT 0,
                total_hours REAL DEFAULT 0,
                rate REAL DEFAULT 0,
                gross_amount REAL DEFAULT 0,
                deduction REAL DEFAULT 0,
                net_amount REAL DEFAULT 0,
                status TEXT DEFAULT 'pending',
                paid_at TEXT,
                notes TEXT DEFAULT '',
                created_by TEXT REFERENCES users(id),
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS commissions (
                id TEXT PRIMARY KEY,
                contract_id TEXT NOT NULL REFERENCES contracts(id),
                user_id TEXT NOT NULL REFERENCES users(id),
                contract_amount REAL DEFAULT 0,
                commission_rate REAL DEFAULT 0,
                commission_amount REAL DEFAULT 0,
                status TEXT DEFAULT 'pending',
                paid_at TEXT,
                notes TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS notifications (
                id TEXT PRIMARY KEY,
                user_id TEXT NOT NULL REFERENCES users(id),
                type TEXT NOT NULL,
                title TEXT NOT NULL,
                content TEXT DEFAULT '',
                related_type TEXT DEFAULT '',
                related_id TEXT DEFAULT '',
                is_read INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_notif_user ON notifications(user_id, is_read);
        """)
        # ALTER TABLE may fail if columns already exist (idempotent re-run)
        for col in ('original_start_time', 'reschedule_reason'):
            try: db.execute(f"ALTER TABLE schedules ADD COLUMN {col} TEXT DEFAULT ''")
            except: pass

    if old_version < 6:
        # v6: completed_at on schedules, contract_id on course_packages, UTM/campaign on leads
        for col in ('completed_at',):
            try: db.execute(f"ALTER TABLE schedules ADD COLUMN {col} TEXT DEFAULT ''")
            except: pass
        for col in ('contract_id',):
            try: db.execute(f"ALTER TABLE course_packages ADD COLUMN {col} TEXT REFERENCES contracts(id)")
            except: pass
        for col in ('utm_source', 'utm_campaign', 'utm_medium', 'campaign', 'ad_source', 'landing_page'):
            try: db.execute(f"ALTER TABLE leads ADD COLUMN {col} TEXT DEFAULT ''")
            except: pass
        # Add schedule_type to schedules (differentiates trial vs regular vs makeup)
        try: db.execute("ALTER TABLE schedules ADD COLUMN schedule_type TEXT DEFAULT 'regular'")
        except: pass

    if old_version < 7:
        # v7: lead enrichment fields + academic manager assignment
        for col in ('rating', 'lead_type', 'service_type', 'academic_manager_id'):
            try: db.execute(f"ALTER TABLE leads ADD COLUMN {col} TEXT DEFAULT ''")
            except: pass
        try: db.execute("CREATE INDEX IF NOT EXISTS idx_leads_academic ON leads(academic_manager_id)")
        except: pass

    if old_version < 8:
        # v8: account_name + tags (JSON array for service needs)
        for col in ('account_name', 'tags'):
            try: db.execute(f"ALTER TABLE leads ADD COLUMN {col} TEXT DEFAULT ''")
            except: pass
        # Migrate existing account names from notes
        db.executescript("""
            UPDATE leads SET account_name =
                CASE
                    WHEN notes LIKE '%账号:宁振亚（视频号1）%' THEN '宁振亚（视频号1）'
                    WHEN notes LIKE '%账号:留学赋能宁老师（视频号2）%' THEN '留学赋能宁老师（视频号2）'
                    WHEN notes LIKE '%账号:宁振亚（抖音1）%' THEN '宁振亚（抖音1）'
                    WHEN notes LIKE '%账号:易维一留学赋能宁老师（抖音2）%' THEN '易维一留学赋能宁老师（抖音2）'
                    ELSE ''
                END;
            UPDATE leads SET tags =
                CASE
                    WHEN service_type != '' THEN '["""' || service_type || '"""]'
                    ELSE '[]'
                END;
        """)


def init_db():
    db = get_db()

    # Create initial tables (v1)
    db.executescript("""
        CREATE TABLE IF NOT EXISTS schema_version (version INTEGER PRIMARY KEY);
        CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY, username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL, name TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'staff', phone TEXT DEFAULT '',
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS leads (
            id TEXT PRIMARY KEY, name TEXT NOT NULL,
            phone TEXT DEFAULT '', wechat TEXT DEFAULT '',
            country TEXT DEFAULT '', grade TEXT DEFAULT '',
            subject TEXT DEFAULT '', source TEXT DEFAULT '',
            status TEXT DEFAULT 'pending', assignee_id TEXT REFERENCES users(id),
            created_by_id TEXT REFERENCES users(id), notes TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime')),
            assigned_at TEXT DEFAULT ''
        );
        CREATE INDEX IF NOT EXISTS idx_leads_status ON leads(status);
        CREATE INDEX IF NOT EXISTS idx_leads_assignee ON leads(assignee_id);
        CREATE INDEX IF NOT EXISTS idx_leads_created ON leads(created_at);
        CREATE TABLE IF NOT EXISTS activities (
            id TEXT PRIMARY KEY, lead_id TEXT NOT NULL REFERENCES leads(id) ON DELETE CASCADE,
            user_id TEXT NOT NULL REFERENCES users(id),
            type TEXT NOT NULL DEFAULT 'note', content TEXT DEFAULT '',
            next_action TEXT DEFAULT '', next_action_date TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE INDEX IF NOT EXISTS idx_activities_lead ON activities(lead_id);
        CREATE INDEX IF NOT EXISTS idx_activities_next ON activities(next_action_date);
    """)

    # Check version and migrate
    cur = db.execute("SELECT MAX(version) FROM schema_version")
    row = cur.fetchone()
    old_version = row[0] if row and row[0] else 0

    if old_version < SCHEMA_VERSION:
        _run_migrations(db, old_version)
        db.execute("INSERT OR REPLACE INTO schema_version (version) VALUES (?)", (SCHEMA_VERSION,))

    # Seed data if empty
    cur = db.execute("SELECT COUNT(*) FROM users")
    if cur.fetchone()[0] == 0:
        _seed_data(db)

    db.commit()
    db.close()


def _seed_data(db):
    """Seed default users and sample data."""
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    users_data = [
        ('u_admin', 'admin', _hash_pw('admin123'), '管理员', 'admin', '13800000000', '', '', 0, '[]', 'Asia/Shanghai', 0),
        ('u_cs01', 'cs01', _hash_pw('123456'), '客服小王', 'cs', '13800000001', '', '', 0, '[]', 'Asia/Shanghai', 0),
        ('u_cs02', 'cs02', _hash_pw('123456'), '客服小李', 'cs', '13800000002', '', '', 0, '[]', 'Asia/Shanghai', 0),
        ('u_con01', 'con01', _hash_pw('123456'), '顾问张老师', 'consultant', '13800000003', '', '', 0, '[]', 'Asia/Shanghai', 0),
        ('u_con02', 'con02', _hash_pw('123456'), '顾问陈老师', 'consultant', '13800000004', '', '', 0, '[]', 'Asia/Shanghai', 0),
        ('u_mgr01', 'mgr01', _hash_pw('123456'), '学管王老师', 'academic', '13800000005', '', '', 0, '[]', 'Asia/Shanghai', 0),
        ('u_sup01', 'sup01', _hash_pw('123456'), '李主管', 'supervisor', '13800000006', '', '', 0, '[]', 'Asia/Shanghai', 0),
        ('u_tut01', 'tut01', _hash_pw('123456'), '王老师', 'tutor', '13800000010', '', '', 1, '["金融数学","数据科学","统计学"]', 'Asia/Shanghai', 200),
        ('u_tut02', 'tut02', _hash_pw('123456'), '陈老师', 'tutor', '13800000011', '', '', 1, '["会计","经济学","CS"]', 'Asia/Shanghai', 200),
        ('u_tut03', 'tut03', _hash_pw('123456'), '李老师', 'tutor', '13800000012', '', '', 1, '["商务管理","英语"]', 'Europe/London', 250),
        ('u_coor01', 'coor01', _hash_pw('123456'), '教务李老师', 'coordinator', '13800000020', '', '', 0, '[]', 'Asia/Shanghai', 0),
    ]
    db.executemany(
        "INSERT INTO users (id,username,password_hash,name,role,phone,classin_account,last_login_at,is_tutor,subjects,timezone,rate) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        users_data
    )

    samples = [
        ('l_001', '张小明', '13811110001', 'zxm_wechat', '英国', '大三', '金融数学', '抖音', 'enrolled', 'u_con01', 'u_cs01'),
        ('l_002', '李华', '13811110002', 'lihua_wx', '澳洲', '研一', '会计', '小红书', 'enrolled', 'u_con01', 'u_cs01'),
        ('l_003', '王芳', '13811110003', 'wangfang', '美国', '大二', 'CS', '转介绍', 'enrolled', 'u_con02', 'u_cs02'),
        ('l_004', '赵磊', '13811110004', 'zhaolei', '加拿大', '高三', '经济学', '抖音', 'trial', 'u_con01', 'u_admin'),
        ('l_005', '陈雪', '13811110005', 'chenxue', '英国', '大一', '商务管理', '小红书', 'following', 'u_con02', 'u_cs01'),
        ('l_006', '刘洋', '13811110006', 'liuyang', '新加坡', '研二', '数据科学', '线下活动', 'enrolled', 'u_con02', 'u_sup01'),
        ('l_007', '孙丽', '13811110007', 'sunli_wx', '澳洲', '预科', '生物', '抖音', 'following', 'u_con01', 'u_cs01'),
        ('l_008', '周杰', '13811110008', 'zhoujie', '美国', '大四', '机械工程', '转介绍', 'enrolled', 'u_mgr01', 'u_con01'),
    ]
    for s in samples:
        assigned = s[9]
        assigned_at = now if assigned else ''
        db.execute(
            "INSERT INTO leads (id,name,phone,wechat,country,grade,subject,source,status,assignee_id,created_by_id,created_at,updated_at,assigned_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (*s, now, now, assigned_at)
        )

    db.execute(
        "INSERT INTO activities (id,lead_id,user_id,type,content,next_action,next_action_date,created_at) VALUES (?,?,?,?,?,?,?,?)",
        ('a_001', 'l_002', 'u_cs01', 'wechat', '已添加微信，初步了解学生情况', '安排试听课', (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d'), now)
    )

    # Sample packages
    packages = [
        ('p_001', 'l_001', '金融数学 30小时', 30, 18.5, 15000, 500, '2026-01-01', '2026-12-31', 'active'),
        ('p_002', 'l_002', '会计 20小时', 20, 8, 10000, 500, '2026-03-01', '2026-09-30', 'active'),
        ('p_003', 'l_003', 'CS 20小时', 20, 17.5, 10000, 500, '2026-02-01', '2026-08-15', 'active'),
        ('p_004', 'l_004', '经济学 15小时', 15, 15, 7500, 500, '2026-01-15', '2026-06-01', 'exhausted'),
        ('p_005', 'l_006', '数据科学 20小时', 20, 6, 12000, 600, '2026-04-01', '2027-01-15', 'active'),
    ]
    db.executemany(
        "INSERT INTO course_packages (id,lead_id,package_name,total_hours,used_hours,price,unit_price,valid_from,valid_until,status) VALUES (?,?,?,?,?,?,?,?,?,?)",
        packages
    )

    # Tutor availability
    avail = [
        ('av_01', 'u_tut01', 0, '19:00', '23:00', 'Asia/Shanghai', '2026-01-01', None),
        ('av_02', 'u_tut01', 2, '19:00', '23:00', 'Asia/Shanghai', '2026-01-01', None),
        ('av_03', 'u_tut01', 4, '20:00', '23:00', 'Asia/Shanghai', '2026-01-01', None),
        ('av_04', 'u_tut01', 5, '14:00', '18:00', 'Asia/Shanghai', '2026-01-01', None),
        ('av_05', 'u_tut02', 1, '19:00', '22:00', 'Asia/Shanghai', '2026-01-01', None),
        ('av_06', 'u_tut02', 3, '19:00', '22:00', 'Asia/Shanghai', '2026-01-01', None),
        ('av_07', 'u_tut02', 5, '10:00', '14:00', 'Asia/Shanghai', '2026-01-01', None),
        ('av_08', 'u_tut03', 0, '09:00', '17:00', 'Europe/London', '2026-01-01', None),
        ('av_09', 'u_tut03', 2, '09:00', '17:00', 'Europe/London', '2026-01-01', None),
        ('av_10', 'u_tut03', 4, '09:00', '17:00', 'Europe/London', '2026-01-01', None),
    ]
    db.executemany(
        "INSERT INTO tutor_availability (id,tutor_id,weekday,start_slot,end_slot,timezone,effective_from,effective_until) VALUES (?,?,?,?,?,?,?,?)",
        avail
    )

    # Sample schedules
    sched = [
        ('sch_01', 'l_001', 'u_tut01', 'p_001', 'u_coor01', 'single', '2026-05-23T19:00:00+01:00', '2026-05-23T20:00:00+01:00', 'Europe/London', 60, 'confirmed', '金融数学 - 微积分', ''),
        ('sch_02', 'l_002', 'u_tut02', 'p_002', 'u_coor01', 'single', '2026-05-23T12:00:00+10:00', '2026-05-23T13:30:00+10:00', 'Australia/Sydney', 90, 'confirmed', '会计 - 财务报表', ''),
        ('sch_03', 'l_006', 'u_tut01', 'p_005', 'u_coor01', 'recurring', '2026-05-22T18:00:00+10:00', '2026-05-22T19:00:00+10:00', 'Australia/Sydney', 60, 'completed', '数据科学 - Python基础', ''),
        ('sch_04', 'l_003', 'u_tut02', 'p_003', 'u_coor01', 'single', '2026-05-21T10:00:00-04:00', '2026-05-21T11:00:00-04:00', 'America/New_York', 60, 'cancelled', 'CS - 算法', ''),
        ('sch_05', 'l_004', 'u_tut02', 'p_004', 'u_coor01', 'single', '2026-05-23T20:00:00-04:00', '2026-05-23T21:00:00-04:00', 'America/New_York', 60, 'confirmed', '经济学 - 宏观', ''),
    ]
    db.executemany(
        "INSERT INTO schedules (id,lead_id,tutor_id,package_id,coordinator_id,type,start_time,end_time,timezone,duration_minutes,status,topic,notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        sched
    )

    # Sample consumption
    cons = [
        ('c_001', 'p_005', 'sch_03', 1.0, 1.03, 1.0, '{"actual_minutes":62}', 'u_coor01', now, 'confirmed'),
        ('c_002', 'p_001', None, 1.0, 0.97, 1.0, '{"actual_minutes":58}', 'u_coor01', now, 'confirmed'),
    ]
    db.executemany(
        "INSERT INTO consumption_log (id,package_id,schedule_id,hours_scheduled,hours_actual,hours_consumed,classin_raw_data,confirmed_by,confirmed_at,status) VALUES (?,?,?,?,?,?,?,?,?,?)",
        cons
    )

    # Sample series
    db.execute(
        "INSERT INTO lesson_series (id,lead_id,tutor_id,coordinator_id,package_id,subject,recurrence,day_of_week,time_of_day,timezone,start_date,end_date,total_lessons,status) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        ('sr_01', 'l_001', 'u_tut01', 'u_coor01', 'p_001', '金融数学', 'weekly', 'Monday,Wednesday', '20:00', 'Europe/London', '2026-05-01', '2026-08-30', 30, 'active')
    )


# ─── Auth & Session ─────────────────────────────────────────────────────────

def _hash_pw(password):
    salt = secrets.token_hex(12)
    h = hashlib.sha256((salt + password).encode()).hexdigest()
    return f"{salt}${h}"


def _check_pw(password, stored):
    try:
        salt, h = stored.split('$', 1)
        return hashlib.sha256((salt + password).encode()).hexdigest() == h
    except (ValueError, AttributeError):
        return False


login_attempts = {}  # "ip:username" -> {"count": int, "window_start": float}
LOGIN_MAX_ATTEMPTS = 5
LOGIN_WINDOW = 900  # 15 minutes


def _check_login_limit(ip, username):
    key = f"{ip}:{username}"
    now = time.time()
    record = login_attempts.get(key)
    if record and record['window_start'] + LOGIN_WINDOW > now:
        if record['count'] >= LOGIN_MAX_ATTEMPTS:
            return False
    else:
        login_attempts[key] = {'count': 0, 'window_start': now}
    return True


def _record_login_attempt(ip, username):
    key = f"{ip}:{username}"
    now = time.time()
    record = login_attempts.get(key)
    if not record or record['window_start'] + LOGIN_WINDOW < now:
        login_attempts[key] = {'count': 1, 'window_start': now}


# Public API rate limiter (IP-based, separate from login limiter)
public_attempts = {}  # "ip" -> {"count": int, "window_start": float}
PUBLIC_MAX_PER_WINDOW = 5
PUBLIC_WINDOW = 300  # 5 minutes

def _check_public_limit(ip):
    """Returns True if request is allowed, False if rate-limited."""
    now = time.time()
    record = public_attempts.get(ip)
    if record and record['window_start'] + PUBLIC_WINDOW > now:
        if record['count'] >= PUBLIC_MAX_PER_WINDOW:
            return False
        record['count'] += 1
    else:
        public_attempts[ip] = {'count': 1, 'window_start': now}
    return True


# Simple math captcha: generate challenge and verify
import random as _random

def _gen_captcha():
    a = _random.randint(10, 99)
    b = _random.randint(1, 9)
    op = _random.choice(['+', '-'])
    if op == '-':
        a, b = max(a, b), min(a, b)  # ensure non-negative result
    answer = a + b if op == '+' else a - b
    challenge = f"{a} {op} {b} = ?"
    return challenge, str(answer)

# Store active captcha challenges keyed by IP
captcha_store = {}  # "ip" -> {"challenge": str, "answer": str, "expires": float}
CAPTCHA_TTL = 300  # 5 minutes

def _gen_captcha_for_ip(ip):
    chal, ans = _gen_captcha()
    captcha_store[ip] = {'challenge': chal, 'answer': ans, 'expires': time.time() + CAPTCHA_TTL}
    return chal

def _verify_captcha(ip, user_answer):
    record = captcha_store.pop(ip, None)
    if not record:
        return False
    if time.time() > record['expires']:
        return False
    return user_answer.strip() == record['answer']


def _reset_login_limit(ip, username):
    login_attempts.pop(f"{ip}:{username}", None)


def _create_session(user_id):
    token = secrets.token_hex(24)
    expiry = int(time.time()) + 86400 * 3  # 3 days
    db = get_db()
    db.execute("INSERT INTO sessions (token, user_id, expiry) VALUES (?, ?, ?)",
               (token, user_id, expiry))
    db.commit()
    db.close()
    return token


def _get_session_user(token):
    db = get_db()
    row = db.execute("SELECT user_id, expiry FROM sessions WHERE token=?", (token,)).fetchone()
    if not row:
        db.close()
        return None
    if row['expiry'] < time.time():
        db.execute("DELETE FROM sessions WHERE token=?", (token,))
        db.commit()
        db.close()
        return None
    # Sliding expiration: extend by 3 days on each use
    new_expiry = int(time.time()) + 86400 * 3
    db.execute("UPDATE sessions SET expiry=? WHERE token=?", (new_expiry, token))
    db.commit()
    db.close()
    return row['user_id']


def _delete_session(token):
    db = get_db()
    db.execute("DELETE FROM sessions WHERE token=?", (token,))
    db.commit()
    db.close()


def _clean_expired_sessions():
    db = get_db()
    db.execute("DELETE FROM sessions WHERE expiry < ?", (int(time.time()),))
    db.commit()
    db.close()


# ─── Audit Log ──────────────────────────────────────────────────────────────

def _audit(user_id, action, resource_type='', resource_id='', detail=''):
    try:
        db = get_db()
        db.execute("INSERT INTO audit_log (id, user_id, action, resource_type, resource_id, detail) VALUES (?,?,?,?,?,?)",
                   (gen_id('a_'), user_id, action, resource_type, resource_id, detail[:500]))
        db.commit()
        db.close()
    except Exception:
        pass


# ─── Pool Config & Auto-Reclaim ──────────────────────────────────────────────

POOL_DEFAULTS = {'timeout_days': '7', 'max_returns': '3', 'silent_pool_timeout': '30', 'source_timeouts': '{}'}

def _get_pool_config():
    """Load pool configuration from DB, with in-memory cache."""
    try:
        db = get_db()
        rows = db.execute("SELECT key, value FROM pool_config").fetchall()
        db.close()
        cfg = {r['key']: r['value'] for r in rows}
        cfg.setdefault('timeout_days', '7')
        cfg.setdefault('max_returns', '3')
        cfg.setdefault('silent_pool_timeout', '30')
        cfg.setdefault('source_timeouts', '{}')
        cfg.setdefault('assignment_strategy', 'manual')
        cfg.setdefault('daily_claim_limit', '10')
        cfg.setdefault('active_lead_limit', '30')
        return cfg
    except Exception:
        return dict(POOL_DEFAULTS)


def _check_pool_reclaim():
    """Auto-reclaim: return overdue leads to public pool.
    Called before each request. Lightweight due to indexed queries."""
    try:
        cfg = _get_pool_config()
        timeout_days = int(cfg.get('timeout_days', 7))
        max_returns = int(cfg.get('max_returns', 3))
        source_timeouts = json.loads(cfg.get('source_timeouts', '{}'))
        now = datetime.now()

        # Find all active leads past their timeout without recent activity
        # Statuses that should auto-reclaim: assigned, following, trial
        db = get_db()
        rows = db.execute(
            "SELECT l.id, l.status, l.source, l.assignee_id, l.pool_return_count, "
            "COALESCE((SELECT MAX(a.created_at) FROM activities a WHERE a.lead_id=l.id), l.assigned_at, l.created_at) as last_active "
            "FROM leads l "
            "WHERE l.status IN ('assigned','following','trial') "
            "AND l.pool_status='active'"
        ).fetchall()

        now_ts = now.timestamp()
        reclaimed = 0
        for r in rows:
            # Determine timeout for this lead's source
            src_timeout = source_timeouts.get(r['source'], timeout_days)

            # Parse last_active
            try:
                last_dt = datetime.strptime(r['last_active'][:19], '%Y-%m-%d %H:%M')
            except:
                last_dt = now - timedelta(days=src_timeout + 1)
            days_since = (now - last_dt).days

            if days_since >= src_timeout:
                lead_id = r['id']
                return_count = (r['pool_return_count'] or 0) + 1

                if return_count >= max_returns:
                    # Move to silent pool — permanently parked
                    new_status = 'lost'
                    pool_status = 'silent'
                else:
                    new_status = 'pending'
                    pool_status = 'pool'

                now_str = now.strftime('%Y-%m-%d %H:%M')
                db.execute(
                    "UPDATE leads SET status=?, pool_status=?, pool_return_count=?, "
                    "assignee_id=NULL, updated_at=? WHERE id=?",
                    (new_status, pool_status, return_count, now_str, lead_id)
                )
                db.execute(
                    "INSERT INTO pool_return_log (id, lead_id, reason, returned_by, created_at) VALUES (?,?,?,?,?)",
                    (gen_id('pr_'), lead_id,
                     f"自动回收: {days_since}天无跟进(超时{src_timeout}天)，第{return_count}次回池",
                     'system', now_str)
                )
                reclaimed += 1

        if reclaimed:
            db.commit()
            print(f"[Pool] 自动回收 {reclaimed} 条线索")
        db.close()
    except Exception as e:
        print(f"[Pool] auto-reclaim error: {e}")


# ─── Helpers ────────────────────────────────────────────────────────────────

def json_resp(handler, data, status=200):
    body = json.dumps(data, ensure_ascii=False).encode('utf-8')
    handler.send_response(status)
    handler.send_header('Content-Type', 'application/json; charset=utf-8')
    handler.send_header('Access-Control-Allow-Origin', '*')
    handler.send_header('Content-Length', str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def read_body(handler):
    length = int(handler.headers.get('Content-Length', 0))
    if length == 0:
        return {}
    raw = handler.rfile.read(length).decode('utf-8')
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return {}


def gen_id(prefix=''):
    return f"{prefix}{uuid.uuid4().hex[:12]}"


def _levenshtein(s1, s2):
    """Compute Levenshtein edit distance between two strings."""
    if len(s1) < len(s2):
        return _levenshtein(s2, s1)
    if len(s2) == 0:
        return len(s1)
    prev_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        curr_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = prev_row[j + 1] + 1
            deletions = curr_row[j] + 1
            substitutions = prev_row[j] + (c1 != c2)
            curr_row.append(min(insertions, deletions, substitutions))
        prev_row = curr_row
    return prev_row[-1]


# ─── Smart Text Parsing ────────────────────────────────────────────────────

_COUNTRIES = ['英国', '美国', '澳洲', '加拿大', '新加坡', '马来西亚', '澳大利亚', '新西兰', '日本', '韩国', '德国', '法国', '意大利', '西班牙', '荷兰', '瑞士', '香港', '澳门']
_GRADES = ['研一', '研二', '研三', '大一', '大二', '大三', '大四', '高三', '高二', '高一', '初三', '初二', '初一', '预科', '硕士', '博士', '研究生', '本科']
_SOURCES = ['抖音', '小红书', '视频号', '转介绍', '线下活动', '线上', '百度', '知乎', 'B站', '微博', '朋友介绍', '其他']

def parse_lead_text(text):
    """Smart parse lead info from unstructured text.
    Handles formats like:
      DingTalk auto-reply:
        自动回复私信【留学辅导-张老师】
        【回复】Lisa: beij2007
        抖音

      Simple:
        '张三 13800138001 英国金融 大三 抖音'
    """
    result = {'name': '', 'phone': '', 'wechat': '', 'country': '', 'grade': '',
              'subject': '', 'source': '', 'notes': '', 'confidence': 0,
              'account_name': ''}

    t = text.strip()

    # ── 1. Try DingTalk auto-reply format ──
    # Extract source channel (抖音/视频号) - usually at end
    source_found = ''
    for s in _SOURCES:
        if s in t:
            source_found = s
            break

    # Extract video account name: 自动回复私信【账号名】
    acct_match = re.search(r'自动回复私信[（(]?【([^】]+)】', t)
    if acct_match:
        result['account_name'] = acct_match.group(1).strip()
        t = t.replace(acct_match.group(0), ' ').strip()

    # Extract from 【回复】block
    reply_match = re.search(r'【回复】\s*([^：:】]+)\s*[：:]\s*([^\s】]+)', t)
    if reply_match:
        result['name'] = reply_match.group(1).strip()
        contact = reply_match.group(2).strip()
        if re.match(r'^1[3-9]\d{8,9}$', contact):
            result['phone'] = contact
        elif re.match(r'^[a-zA-Z0-9_]{4,}$', contact):
            result['wechat'] = contact
        else:
            digits = re.sub(r'\D', '', contact)
            if len(digits) >= 8:
                result['phone'] = contact
            else:
                result['wechat'] = contact
        t = t.replace(reply_match.group(0), ' ').strip()

    # Remove reply content block
    t = re.sub(r'【回复内容】[：:].*', '', t).strip()
    # Remove assistant name tags
    t = re.sub(r'【[^】]+】', '', t).strip()

    if source_found:
        result['source'] = source_found
        t = t.replace(source_found, ' ').strip()

    # ── 2. Simple format fallback ──
    if not result['name']:
        for prefix in ['新线索', '线索', '新增', '客户', '学生', '姓名', '名字']:
            if t.startswith(prefix):
                t = t[len(prefix):].strip().lstrip('：:，, ')

        phone_match = re.search(r'(1[3-9]\d)\s*-?\s*(\d{4})\s*-?\s*(\d{4})', t)
        if phone_match:
            result['phone'] = phone_match.group(1) + phone_match.group(2) + phone_match.group(3)
            t = t.replace(phone_match.group(0), ' ').strip()
        else:
            phone_match = re.search(r'(1[3-9]\d{7,9})', t)
            if phone_match:
                result['phone'] = phone_match.group(1)
                t = t.replace(phone_match.group(0), ' ').strip()

        wx_match = re.search(r'(?:微信|wx|wechat)[：:=\s]*([a-zA-Z0-9_]{4,})', t, re.IGNORECASE)
        if wx_match:
            result['wechat'] = wx_match.group(1)
            t = t.replace(wx_match.group(0), ' ').strip()

    # ── 3. Extract education info ──
    for c in _COUNTRIES:
        if c in t:
            result['country'] = c
            t = t.replace(c, ' ').strip()
            break

    for g in _GRADES:
        if g in t:
            result['grade'] = g
            t = t.replace(g, ' ').strip()
            break

    subjects = ['金融数学', '数据科学', '统计学', '金融', '会计', '经济学', '商务管理',
                '计算机', 'CS', '机械工程', '电子工程', '生物', '化学', '物理', '数学',
                '法律', '医学', '教育', '心理学', '社会学', '传媒', '艺术', '建筑',
                '英语', '语言', '商业分析', '市场营销', '管理']
    for subj in subjects:
        if subj in t:
            result['subject'] = subj
            t = t.replace(subj, ' ').strip()
            break

    # ── 4. Name (simple fallback) ──
    if not result['name']:
        t = re.sub(r'[，,、/；;：:（(）)]', ' ', t).strip()
        t = re.sub(r'\s+', ' ', t).strip()
        parts = [p.strip() for p in t.split() if p.strip()]
        if parts:
            candidate = parts[0]
            if re.match(r'^[一-龥a-zA-Z]{2,}$', candidate):
                result['name'] = candidate
                parts = parts[1:]
        if parts:
            existing = result.get('notes', '') or ''
            result['notes'] = (existing + ' ' + ' '.join(parts)).strip()

    # ── 5. Build notes ──
    notes_parts = []
    if result.get('notes'):
        notes_parts.append(result['notes'])
    if result['account_name']:
        notes_parts.append(f'短视频账号: {result["account_name"]}')
    if notes_parts:
        result['notes'] = ' | '.join(notes_parts)

    # ── 6. Confidence ──
    confidence = 0
    if result['name']: confidence += 0.3
    if result['phone'] or result['wechat']: confidence += 0.3
    if result['source']: confidence += 0.15
    if result['account_name']: confidence += 0.1
    if result['country']: confidence += 0.1
    if result['subject']: confidence += 0.1
    result['confidence'] = min(round(confidence, 2), 1.0)

    return result


def get_token(handler):
    cookie = handler.headers.get('Cookie', '')
    for c in cookie.split(';'):
        c = c.strip()
        if c.startswith('sm_session='):
            return c.split('=', 1)[1].strip()
    return None


def require_auth(handler):
    token = get_token(handler)
    if not token:
        return None
    uid = _get_session_user(token)
    if not uid:
        return None
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=? AND is_active=1", (uid,)).fetchone()
    db.close()
    return dict(user) if user else None


def parse_path(path):
    """Parse RESTful path patterns. Returns (resource, id, action) or None."""
    parts = [p for p in path.rstrip('/').split('/') if p]
    if len(parts) >= 3 and parts[0] == 'api' and parts[1] == 'sm':
        parts = parts[2:]
    else:
        return None

    # Match patterns
    if len(parts) == 1:
        if parts[0] == 'leads': return ('leads', None, 'list')
        if parts[0] == 'academic': return ('academic', None, 'dashboard')
        if parts[0] == 'enrolled': return ('enrolled', None, 'list')
        if parts[0] == 'dashboard': return ('dashboard', None, 'stats')
        if parts[0] == 'users': return ('users', None, 'list')
        if parts[0] == 'me': return ('me', None, 'get')
        if parts[0] == 'login': return ('auth', None, 'login')
        if parts[0] == 'logout': return ('auth', None, 'logout')
        if parts[0] == 'packages': return ('packages', None, 'list')
        if parts[0] == 'schedules': return ('schedules', None, 'list')
        if parts[0] == 'consumptions': return ('consumptions', None, 'list')
        if parts[0] == 'tutors': return ('tutors', None, 'list')
        if parts[0] == 'series': return ('series', None, 'list')
        if parts[0] == 'coordinator': return ('coordinator', None, 'dashboard')
        if parts[0] == 'tutor-availability': return ('tutor_avail', None, 'list')
        if parts[0] == 'ai': return ('ai', None, 'info')
        if parts[0] == 'ai-reports': return ('ai_reports', None, 'list')
        if parts[0] == 'pool': return ('pool', None, 'list')
        if parts[0] == 'contracts': return ('contracts', None, 'list')
        if parts[0] == 'referrals': return ('referrals', None, 'list')
        if parts[0] == 'payments': return ('payments', None, 'list')
        if parts[0] == 'settlements': return ('settlements', None, 'list')
        if parts[0] == 'commissions': return ('commissions', None, 'list')
        if parts[0] == 'notifications': return ('notifications', None, 'list')
        if parts[0] == 'finance': return ('finance', None, 'report')
    if len(parts) == 1:
        if parts[0] == 'backup': return ('backup', None, 'export')
        if parts[0] == 'backup-list': return ('backup', None, 'list')
        if parts[0] == 'export-excel': return ('export', None, 'excel')
    if len(parts) == 2:
        if parts[0] == 'leads' and parts[1] == 'batch': return ('leads', None, 'batch')
        if parts[0] == 'leads' and parts[1] == 'import': return ('leads', None, 'import_leads')
        if parts[0] == 'leads' and parts[1] == 'parse': return ('leads', None, 'parse_text')
        if parts[0] == 'leads': return ('leads', parts[1], 'get')
        if parts[0] == 'packages': return ('packages', parts[1], 'get')
        if parts[0] == 'schedules' and parts[1] == 'calendar': return ('schedules', None, 'calendar')
        if parts[0] == 'schedules': return ('schedules', parts[1], 'get')
        if parts[0] == 'consumptions': return ('consumptions', parts[1], 'get')
        if parts[0] == 'tutor-availability': return ('tutor_avail', parts[1], 'get')
        if parts[0] == 'ai-reports': return ('ai_reports', parts[1], 'get')
        if parts[0] == 'series': return ('series', parts[1], 'get')
        if parts[0] == 'coordinator': return ('coordinator', None, 'dashboard')
        if parts[0] == 'ai' and parts[1] == 'report-types': return ('ai', None, 'report_types')
        if parts[0] == 'pool' and parts[1] == 'list': return ('pool', None, 'list')
        if parts[0] == 'pool' and parts[1] == 'settings': return ('pool', None, 'settings')
        if parts[0] == 'pool' and parts[1] == 'stats': return ('pool', None, 'stats')
        if parts[0] == 'pool' and parts[1] == 'log': return ('pool', None, 'log')
        if parts[0] == 'contracts': return ('contracts', None, 'list')
        if parts[0] == 'referrals': return ('referrals', None, 'list')
        if parts[0] == 'payments' and parts[1] != '': return ('payments', parts[1], 'list')
        if parts[0] == 'settlements' and parts[1] == 'calculate': return ('settlements', None, 'calculate')
        if parts[0] == 'settlements': return ('settlements', parts[1], 'get')
        if parts[0] == 'commissions' and parts[1] == 'calculate': return ('commissions', None, 'calculate')
        if parts[0] == 'commissions': return ('commissions', parts[1], 'get')
        if parts[0] == 'notifications' and parts[1] == 'read-all': return ('notifications', None, 'read_all')
        if parts[0] == 'notifications' and parts[1] == 'read': return ('notifications', parts[1], 'read')
        if parts[0] == 'notifications' and parts[1] == 'check-reminders': return ('notifications', None, 'check_reminders')
        if parts[0] == 'notifications': return ('notifications', parts[1], 'get')
        if parts[0] == 'tutors' and parts[1] == 'match': return ('tutors', None, 'match')
        if parts[0] == 'assignment' and parts[1] == 'dashboard': return ('assignment', None, 'dashboard')
        if parts[0] == 'assignment' and parts[1] == 'rules': return ('assignment', None, 'rules')
        if parts[0] == 'followup' and parts[1] == 'templates': return ('followup', None, 'templates')
        if parts[0] == 'followup' and parts[1] == 'overdue': return ('followup', None, 'overdue')
        if parts[0] == 'followup' and parts[1] == 'stats': return ('followup', None, 'stats')
    if len(parts) == 3:
        if parts[0] == 'leads' and parts[2] == 'activities': return ('leads', parts[1], 'activities')
        if parts[0] == 'leads' and parts[2] == 'assign': return ('leads', parts[1], 'assign')
        if parts[0] == 'leads' and parts[2] == 'reclaim': return ('leads', parts[1], 'reclaim')
        if parts[0] == 'schedules' and parts[2] == 'confirm': return ('schedules', parts[1], 'confirm')
        if parts[0] == 'schedules' and parts[2] == 'cancel': return ('schedules', parts[1], 'cancel')
        if parts[0] == 'ai' and parts[2] == 'generate': return ('ai', parts[1], 'generate')
        if parts[0] == 'ai' and parts[2] == 'report-types': return ('ai', None, 'report_types')
        if parts[0] == 'ai-reports' and parts[2] == 'publish': return ('ai_reports', parts[1], 'publish')
        if parts[0] == 'leads' and parts[2] == 'merge': return ('leads', parts[1], 'merge')
        if parts[0] == 'pool' and parts[2] == 'return': return ('pool', parts[1], 'return_to_pool')
        if parts[0] == 'contracts' and parts[2] == 'pay': return ('contracts', parts[1], 'pay')
        if parts[0] == 'series' and parts[2] == 'generate': return ('series', parts[1], 'generate')
        if parts[0] == 'schedules' and parts[2] == 'reschedule': return ('schedules', parts[1], 'reschedule')
        if parts[0] == 'schedules' and parts[2] == 'complete': return ('schedules', parts[1], 'complete')
        if parts[0] == 'leads' and parts[2] == 'self-claim': return ('leads', parts[1], 'self_claim')
        if parts[0] == 'leads' and parts[2] == 'assign-academic': return ('leads', parts[1], 'assign_academic')
        if parts[0] == 'notifications' and parts[2] == 'read': return ('notifications', parts[1], 'read')
        if parts[0] == 'followup' and parts[1] == 'templates': return ('followup_templates', parts[2], 'get')


def get_page_params(params):
    try: page = max(1, int(params.get('page', ['1'])[0]))
    except: page = 1
    try: size = min(200, max(1, int(params.get('page_size', ['50'])[0])))
    except: size = 50
    offset = (page - 1) * size
    return page, size, offset


def require_role(user, roles):
    if user['role'] not in roles:
        return False
    return True


# ─── HTTP Handler ───────────────────────────────────────────────────────────

class TMSHandler(BaseHTTPRequestHandler):

    def _send_cors_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, Cookie')

    def do_OPTIONS(self):
        self.send_response(204)
        self._send_cors_headers()
        self.end_headers()

    _tables_ensured = False
    _last_reminder_check = 0

    def _ensure_tables(self):
        if TMSHandler._tables_ensured:
            return
        TMSHandler._tables_ensured = True
        try:
            db = get_db()
            db.execute("CREATE TABLE IF NOT EXISTS assignment_rules (\n"
                "id TEXT PRIMARY KEY, priority INTEGER DEFAULT 0,\n"
                "condition_field TEXT, condition_value TEXT,\n"
                "assignee_id TEXT REFERENCES users(id),\n"
                "is_active INTEGER DEFAULT 1,\n"
                "created_at TEXT DEFAULT '')")
            db.execute("CREATE TABLE IF NOT EXISTS followup_templates (\n"
                "id TEXT PRIMARY KEY, name TEXT, type TEXT,\n"
                "content_template TEXT, sort_order INTEGER DEFAULT 0,\n"
                "is_active INTEGER DEFAULT 1,\n"
                "created_at TEXT DEFAULT '')")
            # Seed default followup templates if empty
            cnt = db.execute("SELECT COUNT(*) FROM followup_templates").fetchone()[0]
            if cnt == 0:
                now = datetime.now().strftime('%Y-%m-%d %H:%M')
                templates = [
                    ('ft_1', '发送资料', 'material', '发送了 {subject} 课程资料，家长表示会考虑', 0),
                    ('ft_2', '意向确认', 'call', '确认了 {country} {subject} 的辅导意向，预算约 XX 元/课时', 1),
                    ('ft_3', '试听后回访', 'trial', '试听 {subject} 课程后反馈：优点：...，待改进：...，下次安排：...', 2),
                    ('ft_4', '签约跟进', 'call', '沟通合同条款，家长 concern：{concern}，已解释/待跟进', 3),
                    ('ft_5', '节后回访', 'wechat', '假期后回访，了解学习进度，下次考试时间：...', 4),
                    ('ft_6', '价格咨询', 'wechat', '家长咨询 {subject} 课程价格，已发送价目表，跟进意向', 5),
                ]
                for t in templates:
                    db.execute("INSERT OR IGNORE INTO followup_templates (id,name,type,content_template,sort_order,is_active,created_at) VALUES (?,?,?,?,?,1,?)",
                               (t[0], t[1], t[2], t[3], t[4], now))
            db.commit()
            db.close()
        except:
            pass

    def _serve_static(self, path):
        if not path or path == '/' or path == '/sm.html':
            filepath = os.path.join(STATIC_DIR, 'sm.html')
        else:
            safe = path.lstrip('/')
            if safe.startswith('static/'):
                safe = safe[7:]
            filepath = os.path.join(STATIC_DIR, safe)
        real = os.path.realpath(filepath)
        if not real.startswith(os.path.realpath(STATIC_DIR)):
            return json_resp(self, {"error": "Forbidden"}, 403)
        if not os.path.exists(real) or os.path.isdir(real):
            return json_resp(self, {"error": "Not Found"}, 404)
        ext = os.path.splitext(real)[1].lower()
        mime_map = {
            '.html': 'text/html; charset=utf-8', '.css': 'text/css; charset=utf-8',
            '.js': 'application/javascript; charset=utf-8', '.json': 'application/json; charset=utf-8',
            '.png': 'image/png', '.jpg': 'image/jpeg', '.svg': 'image/svg+xml', '.ico': 'image/x-icon',
        }
        ct = mime_map.get(ext, 'application/octet-stream')
        try:
            with open(real, 'rb') as f:
                content = f.read()
            self.send_response(200)
            self.send_header('Content-Type', ct)
            self.send_header('Content-Length', str(len(content)))
            self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
            self.send_header('Pragma', 'no-cache')
            self.send_header('Expires', '0')
            self.end_headers()
            self.wfile.write(content)
        except Exception:
            json_resp(self, {"error": "Internal"}, 500)

    def _route(self):
        _clean_expired_sessions()
        self._ensure_tables()
        _check_pool_reclaim()
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        params = urllib.parse.parse_qs(parsed.query)
        method = self.command

        if not path.startswith('/api/'):
            return self._serve_static(path)

        # Public API routes (no auth required)
        if path.startswith('/api/public/'):
            if method == 'POST' and path == '/api/public/lead':
                data = read_body(self)
                return self._handle_public_lead(data)
            if method == 'GET' and path == '/api/public/captcha':
                return self._handle_public_captcha()
            return json_resp(self, {"error": "Not Found"}, 404)

        # DingTalk webhook (no session auth, uses its own signature)
        if path == '/api/dingtalk/webhook':
            data = read_body(self)
            return self._handle_dingtalk_webhook(data)

        r = parse_path(path)
        if not r:
            return json_resp(self, {"error": "Not Found"}, 404)

        resource, res_id, action = r

        # Auth routes (no login required)
        if resource == 'auth' and action == 'login':
            return self._handle_login()
        if resource == 'auth' and action == 'logout':
            return self._handle_logout()

        # All other routes require auth
        user = require_auth(self)
        if not user:
            return json_resp(self, {"error": "Unauthorized"}, 401)

        # Auto-check reminders every 5 minutes (silent, no response)
        now_ts = time.time()
        if now_ts - TMSHandler._last_reminder_check > 300:
            TMSHandler._last_reminder_check = now_ts
            try:
                self._check_reminders_silent(user)
            except:
                pass

        try:
            if method == 'GET':
                return self._handle_get(resource, res_id, action, user, params)
            elif method == 'POST':
                data = read_body(self)
                return self._handle_post(resource, res_id, action, user, data)
            elif method == 'PUT':
                data = read_body(self)
                return self._handle_put(resource, res_id, action, user, data)
            elif method == 'DELETE':
                return self._handle_delete(resource, res_id, action, user)
        except Exception as e:
            return json_resp(self, {"error": str(e)}, 500)

        return json_resp(self, {"error": "Method Not Allowed"}, 405)

    def do_GET(self):
        self._route()

    def do_POST(self):
        self._route()

    def do_PUT(self):
        self._route()

    def do_DELETE(self):
        self._route()

    # ─── GET Router ──────────────────────────────────────────────────────

    def _handle_get(self, resource, res_id, action, user, params):
        handlers = {
            'me': lambda: json_resp(self, {"user": user}),
            'dashboard': lambda: self._handle_dashboard(user),
            'users': lambda: self._handle_list_users(user),
            'leads@list': lambda: self._handle_list_leads(user, params),
            'leads@get': lambda: self._handle_get_lead(user, res_id),
            'leads@activities': lambda: self._handle_list_activities(user, res_id),
            'packages@list': lambda: self._handle_list_packages(user, params),
            'packages@get': lambda: self._handle_get_package(user, res_id),
            'schedules@list': lambda: self._handle_list_schedules(user, params),
            'schedules@get': lambda: self._handle_get_schedule(user, res_id),
            'consumptions@list': lambda: self._handle_list_consumptions(user, params),
            'tutors@list': lambda: self._handle_list_tutors(user),
            'series@list': lambda: self._handle_list_series(user, params),
            'series@get': lambda: self._handle_get_series(user, res_id),
            'coordinator@dashboard': lambda: self._handle_coordinator_dashboard(user),
            'tutor_avail@list': lambda: self._handle_list_tutor_avail(user, params),
            'tutor_avail@get': lambda: self._handle_get_tutor_avail(user, res_id),
            'ai@info': lambda: self._handle_ai_info(user),
            'ai@report_types': lambda: self._handle_ai_report_types(user),
            'ai_reports@list': lambda: self._handle_list_ai_reports(user, params),
            'ai_reports@get': lambda: self._handle_get_ai_report(user, res_id),
            'pool@list': lambda: self._handle_pool_list(user, params),
            'pool@settings': lambda: self._handle_pool_settings(user),
            'pool@stats': lambda: self._handle_pool_stats(user),
            'pool@log': lambda: self._handle_pool_log(user, params),
            'contracts@list': lambda: self._handle_list_contracts(user, params),
            'contracts@get': lambda: self._handle_get_contract(user, res_id),
            'referrals@list': lambda: self._handle_list_referrals(user, params),
            'referrals@get': lambda: self._handle_get_referral(user, res_id),
            'payments@list': lambda: self._handle_list_payments(user, res_id, params),
            'settlements@list': lambda: self._handle_list_settlements(user, params),
            'settlements@get': lambda: self._handle_get_settlement(user, res_id),
            'commissions@list': lambda: self._handle_list_commissions(user, params),
            'commissions@get': lambda: self._handle_get_commission(user, res_id),
            'finance@report': lambda: self._handle_finance_report(user, params),
            'schedules@calendar': lambda: self._handle_calendar_view(user, params),
            'tutors@match': lambda: self._handle_match_tutors(user, params),
            'notifications@list': lambda: self._handle_list_notifications(user, params),
            'notifications@get': lambda: self._handle_get_notification(user, res_id),
            'academic@dashboard': lambda: self._handle_academic_dashboard(user),
            'enrolled@list': lambda: self._handle_enrolled_list(user, params),
            'assignment@dashboard': lambda: self._handle_assignment_dashboard(user),
            'assignment@rules': lambda: self._handle_assignment_rules(user),
            'backup@export': lambda: self._handle_backup_export(user),
            'backup@list': lambda: self._handle_backup_list(user),
            'export@excel': lambda: self._handle_export_excel(user),
            'followup@templates': lambda: self._handle_list_followup_templates(user),
            'followup@overdue': lambda: self._handle_followup_overdue(user, params),
            'followup@stats': lambda: self._handle_followup_stats(user, params),
            'followup_templates@get': lambda: self._handle_get_followup_template(user, res_id),
        }
        key = f"{resource}@{action}"
        alt_key = resource  # fallback: some endpoints register as plain 'me', 'users', etc.
        handler = handlers.get(key) or handlers.get(alt_key)
        if not handler:
            return json_resp(self, {"error": "Not Found"}, 404)
        return handler()

    # ─── POST Router ─────────────────────────────────────────────────────

    def _handle_post(self, resource, res_id, action, user, data):
        handlers = {
            'leads@list': lambda: self._handle_create_lead(user, data),
            'leads@assign': lambda: self._handle_assign_lead(user, res_id, data),
            'leads@activities': lambda: self._handle_add_activity(user, res_id, data),
            'leads@reclaim': lambda: self._handle_reclaim_lead(user, res_id),
            'packages@list': lambda: self._handle_create_package(user, data),
            'schedules@list': lambda: self._handle_create_schedule(user, data),
            'schedules@confirm': lambda: self._handle_confirm_schedule(user, res_id, data),
            'schedules@cancel': lambda: self._handle_cancel_schedule(user, res_id, data),
            'consumptions@list': lambda: self._handle_confirm_consumption(user, res_id, data),
            'tutor_avail@list': lambda: self._handle_create_tutor_avail(user, data),
            'series@list': lambda: self._handle_create_series(user, data),
            'series@generate': lambda: self._handle_generate_series_schedules(user, res_id, data),
            'ai@generate': lambda: self._handle_ai_generate(user, res_id, data),
            'ai_reports@publish': lambda: self._handle_publish_ai_report(user, res_id, data),
            'pool@return_to_pool': lambda: self._handle_pool_return(user, res_id, data),
            'contracts@list': lambda: self._handle_create_contract(user, data),
            'contracts@pay': lambda: self._handle_contract_pay(user, res_id, data),
            'referrals@list': lambda: self._handle_create_referral(user, data),
            'leads@merge': lambda: self._handle_merge_leads(user, res_id, data),
            'leads@parse_text': lambda: self._handle_parse_lead_text(user, data),
            'settlements@calculate': lambda: self._handle_calculate_settlements(user),
            'commissions@calculate': lambda: self._handle_calculate_commissions(user),
            'schedules@reschedule': lambda: self._handle_reschedule_schedule(user, res_id, data),
            'schedules@complete': lambda: self._handle_complete_schedule(user, res_id, data),
            'leads@self_claim': lambda: self._handle_self_claim(user, res_id),
            'leads@assign_academic': lambda: self._handle_assign_academic(user, res_id, data),
            'leads@batch': lambda: self._handle_batch_leads(user, data),
            'leads@import_leads': lambda: self._handle_import_leads(user, data),
            'notifications@read': lambda: self._handle_mark_read(user, res_id),
            'notifications@read_all': lambda: self._handle_mark_read_all(user),
            'notifications@check_reminders': lambda: self._handle_check_reminders(user),
            'assignment@rules': lambda: self._handle_create_assignment_rule(user, data),
            'followup@templates': lambda: self._handle_create_followup_template(user, data),
        }
        key = f"{resource}@{action}"
        handler = handlers.get(key)
        if not handler:
            return json_resp(self, {"error": "Not Found"}, 404)
        return handler()

    # ─── PUT Router ──────────────────────────────────────────────────────

    def _handle_put(self, resource, res_id, action, user, data):
        handlers = {
            'leads@get': lambda: self._handle_update_lead(user, res_id, data),
            'packages@get': lambda: self._handle_update_package(user, res_id, data),
            'schedules@get': lambda: self._handle_update_schedule(user, res_id, data),
            'tutor_avail@get': lambda: self._handle_update_tutor_avail(user, res_id, data),
            'series@get': lambda: self._handle_update_series(user, res_id, data),
            'ai_reports@get': lambda: self._handle_update_ai_report(user, res_id, data),
            'pool@settings': lambda: self._handle_pool_update_settings(user, data),
            'settlements@get': lambda: self._handle_update_settlement(user, res_id, data),
            'commissions@get': lambda: self._handle_update_commission(user, res_id, data),
            'assignment@rules': lambda: self._handle_update_assignment_rule(user, data),
            'followup_templates@get': lambda: self._handle_update_followup_template(user, res_id, data),
        }
        key = f"{resource}@{action}"
        handler = handlers.get(key)
        if not handler:
            return json_resp(self, {"error": "Not Found"}, 404)
        return handler()

    def _handle_delete(self, resource, res_id, action, user):
        handlers = {
            'tutor_avail@get': lambda: self._handle_delete_tutor_avail(user, res_id),
            'series@get': lambda: self._handle_delete_series(user, res_id),
            'followup_templates@get': lambda: self._handle_delete_followup_template(user, res_id),
        }
        key = f"{resource}@{action}"
        handler = handlers.get(key)
        if not handler:
            return json_resp(self, {"error": "Not Found"}, 404)
        return handler()

    # ═══════════════════ AUTH ═══════════════════════════════════════════════

    def _handle_login(self):
        data = read_body(self)
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        if not username or not password:
            return json_resp(self, {"error": "用户名和密码不能为空"}, 400)

        ip = self.client_address[0]
        if not _check_login_limit(ip, username):
            return json_resp(self, {"error": "登录失败次数过多，请15分钟后再试"}, 429)

        db = get_db()
        user = db.execute("SELECT * FROM users WHERE username=? AND is_active=1", (username,)).fetchone()
        db.close()

        if not user or not _check_pw(password, user['password_hash']):
            _record_login_attempt(ip, username)
            return json_resp(self, {"error": "用户名或密码错误"}, 401)

        _reset_login_limit(ip, username)
        token = _create_session(user['id'])

        # Update last login
        db = get_db()
        db.execute("UPDATE users SET last_login_at=? WHERE id=?", (datetime.now().strftime('%Y-%m-%d %H:%M'), user['id']))
        db.commit()
        db.close()

        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Set-Cookie', f'sm_session={token}; Path=/; Max-Age=259200; SameSite=Lax')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps({
            "token": token,
            "user": {"id": user['id'], "name": user['name'], "role": user['role'], "username": user['username']}
        }, ensure_ascii=False).encode('utf-8'))

    def _handle_logout(self):
        token = get_token(self)
        if token:
            _delete_session(token)
        self.send_response(200)
        self.send_header('Set-Cookie', 'sm_session=; Path=/; Max-Age=0')
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        json_resp(self, {"status": "ok"})

    # ═══════════════════ DASHBOARD ═════════════════════════════════════════

    def _handle_dashboard(self, user):
        db = get_db()
        role, uid = user['role'], user['id']

        scope_field, scope_val = ('assignee_id', uid) if role in ('cs', 'consultant', 'academic') else (None, None)

        MERGED_EXCLUDE = " AND l.merge_target_id IS NULL"

        def _scoped_count(extra_where='', extra_params=()):
            if scope_field:
                w = f"WHERE {scope_field}=?" + (f" AND {extra_where}" if extra_where else "")
                p = (scope_val,) + extra_params
            else:
                w = "WHERE 1=1" + (f" AND {extra_where}" if extra_where else "")
                p = extra_params
            return db.execute(f"SELECT COUNT(*) FROM leads l {w} {MERGED_EXCLUDE}", p).fetchone()[0]

        total = _scoped_count()
        pending = _scoped_count("l.status='pending'")
        following = _scoped_count("l.status IN ('assigned','following','trial')")
        enrolled = _scoped_count("l.status='enrolled'")

        today = date.today().isoformat()
        today_fups = db.execute(
            "SELECT a.id, a.next_action, a.next_action_date, a.lead_id, l.name as lead_name "
            "FROM activities a JOIN leads l ON a.lead_id = l.id "
            "WHERE a.next_action_date=? AND a.next_action!='' AND l.merge_target_id IS NULL "
            "ORDER BY a.created_at DESC LIMIT 20", (today,)
        ).fetchall()

        recent = db.execute(
            "SELECT l.*, u.name as assignee_name FROM leads l "
            "LEFT JOIN users u ON l.assignee_id = u.id "
            "WHERE l.merge_target_id IS NULL "
            "ORDER BY l.created_at DESC LIMIT 10"
        ).fetchall()

        pending_count = 0
        pool_count = 0
        if role in ('admin', 'supervisor'):
            pending_count = db.execute("SELECT COUNT(*) FROM leads WHERE status='pending' AND merge_target_id IS NULL").fetchone()[0]
            pool_count = db.execute("SELECT COUNT(*) FROM leads WHERE pool_status='pool' AND merge_target_id IS NULL").fetchone()[0]

        statuses = db.execute(
            "SELECT status, COUNT(*) as cnt FROM leads WHERE merge_target_id IS NULL GROUP BY status ORDER BY cnt DESC"
        ).fetchall()

        # Coordinator quick stats
        coordinator_stats = {}
        if role in ('admin', 'supervisor', 'coordinator'):
            coordinator_stats['pending_schedules'] = db.execute(
                "SELECT COUNT(*) FROM schedules WHERE status='pending'"
            ).fetchone()[0]
            coordinator_stats['pending_consumptions'] = db.execute(
                "SELECT COUNT(*) FROM consumption_log WHERE status='pending_confirm'"
            ).fetchone()[0]
            coordinator_stats['today_lessons'] = db.execute(
                "SELECT COUNT(*) FROM schedules WHERE status='confirmed' AND date(start_time)=?",
                (datetime.now().strftime('%Y-%m-%d'),)
            ).fetchone()[0]

        # Trend: weekly lead creation (past 8 weeks)
        weekly_trend = []
        for w in range(7, -1, -1):
            ws = (datetime.now() - timedelta(days=w*7)).strftime('%Y-%m-%d')
            we = (datetime.now() - timedelta(days=w*7-6)).strftime('%Y-%m-%d')
            cnt = db.execute(
                "SELECT COUNT(*) FROM leads WHERE merge_target_id IS NULL AND date(created_at) BETWEEN ? AND ?",
                (ws, we)
            ).fetchone()[0]
            weekly_trend.append({"week": ws, "count": cnt})

        # Conversion funnel
        funnel = {}
        for s in ('pending', 'assigned', 'following', 'trial', 'enrolled'):
            funnel[s] = db.execute(
                "SELECT COUNT(*) FROM leads WHERE status=? AND merge_target_id IS NULL", (s,)
            ).fetchone()[0]

        # Team performance (admin/supervisor)
        team_perf = []
        if role in ('admin', 'supervisor'):
            team = db.execute(
                "SELECT u.id, u.name, u.role, "
                "(SELECT COUNT(*) FROM leads l WHERE l.assignee_id=u.id AND l.merge_target_id IS NULL) as total_leads, "
                "(SELECT COUNT(*) FROM leads l WHERE l.assignee_id=u.id AND l.status='enrolled' AND l.merge_target_id IS NULL) as enrolled, "
                "(SELECT COUNT(*) FROM activities a WHERE a.user_id=u.id AND a.created_at >= ?) as recent_activities "
                "FROM users u WHERE u.role IN ('cs','consultant','academic','coordinator') AND u.is_active=1 "
                "ORDER BY total_leads DESC",
                (today,)
            ).fetchall()
            team_perf = [dict(t) for t in team]

        # Followup stats for overdue banner
        overdue_count = db.execute(
            "SELECT COUNT(DISTINCT l.id) FROM leads l "
            "WHERE l.id IN (SELECT lead_id FROM activities WHERE next_action_date IS NOT NULL AND next_action_date != '' AND next_action_date < ?) "
            "AND l.status NOT IN ('enrolled','closed','lost') AND l.merge_target_id IS NULL",
            (today,)
        ).fetchone()[0]

        db.close()
        return json_resp(self, {
            "stats": {
                "total": total, "pending": pending, "following": following, "enrolled": enrolled,
                "pending_unassigned": pending_count,
                "pool_count": pool_count,
                **coordinator_stats,
            },
            "followup_stats": {"overdue": overdue_count, "today": len(today_fups)},
            "today_followups": [dict(r) for r in today_fups],
            "recent_leads": [dict(r) for r in recent],
            "status_breakdown": [dict(r) for r in statuses],
            "weekly_trend": weekly_trend,
            "funnel": funnel,
            "team_performance": team_perf,
        })

    # ═══════════════════ PUBLIC POOL ═══════════════════════════════════════

    def _handle_pool_list(self, user, params):
        """List leads currently in the public pool or silent pool."""
        db = get_db()
        pool_status = params.get('pool_status', ['pool'])[0]
        status_filter = params.get('status', ['all'])[0]

        conditions = ["l.pool_status=?"]
        args = [pool_status]
        if status_filter != 'all':
            conditions.append("l.status=?")
            args.append(status_filter)

        search = params.get('search', [None])[0]
        if search:
            conditions.append("(l.name LIKE ? OR l.phone LIKE ?)")
            s = f"%{search}%"
            args.extend([s, s])

        where = "WHERE " + " AND ".join(conditions) + " AND l.merge_target_id IS NULL"
        page, size, offset = get_page_params(params)

        rows = db.execute(
            f"SELECT l.*, (SELECT COUNT(*) FROM pool_return_log pr WHERE pr.lead_id=l.id) as return_count "
            f"FROM leads l {where} ORDER BY l.updated_at DESC LIMIT ? OFFSET ?",
            args + [size, offset]
        ).fetchall()
        total = db.execute(f"SELECT COUNT(*) FROM leads l {where}", args).fetchone()[0]
        db.close()
        return json_resp(self, {"leads": [dict(r) for r in rows], "total": total})

    def _handle_pool_return(self, user, lead_id, data):
        """Manually return a lead to the public pool."""
        db = get_db()
        lead = db.execute("SELECT * FROM leads WHERE id=?", (lead_id,)).fetchone()
        if not lead:
            db.close()
            return json_resp(self, {"error": "Not Found"}, 404)
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        reason = data.get('reason', '手动回池')
        return_count = (lead['pool_return_count'] or 0) + 1
        # Check max returns
        cfg = _get_pool_config()
        max_returns = int(cfg.get('max_returns', 3))
        new_status = 'lost' if return_count >= max_returns else 'pending'
        pool_status = 'silent' if return_count >= max_returns else 'pool'

        db.execute(
            "UPDATE leads SET status=?, pool_status=?, pool_return_count=?, assignee_id=NULL, updated_at=? WHERE id=?",
            (new_status, pool_status, return_count, now, lead_id)
        )
        db.execute(
            "INSERT INTO pool_return_log (id, lead_id, reason, returned_by, created_at) VALUES (?,?,?,?,?)",
            (gen_id('pr_'), lead_id, reason, user['id'], now)
        )
        db.commit()
        db.close()
        _audit(user['id'], 'pool_return', 'lead', lead_id, f'手动回池: {reason}')
        return json_resp(self, {"status": "returned", "pool_status": pool_status})

    def _handle_pool_settings(self, user):
        """GET current pool configuration."""
        cfg = _get_pool_config()
        # Parse source_timeouts JSON
        try:
            cfg['source_timeouts'] = json.loads(cfg.get('source_timeouts', '{}'))
        except:
            cfg['source_timeouts'] = {}
        return json_resp(self, {"settings": cfg})

    def _handle_pool_update_settings(self, user, data):
        """PUT update pool configuration."""
        allowed = ('timeout_days', 'max_returns', 'silent_pool_timeout', 'source_timeouts',
                    'assignment_strategy', 'daily_claim_limit', 'active_lead_limit')
        db = get_db()
        for key in allowed:
            if key in data:
                val = data[key]
                if isinstance(val, (dict, list)):
                    val = json.dumps(val, ensure_ascii=False)
                else:
                    val = str(val)
                db.execute("INSERT OR REPLACE INTO pool_config (key, value) VALUES (?,?)", (key, val))
        db.commit()
        db.close()
        _audit(user['id'], 'pool_settings', 'pool', '', '更新公海池配置')
        return json_resp(self, {"status": "updated"})

    def _handle_pool_stats(self, user):
        """Statistics about pool status."""
        db = get_db()
        stats = {
            'in_pool': db.execute("SELECT COUNT(*) FROM leads WHERE pool_status='pool' AND merge_target_id IS NULL").fetchone()[0],
            'silent_pool': db.execute("SELECT COUNT(*) FROM leads WHERE pool_status='silent' AND merge_target_id IS NULL").fetchone()[0],
            'active': db.execute("SELECT COUNT(*) FROM leads WHERE pool_status='active' AND merge_target_id IS NULL").fetchone()[0],
            'total_returns': db.execute("SELECT COUNT(*) FROM pool_return_log").fetchone()[0],
        }
        # Recent return activity
        recent = db.execute(
            "SELECT pr.*, l.name as lead_name FROM pool_return_log pr "
            "JOIN leads l ON pr.lead_id = l.id ORDER BY pr.created_at DESC LIMIT 20"
        ).fetchall()
        db.close()
        return json_resp(self, {"stats": stats, "recent_returns": [dict(r) for r in recent]})

    def _handle_pool_log(self, user, params):
        """Full pool return log with pagination."""
        db = get_db()
        lead_id = params.get('lead_id', [None])[0]
        page, size, offset = get_page_params(params)
        if lead_id:
            rows = db.execute(
                "SELECT pr.*, l.name as lead_name FROM pool_return_log pr "
                "JOIN leads l ON pr.lead_id = l.id WHERE pr.lead_id=? "
                "ORDER BY pr.created_at DESC LIMIT ? OFFSET ?",
                (lead_id, size, offset)
            ).fetchall()
            total = db.execute("SELECT COUNT(*) FROM pool_return_log WHERE lead_id=?", (lead_id,)).fetchone()[0]
        else:
            rows = db.execute(
                "SELECT pr.*, l.name as lead_name FROM pool_return_log pr "
                "JOIN leads l ON pr.lead_id = l.id ORDER BY pr.created_at DESC LIMIT ? OFFSET ?",
                (size, offset)
            ).fetchall()
            total = db.execute("SELECT COUNT(*) FROM pool_return_log").fetchone()[0]
        db.close()
        return json_resp(self, {"logs": [dict(r) for r in rows], "total": total})

    # ═══════════════════ CONTRACTS & PAYMENTS ══════════════════════════════

    def _handle_list_contracts(self, user, params):
        db = get_db()
        conditions, args = [], []
        lead_id = params.get('lead_id', [None])[0]
        if lead_id:
            conditions.append("c.lead_id=?")
            args.append(lead_id)
        status = params.get('status', [None])[0]
        if status and status != 'all':
            conditions.append("c.status=?")
            args.append(status)
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        page, size, offset = get_page_params(params)
        rows = db.execute(
            f"SELECT c.*, l.name as lead_name, u.name as created_by_name "
            f"FROM contracts c JOIN leads l ON c.lead_id = l.id "
            f"LEFT JOIN users u ON c.created_by = u.id {where} "
            f"ORDER BY c.created_at DESC LIMIT ? OFFSET ?", args + [size, offset]
        ).fetchall()
        total = db.execute(f"SELECT COUNT(*) FROM contracts c {where}", args).fetchone()[0]
        db.close()
        return json_resp(self, {"contracts": [dict(r) for r in rows], "total": total})

    def _handle_get_contract(self, user, contract_id):
        db = get_db()
        c = db.execute(
            "SELECT c.*, l.name as lead_name, u.name as created_by_name "
            "FROM contracts c JOIN leads l ON c.lead_id = l.id "
            "LEFT JOIN users u ON c.created_by = u.id WHERE c.id=?", (contract_id,)
        ).fetchone()
        if not c:
            db.close()
            return json_resp(self, {"error": "Not Found"}, 404)
        payments = db.execute(
            "SELECT p.*, u.name as created_by_name FROM payments p "
            "LEFT JOIN users u ON p.created_by = u.id WHERE p.contract_id=? ORDER BY p.created_at", (contract_id,)
        ).fetchall()
        db.close()
        return json_resp(self, {"contract": dict(c), "payments": [dict(p) for p in payments]})

    def _handle_create_contract(self, user, data):
        lead_id = data.get('lead_id', '').strip()
        if not lead_id:
            return json_resp(self, {"error": "请选择学生"}, 400)
        contract_id = gen_id('ct_')
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        # Generate contract number
        seq = uuid.uuid4().hex[:6].upper()
        contract_no = f"CT-{datetime.now().strftime('%Y%m')}-{seq}"
        db = get_db()

        # Get lead info for notification
        lead = db.execute("SELECT * FROM leads WHERE id=?", (lead_id,)).fetchone()
        if not lead:
            db.close()
            return json_resp(self, {"error": "学生不存在"}, 404)

        total_amount = float(data.get('total_amount', 0))
        paid_amount = float(data.get('paid_amount', 0))
        db.execute(
            "INSERT INTO contracts (id,lead_id,contract_no,total_amount,paid_amount,status,signed_at,valid_from,valid_until,notes,created_by,created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (contract_id, lead_id, contract_no, total_amount,
             paid_amount, data.get('status', 'active'),
             data.get('signed_at', now), data.get('valid_from', ''),
             data.get('valid_until', ''), data.get('notes', ''), user['id'], now)
        )

        # Auto-transition lead to enrolled status
        old_status = lead['status']
        if old_status != 'enrolled':
            db.execute("UPDATE leads SET status='enrolled', updated_at=? WHERE id=?",
                       (now, lead_id))
            # Log status change in activity
            db.execute(
                "INSERT INTO activities (id,lead_id,user_id,type,content,created_at) VALUES (?,?,?,?,?,?)",
                (gen_id('a_'), lead_id, user['id'], 'note',
                 f"系统：创建合同后线索状态由 {old_status} 变更为 enrolled", now)
            )

        # Auto-create course_package if requested
        if data.get('auto_create_package') and float(data.get('estimated_hours', 0)) > 0:
            est_hours = float(data['estimated_hours'])
            pkg_id = gen_id('p_')
            unit_price = round(total_amount / est_hours, 2) if est_hours > 0 else 0
            pkg_name = data.get('package_name', f'合同{contract_no} 课时包')
            valid_from = data.get('valid_from', now[:10])
            valid_until = data.get('valid_until', '')
            # Auto-calculate valid_until from period_months
            period_months = int(data.get('period_months', 0))
            if period_months > 0 and not valid_until:
                try:
                    from_dt = datetime.strptime(valid_from, '%Y-%m-%d') if valid_from else datetime.now()
                    # Add months manually (no external deps)
                    m = from_dt.month - 1 + period_months
                    y = from_dt.year + m // 12
                    m = m % 12 + 1
                    import calendar
                    d = min(from_dt.day, calendar.monthrange(y, m)[1])
                    valid_until = f'{y:04d}-{m:02d}-{d:02d}'
                except:
                    valid_until = ''
            db.execute(
                "INSERT INTO course_packages (id,lead_id,package_name,total_hours,used_hours,price,unit_price,valid_from,valid_until,status,contract_id,created_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (pkg_id, lead_id, pkg_name, est_hours, 0,
                 total_amount, unit_price, valid_from, valid_until,
                 'active', contract_id, now)
            )

        db.commit()

        # Notify the lead's assignee consultant
        if lead['assignee_id']:
            self._create_notification_simple(db, contract_id, 'contract',
                f'学生 {lead["name"]} 已签约',
                f'合同 {contract_no} 已创建，线索状态已更新为已签约',
                'contract', lead['assignee_id'])

        db.close()
        _audit(user['id'], 'create_contract', 'contract', contract_id, f'创建合同: {contract_no}')
        return json_resp(self, {"id": contract_id, "contract_no": contract_no, "status": "created"}, 201)

    def _handle_contract_pay(self, user, contract_id, data):
        """Record a payment against a contract."""
        amount = float(data.get('amount', 0))
        if amount <= 0:
            return json_resp(self, {"error": "金额必须大于0"}, 400)
        pay_id = gen_id('pay_')
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db = get_db()
        contract = db.execute("SELECT * FROM contracts WHERE id=?", (contract_id,)).fetchone()
        if not contract:
            db.close()
            return json_resp(self, {"error": "合同不存在"}, 404)
        db.execute(
            "INSERT INTO payments (id,contract_id,amount,pay_method,pay_at,notes,created_by,created_at) VALUES (?,?,?,?,?,?,?,?)",
            (pay_id, contract_id, amount, data.get('pay_method', ''), data.get('pay_at', now),
             data.get('notes', ''), user['id'], now)
        )
        new_paid = contract['paid_amount'] + amount
        new_status = 'completed' if new_paid >= contract['total_amount'] else 'active'
        db.execute("UPDATE contracts SET paid_amount=?, status=? WHERE id=?",
                   (new_paid, new_status, contract_id))
        db.commit()
        db.close()
        _audit(user['id'], 'pay', 'contract', contract_id, f'收款 ¥{amount}')
        return json_resp(self, {"id": pay_id, "status": "paid", "paid_amount": new_paid})

    def _handle_list_payments(self, user, contract_id, params):
        """List payments for a contract."""
        db = get_db()
        rows = db.execute(
            "SELECT p.*, u.name as created_by_name FROM payments p "
            "LEFT JOIN users u ON p.created_by = u.id WHERE p.contract_id=? ORDER BY p.created_at", (contract_id,)
        ).fetchall()
        db.close()
        return json_resp(self, {"payments": [dict(r) for r in rows]})

    # ═══════════════════ REFERRALS ═════════════════════════════════════════

    def _handle_list_referrals(self, user, params):
        db = get_db()
        conditions, args = [], []
        lead_id = params.get('lead_id', [None])[0]
        if lead_id:
            conditions.append("(r.referrer_lead_id=? OR r.referred_lead_id=?)")
            args.extend([lead_id, lead_id])
        status = params.get('status', [None])[0]
        if status and status != 'all':
            conditions.append("r.status=?")
            args.append(status)
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        page, size, offset = get_page_params(params)
        rows = db.execute(
            f"SELECT r.*, ref.name as referrer_name, red.name as referred_name "
            f"FROM referrals r "
            f"JOIN leads ref ON r.referrer_lead_id = ref.id "
            f"LEFT JOIN leads red ON r.referred_lead_id = red.id "
            f"{where} ORDER BY r.created_at DESC LIMIT ? OFFSET ?", args + [size, offset]
        ).fetchall()
        total = db.execute(f"SELECT COUNT(*) FROM referrals r {where}", args).fetchone()[0]
        db.close()
        return json_resp(self, {"referrals": [dict(r) for r in rows], "total": total})

    def _handle_get_referral(self, user, referral_id):
        db = get_db()
        r = db.execute(
            "SELECT r.*, ref.name as referrer_name, ref.phone as referrer_phone, "
            "red.name as referred_name, red.phone as referred_phone "
            "FROM referrals r "
            "JOIN leads ref ON r.referrer_lead_id = ref.id "
            "LEFT JOIN leads red ON r.referred_lead_id = red.id WHERE r.id=?", (referral_id,)
        ).fetchone()
        if not r:
            db.close()
            return json_resp(self, {"error": "Not Found"}, 404)
        db.close()
        return json_resp(self, {"referral": dict(r)})

    def _handle_create_referral(self, user, data):
        """Create a referral link between two leads."""
        referrer_id = data.get('referrer_lead_id', '').strip()
        referred_id = data.get('referred_lead_id', '').strip()
        if not referrer_id or not referred_id:
            return json_resp(self, {"error": "请选择推荐人和被推荐人"}, 400)
        if referrer_id == referred_id:
            return json_resp(self, {"error": "推荐人和被推荐人不能相同"}, 400)
        ref_id = gen_id('ref_')
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db = get_db()
        # Check if referral already exists
        existing = db.execute(
            "SELECT id FROM referrals WHERE referrer_lead_id=? AND referred_lead_id=?",
            (referrer_id, referred_id)
        ).fetchone()
        if existing:
            db.close()
            return json_resp(self, {"warning": "该转介绍关系已存在", "id": existing['id']}, 409)
        db.execute(
            "INSERT INTO referrals (id,referrer_lead_id,referred_lead_id,status,reward_type,reward_amount,created_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (ref_id, referrer_id, referred_id, 'pending', data.get('reward_type', ''), float(data.get('reward_amount', 0)), now)
        )
        # Also update the referred lead's source to '转介绍' if not set
        db.execute("UPDATE leads SET source='转介绍', referral_source_id=?, updated_at=? WHERE id=? AND (source IS NULL OR source='')",
                   (referrer_id, now, referred_id))
        db.commit()
        db.close()
        _audit(user['id'], 'create_referral', 'referral', ref_id, '创建转介绍关系')
        return json_resp(self, {"id": ref_id, "status": "created"}, 201)

    # ═══════════════════ LEAD MERGE ════════════════════════════════════════

    def _handle_merge_leads(self, user, target_lead_id, data):
        """Merge a duplicate lead into the target lead.
        Source lead is marked as merged, activities/packages are re-assigned."""
        source_lead_id = data.get('source_lead_id', '').strip()
        if not source_lead_id:
            return json_resp(self, {"error": "请选择要合并的源线索"}, 400)
        if source_lead_id == target_lead_id:
            return json_resp(self, {"error": "不能合并自己"}, 400)

        db = get_db()
        target = db.execute("SELECT * FROM leads WHERE id=?", (target_lead_id,)).fetchone()
        source = db.execute("SELECT * FROM leads WHERE id=?", (source_lead_id,)).fetchone()
        if not target or not source:
            db.close()
            return json_resp(self, {"error": "线索不存在"}, 404)

        now = datetime.now().strftime('%Y-%m-%d %H:%M')

        # Merge fields: fill empty target fields from source
        for field in ('phone', 'wechat', 'country', 'grade', 'subject', 'notes'):
            if not target[field] and source[field]:
                db.execute(f"UPDATE leads SET {field}=? WHERE id=?", (source[field], target_lead_id))

        # Reassign activities from source to target
        db.execute("UPDATE activities SET lead_id=? WHERE lead_id=?", (target_lead_id, source_lead_id))

        # Reassign packages
        db.execute("UPDATE course_packages SET lead_id=? WHERE lead_id=?", (target_lead_id, source_lead_id))

        # Reassign schedules
        db.execute("UPDATE schedules SET lead_id=? WHERE lead_id=?", (target_lead_id, source_lead_id))

        # Reassign AI reports
        db.execute("UPDATE ai_reports SET lead_id=? WHERE lead_id=?", (target_lead_id, source_lead_id))

        # Reassign referrals
        db.execute("UPDATE referrals SET referrer_lead_id=? WHERE referrer_lead_id=?", (target_lead_id, source_lead_id))
        db.execute("UPDATE referrals SET referred_lead_id=? WHERE referred_lead_id=?", (target_lead_id, source_lead_id))

        # Mark source as merged
        db.execute("UPDATE leads SET merge_target_id=?, merged_at=?, status='closed', pool_status='silent', updated_at=? WHERE id=?",
                   (target_lead_id, now, now, source_lead_id))

        # Add activity note about the merge
        db.execute(
            "INSERT INTO activities (id,lead_id,user_id,type,content,created_at) VALUES (?,?,?,?,?,?)",
            (gen_id('a_'), target_lead_id, user['id'], 'note',
             f"系统：已合并重复线索 [{source['name']}] 至本线索", now)
        )

        db.commit()
        db.close()
        _audit(user['id'], 'merge', 'lead', target_lead_id, f'合并线索 {source_lead_id} 至 {target_lead_id}')
        return json_resp(self, {"status": "merged", "target_id": target_lead_id})

    # ═══════════════════ USERS ═════════════════════════════════════════════

    def _handle_list_users(self, user):
        db = get_db()
        rows = db.execute(
            "SELECT id, name, role, phone, is_active, subjects, timezone, rate, classin_account, is_tutor FROM users WHERE is_active=1 ORDER BY role, name"
        ).fetchall()
        db.close()
        return json_resp(self, {"users": [dict(r) for r in rows]})

    # ═══════════════════ LEADS (CRM) ════════════════════════════════════════

    def _handle_list_leads(self, user, params):
        db = get_db()
        role, uid = user['role'], user['id']
        conditions, args = [], []

        if role not in ('admin', 'supervisor'):
            conditions.append("l.assignee_id=?")
            args.append(uid)

        for key, col in [('status', 'l.status'), ('source', 'l.source'), ('pool_status', 'l.pool_status'), ('service_type', 'l.service_type')]:
            v = params.get(key, [None])[0]
            if v and v != 'all':
                conditions.append(f"{col}=?")
                args.append(v)

        assignee_filter = params.get('assignee', [None])[0]
        if assignee_filter and assignee_filter != 'all':
            conditions.append("l.assignee_id=?")
            args.append(assignee_filter)

        academic_filter = params.get('academic_manager', [None])[0]
        if academic_filter and academic_filter != 'all':
            conditions.append("l.academic_manager_id=?")
            args.append(academic_filter)

        search = params.get('search', [None])[0]
        if search:
            conditions.append("(l.name LIKE ? OR l.phone LIKE ? OR l.wechat LIKE ?)")
            s = f"%{search}%"
            args.extend([s, s, s])

        # Exclude merged leads by default
        if params.get('include_merged', [None])[0] != '1':
            conditions.append("l.merge_target_id IS NULL")

        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        page, size, offset = get_page_params(params)

        rows = db.execute(
            f"SELECT l.*, u.name as assignee_name, cr.name as creator_name, ac.name as academic_name "
            f"FROM leads l LEFT JOIN users u ON l.assignee_id = u.id "
            f"LEFT JOIN users cr ON l.created_by_id = cr.id "
            f"LEFT JOIN users ac ON l.academic_manager_id = ac.id {where} "
            f"ORDER BY l.created_at DESC LIMIT ? OFFSET ?", args + [size, offset]
        ).fetchall()

        count = db.execute(f"SELECT COUNT(*) FROM leads l {where}", args).fetchone()[0]
        db.close()
        return json_resp(self, {"leads": [dict(r) for r in rows], "total": count, "page": page, "page_size": size})

    def _handle_get_lead(self, user, lead_id):
        db = get_db()
        lead = db.execute(
            "SELECT l.*, u.name as assignee_name, cr.name as creator_name, ac.name as academic_name "
            "FROM leads l LEFT JOIN users u ON l.assignee_id = u.id "
            "LEFT JOIN users cr ON l.created_by_id = cr.id "
            "LEFT JOIN users ac ON l.academic_manager_id = ac.id WHERE l.id=?", (lead_id,)
        ).fetchone()
        if not lead:
            db.close()
            return json_resp(self, {"error": "Not Found"}, 404)

        acts = db.execute(
            "SELECT a.*, u.name as user_name FROM activities a "
            "LEFT JOIN users u ON a.user_id = u.id WHERE a.lead_id=? "
            "ORDER BY a.created_at DESC LIMIT 100", (lead_id,)
        ).fetchall()

        # Packages for this lead
        pkgs = db.execute(
            "SELECT * FROM course_packages WHERE lead_id=? ORDER BY created_at DESC", (lead_id,)
        ).fetchall()

        # Contracts for this lead
        contracts = db.execute(
            "SELECT * FROM contracts WHERE lead_id=? ORDER BY created_at DESC", (lead_id,)
        ).fetchall()

        # Referral info
        referrals_as_referrer = db.execute(
            "SELECT r.*, l.name as referred_name FROM referrals r "
            "JOIN leads l ON r.referred_lead_id = l.id WHERE r.referrer_lead_id=?", (lead_id,)
        ).fetchall()
        referrals_as_referred = db.execute(
            "SELECT r.*, l.name as referrer_name FROM referrals r "
            "JOIN leads l ON r.referrer_lead_id = l.id WHERE r.referred_lead_id=?", (lead_id,)
        ).fetchall()

        db.close()
        return json_resp(self, {
            "lead": dict(lead),
            "activities": [dict(a) for a in acts],
            "packages": [dict(p) for p in pkgs],
            "contracts": [dict(c) for c in contracts],
            "referrals_as_referrer": [dict(r) for r in referrals_as_referrer],
            "referrals_as_referred": [dict(r) for r in referrals_as_referred],
        })

    def _handle_public_captcha(self):
        """Generate a simple math captcha for public lead form."""
        ip = self.client_address[0]
        challenge = _gen_captcha_for_ip(ip)
        return json_resp(self, {"challenge": challenge, "hint": "输入计算结果"})

    def _handle_public_lead(self, data):
        """Public lead submission endpoint - no auth required."""
        ip = self.client_address[0]

        # Rate limit check
        if not _check_public_limit(ip):
            return json_resp(self, {"error": "提交过于频繁，请稍后再试"}, 429)

        # Captcha verification
        captcha_answer = data.get('captcha', '').strip()
        if not captcha_answer:
            return json_resp(self, {"error": "请完成验证"}, 400)
        if not _verify_captcha(ip, captcha_answer):
            return json_resp(self, {"error": "验证码错误，请重试"}, 400)

        name = data.get('name', '').strip()
        if not name:
            return json_resp(self, {"error": "姓名不能为空"}, 400)

        lead_id = gen_id('l_')
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db = get_db()
        phone = data.get('phone', '').strip()
        wechat = data.get('wechat', '').strip()

        # Duplicate detection
        if phone:
            dup = db.execute("SELECT id, name, status FROM leads WHERE phone=? AND status NOT IN ('lost','closed')", (phone,)).fetchone()
            if dup:
                db.close()
                return json_resp(self, {"warning": f"该手机号已有线索: {dup['name']}", "dup_id": dup['id']}, 409)
        if wechat:
            dup = db.execute("SELECT id, name, status FROM leads WHERE wechat=? AND status NOT IN ('lost','closed')", (wechat,)).fetchone()
            if dup:
                db.close()
                return json_resp(self, {"warning": f"该微信号已有线索: {dup['name']}", "dup_id": dup['id']}, 409)

        # Name fuzzy dedup (Levenshtein distance <= 2)
        all_leads = db.execute(
            "SELECT id, name, status FROM leads WHERE status NOT IN ('lost','closed')"
        ).fetchall()
        close_matches = []
        for ld in all_leads:
            if _levenshtein(name.lower(), ld['name'].lower()) <= 2:
                close_matches.append(ld)
        if close_matches:
            match_names = ', '.join([m['name'] for m in close_matches[:3]])
            db.close()
            return json_resp(self, {"warning": f"存在相似姓名线索: {match_names}"}, 409)

        # Insert with UTM fields
        utm_source = data.get('utm_source', '').strip()
        utm_campaign = data.get('utm_campaign', '').strip()
        utm_medium = data.get('utm_medium', '').strip()
        campaign = data.get('campaign', '').strip()
        ad_source = data.get('ad_source', '').strip()
        landing_page = data.get('landing_page', '').strip()
        notes = data.get('notes', '').strip()

        db.execute(
            "INSERT INTO leads (id,name,phone,wechat,country,grade,subject,source,status,"
            "notes,created_at,updated_at,"
            "utm_source,utm_campaign,utm_medium,campaign,ad_source,landing_page) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (lead_id, name, phone, wechat,
             data.get('country', '').strip(), data.get('grade', '').strip(),
             data.get('subject', '').strip(), '线上', 'pending',
             notes, now, now,
             utm_source, utm_campaign, utm_medium, campaign, ad_source, landing_page)
        )
        # Auto-create activity for source tracking
        source_note = f"线上咨询提交"
        if utm_source:
            source_note += f" (来源: {utm_source}"
            if utm_campaign:
                source_note += f" / {utm_campaign}"
            source_note += ")"
        db.execute(
            "INSERT INTO activities (id,lead_id,user_id,type,content,created_at) VALUES (?,?,?,?,?,?)",
            (gen_id('act_'), lead_id, 'u_admin', 'note', source_note, now)
        )
        # Auto-assign if strategy is configured
        assignee_id = self._auto_assign_lead(db, lead_id, data)
        if assignee_id:
            db.execute(
                "UPDATE leads SET assignee_id=?, status='assigned', assigned_at=? WHERE id=?",
                (assignee_id, now, lead_id)
            )
            assignee_name = db.execute("SELECT name FROM users WHERE id=?", (assignee_id,)).fetchone()
            aname = assignee_name['name'] if assignee_name else assignee_id
            db.execute(
                "INSERT INTO activities (id,lead_id,user_id,type,content,created_at) VALUES (?,?,?,?,?,?)",
                (gen_id('act_'), lead_id, 'u_admin', 'note', f'系统自动分配给 {aname}', now)
            )
        db.commit()
        db.close()
        _audit('public', 'create', 'lead', lead_id, f'线上提交: {name}')
        return json_resp(self, {"id": lead_id, "status": "created", "auto_assigned": bool(assignee_id)}, 201)

    def _auto_assign_lead(self, db, lead_id, lead_data):
        """Auto-assign a lead based on strategy and rules. Returns assignee_id or None."""
        cfg = _get_pool_config()
        strategy = cfg.get('assignment_strategy', 'manual')
        if strategy == 'manual':
            return None

        # 1. Check assignment rules (source, country, subject → specific assignee)
        rules = db.execute(
            "SELECT * FROM assignment_rules WHERE is_active=1 ORDER BY priority ASC"
        ).fetchall()
        for rule in rules:
            field_val = lead_data.get(rule['condition_field'], '')
            if field_val and field_val.lower() == rule['condition_value'].lower():
                assignee = db.execute(
                    "SELECT id FROM users WHERE id=? AND is_active=1 AND role IN ('cs','consultant','academic')",
                    (rule['assignee_id'],)
                ).fetchone()
                if assignee:
                    return assignee['id']

        # 2. Strategy-based assignment
        sales_roles = "'cs','consultant','academic'"
        if strategy == 'round_robin':
            # Get the assignee who got the last assignment
            last = db.execute(
                f"SELECT assignee_id FROM leads WHERE assignee_id IS NOT NULL "
                f"AND assignee_id IN (SELECT id FROM users WHERE role IN ({sales_roles}) AND is_active=1) "
                "ORDER BY assigned_at DESC LIMIT 1"
            ).fetchone()
            sales_users = db.execute(
                "SELECT id FROM users WHERE role IN ('cs','consultant','academic') AND is_active=1 ORDER BY id"
            ).fetchall()
            if not sales_users:
                return None
            if not last:
                return sales_users[0]['id']
            # Find next in round
            ids = [u['id'] for u in sales_users]
            try:
                last_idx = ids.index(last['assignee_id'])
                next_idx = (last_idx + 1) % len(ids)
                return ids[next_idx]
            except ValueError:
                return ids[0]

        elif strategy == 'least_load':
            # Assign to the user with fewest active (non-closed/lost) leads
            sales_users = db.execute(
                "SELECT u.id, (SELECT COUNT(*) FROM leads WHERE assignee_id=u.id "
                "AND status NOT IN ('closed','lost') AND merge_target_id IS NULL) as cnt "
                "FROM users u WHERE u.role IN ('cs','consultant','academic') AND u.is_active=1 "
                "ORDER BY cnt ASC LIMIT 1"
            ).fetchone()
            return sales_users['id'] if sales_users else None

        return None

    def _handle_create_lead(self, user, data):
        name = data.get('name', '').strip()
        if not name:
            return json_resp(self, {"error": "姓名不能为空"}, 400)
        lead_id = gen_id('l_')
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db = get_db()
        phone = data.get('phone', '').strip()
        wechat = data.get('wechat', '').strip()
        if phone:
            dup = db.execute("SELECT id, name, status FROM leads WHERE phone=? AND status NOT IN ('lost','closed')", (phone,)).fetchone()
            if dup:
                db.close()
                return json_resp(self, {"warning": f"该手机号已有线索: {dup['name']}({dup['status']})", "dup_id": dup['id']}, 409)
        if wechat:
            dup = db.execute("SELECT id, name, status FROM leads WHERE wechat=? AND status NOT IN ('lost','closed')", (wechat,)).fetchone()
            if dup:
                db.close()
                return json_resp(self, {"warning": f"该微信号已有线索: {dup['name']}({dup['status']})", "dup_id": dup['id']}, 409)

        db.execute(
            "INSERT INTO leads (id,name,phone,wechat,country,grade,subject,source,status,created_by_id,notes,created_at,updated_at,"
            "utm_source,utm_campaign,utm_medium,campaign,ad_source,landing_page,"
            "rating,lead_type,service_type,account_name,tags) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (lead_id, name, phone, wechat, data.get('country', '').strip(), data.get('grade', '').strip(),
             data.get('subject', '').strip(), data.get('source', '').strip(), 'pending', user['id'],
             data.get('notes', '').strip(), now, now,
             data.get('utm_source', '').strip(), data.get('utm_campaign', '').strip(),
             data.get('utm_medium', '').strip(), data.get('campaign', '').strip(),
             data.get('ad_source', '').strip(), data.get('landing_page', '').strip(),
             data.get('rating', '').strip(), data.get('lead_type', '').strip(),
             data.get('service_type', '').strip(),
             data.get('account_name', '').strip(),
             data.get('tags', '[]').strip())
        )
        # Auto-assign if strategy is configured
        assignee_id = self._auto_assign_lead(db, lead_id, data)
        if assignee_id:
            db.execute(
                "UPDATE leads SET assignee_id=?, status='assigned', assigned_at=? WHERE id=?",
                (assignee_id, now, lead_id)
            )
            assignee_name = db.execute("SELECT name FROM users WHERE id=?", (assignee_id,)).fetchone()
            aname = assignee_name['name'] if assignee_name else assignee_id
            db.execute(
                "INSERT INTO activities (id,lead_id,user_id,type,content,created_at) VALUES (?,?,?,?,?,?)",
                (gen_id('act_'), lead_id, user['id'], 'note', f'系统自动分配给 {aname}', now)
            )
        db.commit()
        db.close()
        _audit(user['id'], 'create', 'lead', lead_id, f'创建线索: {name}')
        return json_resp(self, {"id": lead_id, "status": "created", "auto_assigned": bool(assignee_id)}, 201)

    def _handle_update_lead(self, user, lead_id, data):
        db = get_db()
        lead = db.execute("SELECT * FROM leads WHERE id=?", (lead_id,)).fetchone()
        if not lead:
            db.close()
            return json_resp(self, {"error": "Not Found"}, 404)
        role, uid = user['role'], user['id']
        if role not in ('admin', 'supervisor', 'consultant') and lead['assignee_id'] != uid:
            db.close()
            return json_resp(self, {"error": "无权限修改此线索"}, 403)
        updates = {}
        for field in ('name', 'phone', 'wechat', 'country', 'grade', 'subject', 'source', 'notes', 'status',
                       'classin_account', 'referral_source_id', 'rating', 'lead_type', 'service_type',
                       'academic_manager_id', 'account_name', 'tags'):
            if field in data:
                updates[field] = data[field].strip() if isinstance(data[field], str) else data[field]
        if not updates:
            db.close()
            return json_resp(self, {"error": "No fields to update"}, 400)
        updates['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        set_clause = ", ".join(f"{k}=?" for k in updates)
        vals = list(updates.values()) + [lead_id]
        db.execute(f"UPDATE leads SET {set_clause} WHERE id=?", vals)
        db.commit()
        db.close()
        _audit(user['id'], 'update', 'lead', lead_id, f'更新字段: {",".join(updates.keys())}')
        return json_resp(self, {"status": "updated"})

    def _handle_assign_lead(self, user, lead_id, data):
        assignee_id = data.get('assignee_id', '').strip()
        if not assignee_id:
            return json_resp(self, {"error": "请选择跟进人"}, 400)
        db = get_db()
        assignee = db.execute("SELECT id, name, role FROM users WHERE id=? AND is_active=1", (assignee_id,)).fetchone()
        if not assignee:
            db.close()
            return json_resp(self, {"error": "跟进人不存在"}, 404)
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        new_status = 'assigned' if data.get('keep_status') != 'yes' else None
        if new_status:
            db.execute("UPDATE leads SET assignee_id=?, status=?, assigned_at=?, updated_at=? WHERE id=?",
                       (assignee_id, new_status, now, now, lead_id))
        else:
            db.execute("UPDATE leads SET assignee_id=?, assigned_at=?, updated_at=? WHERE id=?",
                       (assignee_id, now, now, lead_id))
        db.execute(
            "INSERT INTO activities (id,lead_id,user_id,type,content,created_at) VALUES (?,?,?,?,?,?)",
            (gen_id('a_'), lead_id, user['id'], 'note', f"系统：线索已分配给 {assignee['name']}({assignee['role']})", now)
        )
        db.commit()
        # Notify assignee
        lead_name = db.execute("SELECT name FROM leads WHERE id=?", (lead_id,)).fetchone()
        if lead_name:
            self._create_notification_simple(db, lead_id, 'lead_assignment',
                '新线索分配', f'线索 {lead_name["name"]} 已分配给您跟进',
                'lead', assignee_id)
        db.close()
        _audit(user['id'], 'assign', 'lead', lead_id, f'分配给 {assignee["name"]}')
        return json_resp(self, {"status": "assigned", "assignee_name": assignee['name']})

    def _handle_reclaim_lead(self, user, lead_id):
        # Accept data from POST body
        data = read_body(self)
        db = get_db()
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        assignee_id = data.get('assignee_id', '').strip() if data else ''
        if assignee_id:
            # Check assignee exists
            assignee = db.execute("SELECT id, name FROM users WHERE id=? AND is_active=1", (assignee_id,)).fetchone()
            if assignee:
                db.execute(
                    "UPDATE leads SET pool_status='active', status='assigned', assignee_id=?, updated_at=? WHERE id=?",
                    (assignee_id, now, lead_id)
                )
                db.execute(
                    "INSERT INTO activities (id,lead_id,user_id,type,content,created_at) VALUES (?,?,?,?,?,?)",
                    (gen_id('a_'), lead_id, user['id'], 'note',
                     f"系统：从公海池捞回，重新分配给 {assignee['name']}", now)
                )
                _audit(user['id'], 'reclaim', 'lead', lead_id, f'从公海池捞回，重分配给 {assignee["name"]}')
            else:
                db.close()
                return json_resp(self, {"error": "跟进人不存在"}, 404)
        else:
            db.execute(
                "UPDATE leads SET pool_status='active', status='assigned', assignee_id=?, updated_at=? WHERE id=?",
                (user['id'], now, lead_id)
            )
            db.execute(
                "INSERT INTO activities (id,lead_id,user_id,type,content,created_at) VALUES (?,?,?,?,?,?)",
                (gen_id('a_'), lead_id, user['id'], 'note', f"系统：从公海池捞回，由 {user['name']} 跟进", now)
            )
            _audit(user['id'], 'reclaim', 'lead', lead_id, '从公海池捞回')
        # Notify the assignee of the reclaim
        notify_uid = assignee_id if assignee_id else user['id']
        lead_name_r = db.execute("SELECT name FROM leads WHERE id=?", (lead_id,)).fetchone()
        if lead_name_r:
            self._create_notification_simple(db, lead_id, 'lead_reclaim',
                '线索捞回', f'线索 {lead_name_r["name"]} 已从公海池捞回分配给您',
                'lead', notify_uid)
        db.commit()
        db.close()
        return json_resp(self, {"status": "reclaimed"})

    def _handle_assignment_dashboard(self, user):
        """Assignment dashboard for supervisors."""
        if user['role'] not in ('admin', 'supervisor'):
            return json_resp(self, {"error": "无权限"}, 403)
        db = get_db()
        cfg = _get_pool_config()
        pending = db.execute(
            "SELECT COUNT(*) FROM leads WHERE status='pending' AND pool_status='active' AND merge_target_id IS NULL"
        ).fetchone()[0]
        in_pool = db.execute(
            "SELECT COUNT(*) FROM leads WHERE pool_status='pool' AND merge_target_id IS NULL"
        ).fetchone()[0]
        # Sales team load
        sales = db.execute(
            "SELECT u.id, u.name, u.role, "
            "(SELECT COUNT(*) FROM leads WHERE assignee_id=u.id AND status NOT IN ('closed','lost') AND merge_target_id IS NULL) as active_leads, "
            "(SELECT COUNT(*) FROM leads WHERE assignee_id=u.id AND merge_target_id IS NULL) as total_leads, "
            "(SELECT COUNT(*) FROM activities WHERE user_id=u.id AND created_at >= date('now','start of day')) as today_activities "
            "FROM users u WHERE u.role IN ('cs','consultant','academic') AND u.is_active=1 ORDER BY u.name"
        ).fetchall()
        # Today's assignments
        today = datetime.now().strftime('%Y-%m-%d')
        today_assign = db.execute(
            "SELECT COUNT(*) FROM leads WHERE assigned_at LIKE ?", (today + '%',)
        ).fetchone()[0]
        db.close()
        return json_resp(self, {
            "pending": pending,
            "in_pool": in_pool,
            "strategy": cfg.get('assignment_strategy', 'manual'),
            "daily_claim_limit": cfg.get('daily_claim_limit', '10'),
            "active_lead_limit": cfg.get('active_lead_limit', '30'),
            "team": [dict(r) for r in sales],
            "today_assignments": today_assign,
        })

    def _handle_assignment_rules(self, user):
        """GET all assignment rules."""
        db = get_db()
        rules = db.execute(
            "SELECT r.*, u.name as assignee_name FROM assignment_rules r "
            "LEFT JOIN users u ON r.assignee_id=u.id ORDER BY r.priority ASC"
        ).fetchall()
        db.close()
        return json_resp(self, {"rules": [dict(r) for r in rules]})

    def _handle_create_assignment_rule(self, user, data):
        """POST create a new assignment rule."""
        if user['role'] not in ('admin', 'supervisor'):
            return json_resp(self, {"error": "无权限"}, 403)
        db = get_db()
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        rid = gen_id('ar_')
        db.execute(
            "INSERT INTO assignment_rules (id,priority,condition_field,condition_value,assignee_id,is_active,created_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (rid, int(data.get('priority', 0)), data.get('condition_field', '').strip(),
             data.get('condition_value', '').strip(), data.get('assignee_id', '').strip(),
             1, now)
        )
        db.commit()
        db.close()
        return json_resp(self, {"id": rid, "status": "created"}, 201)

    def _handle_update_assignment_rule(self, user, data):
        """PUT update assignment rules (bulk replace)."""
        if user['role'] not in ('admin', 'supervisor'):
            return json_resp(self, {"error": "无权限"}, 403)
        db = get_db()
        # If rules[] array provided, bulk replace
        rules = data.get('rules', [])
        if rules:
            db.execute("DELETE FROM assignment_rules")
            now = datetime.now().strftime('%Y-%m-%d %H:%M')
            for r in rules:
                rid = r.get('id') or gen_id('ar_')
                db.execute(
                    "INSERT INTO assignment_rules (id,priority,condition_field,condition_value,assignee_id,is_active,created_at) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (rid, int(r.get('priority', 0)), r.get('condition_field', '').strip(),
                     r.get('condition_value', '').strip(), r.get('assignee_id', '').strip(),
                     int(r.get('is_active', 1)), now)
                )
        else:
            # Single rule update by ID
            rid = data.get('id', '').strip()
            if not rid:
                db.close()
                return json_resp(self, {"error": "缺少规则ID"}, 400)
            updates = {}
            for k in ('priority', 'condition_field', 'condition_value', 'assignee_id', 'is_active'):
                if k in data:
                    updates[k] = data[k]
            if updates:
                updates['id'] = rid
                set_clause = ", ".join(f"{k}=?" for k in updates)
                db.execute(f"UPDATE assignment_rules SET {set_clause} WHERE id=?", list(updates.values()) + [rid])
        db.commit()
        db.close()
        return json_resp(self, {"status": "updated"})

    # ─── Follow-up Templates ─────────────────────────────────────────────────

    def _handle_list_followup_templates(self, user):
        db = get_db()
        templates = db.execute(
            "SELECT id, name, type, content_template, sort_order, is_active, created_at "
            "FROM followup_templates WHERE is_active=1 ORDER BY sort_order ASC"
        ).fetchall()
        db.close()
        return json_resp(self, {"templates": [dict(r) for r in templates]})

    def _handle_get_followup_template(self, user, template_id):
        db = get_db()
        t = db.execute("SELECT * FROM followup_templates WHERE id=?", (template_id,)).fetchone()
        db.close()
        if not t:
            return json_resp(self, {"error": "Not Found"}, 404)
        return json_resp(self, {"template": dict(t)})

    def _handle_create_followup_template(self, user, data):
        if user['role'] not in ('admin', 'supervisor'):
            return json_resp(self, {"error": "无权限"}, 403)
        db = get_db()
        tid = gen_id('ft_')
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db.execute(
            "INSERT INTO followup_templates (id,name,type,content_template,sort_order,is_active,created_at) VALUES (?,?,?,?,?,1,?)",
            (tid, data.get('name','').strip(), data.get('type','note').strip(),
             data.get('content_template','').strip(), int(data.get('sort_order', 0)), now)
        )
        db.commit()
        db.close()
        return json_resp(self, {"id": tid, "status": "created"}, 201)

    def _handle_update_followup_template(self, user, template_id, data):
        if user['role'] not in ('admin', 'supervisor'):
            return json_resp(self, {"error": "无权限"}, 403)
        db = get_db()
        existing = db.execute("SELECT id FROM followup_templates WHERE id=?", (template_id,)).fetchone()
        if not existing:
            db.close()
            return json_resp(self, {"error": "Not Found"}, 404)
        updates = []
        params = []
        for field in ('name', 'type', 'content_template'):
            if field in data:
                updates.append(f"{field}=?")
                params.append(data[field].strip())
        if 'sort_order' in data:
            updates.append("sort_order=?")
            params.append(int(data['sort_order']))
        if 'is_active' in data:
            updates.append("is_active=?")
            params.append(1 if data['is_active'] else 0)
        if updates:
            params.append(template_id)
            db.execute(f"UPDATE followup_templates SET {','.join(updates)} WHERE id=?", params)
            db.commit()
        db.close()
        return json_resp(self, {"status": "updated"})

    def _handle_delete_followup_template(self, user, template_id):
        if user['role'] not in ('admin', 'supervisor'):
            return json_resp(self, {"error": "无权限"}, 403)
        db = get_db()
        db.execute("UPDATE followup_templates SET is_active=0 WHERE id=?", (template_id,))
        db.commit()
        db.close()
        return json_resp(self, {"status": "deleted"})

    # ─── Overdue Follow-ups ──────────────────────────────────────────────────

    def _handle_followup_overdue(self, user, params):
        range_type = params.get('range', ['overdue'])[0]
        db = get_db()
        today = datetime.now().strftime('%Y-%m-%d')
        if range_type == 'overdue':
            rows = db.execute(
                "SELECT l.id as lead_id, l.name as lead_name, a.next_action, a.next_action_date, "
                "l.assignee_id, u.name as assignee_name, "
                "julianday(?) - julianday(a.next_action_date) as overdue_days "
                "FROM activities a JOIN leads l ON a.lead_id = l.id "
                "LEFT JOIN users u ON l.assignee_id = u.id "
                "WHERE a.next_action_date IS NOT NULL AND a.next_action_date != '' "
                "AND a.next_action_date < ? AND l.status NOT IN ('enrolled','closed','lost') "
                "AND l.merge_target_id IS NULL "
                "AND a.id = (SELECT a2.id FROM activities a2 WHERE a2.lead_id=l.id AND a2.next_action_date IS NOT NULL AND a2.next_action_date != '' ORDER BY a2.created_at DESC LIMIT 1) "
                "ORDER BY overdue_days DESC",
                (today, today)
            ).fetchall()
        elif range_type == 'today':
            rows = db.execute(
                "SELECT l.id as lead_id, l.name as lead_name, a.next_action, a.next_action_date, "
                "l.assignee_id, u.name as assignee_name, 0 as overdue_days "
                "FROM activities a JOIN leads l ON a.lead_id = l.id "
                "LEFT JOIN users u ON l.assignee_id = u.id "
                "WHERE a.next_action_date = ? AND l.merge_target_id IS NULL "
                "AND a.id = (SELECT a2.id FROM activities a2 WHERE a2.lead_id=l.id AND a2.next_action_date IS NOT NULL ORDER BY a2.created_at DESC LIMIT 1) "
                "ORDER BY l.name",
                (today,)
            ).fetchall()
        else:  # upcoming
            rows = db.execute(
                "SELECT l.id as lead_id, l.name as lead_name, a.next_action, a.next_action_date, "
                "l.assignee_id, u.name as assignee_name, 0 as overdue_days "
                "FROM activities a JOIN leads l ON a.lead_id = l.id "
                "LEFT JOIN users u ON l.assignee_id = u.id "
                "WHERE a.next_action_date > ? AND l.merge_target_id IS NULL "
                "AND a.id = (SELECT a2.id FROM activities a2 WHERE a2.lead_id=l.id AND a2.next_action_date IS NOT NULL ORDER BY a2.created_at DESC LIMIT 1) "
                "ORDER BY a.next_action_date ASC",
                (today,)
            ).fetchall()
        db.close()
        return json_resp(self, {"followups": [dict(r) for r in rows]})

    # ─── Follow-up Stats ─────────────────────────────────────────────────────

    def _handle_followup_stats(self, user, params):
        db = get_db()
        today = datetime.now().strftime('%Y-%m-%d')
        # Determine scope: admin/supervisor see all, others see own
        if user['role'] in ('admin', 'supervisor') and params.get('user_id', [None])[0]:
            user_filter = params['user_id'][0]
            user_cond = "AND l.assignee_id=?"
            user_params = [user_filter]
        elif user['role'] in ('admin', 'supervisor'):
            user_cond = ""
            user_params = []
        else:
            user_cond = "AND l.assignee_id=?"
            user_params = [user['id']]

        # Today followups count
        today_count = db.execute(
            f"SELECT COUNT(DISTINCT a.id) FROM activities a JOIN leads l ON a.lead_id=l.id "
            f"WHERE a.created_at >= ? AND a.created_at < ? {user_cond}",
            [today, (datetime.now()+timedelta(days=1)).strftime('%Y-%m-%d')] + user_params
        ).fetchone()[0]

        # Overdue followups count
        overdue_count = db.execute(
            f"SELECT COUNT(DISTINCT l.id) FROM leads l "
            f"WHERE l.id IN (SELECT lead_id FROM activities WHERE next_action_date IS NOT NULL AND next_action_date != '' AND next_action_date < ?) "
            f"AND l.status NOT IN ('enrolled','closed','lost') AND l.merge_target_id IS NULL {user_cond}",
            [today] + user_params
        ).fetchone()[0]

        # Total active leads assigned to user
        total_assigned = db.execute(
            f"SELECT COUNT(*) FROM leads WHERE status NOT IN ('closed','lost') AND merge_target_id IS NULL AND assignee_id IS NOT NULL {user_cond}",
            user_params
        ).fetchone()[0] if user_params else 0

        if not user_params:
            total_assigned = db.execute(
                "SELECT COUNT(*) FROM leads WHERE status NOT IN ('closed','lost') AND merge_target_id IS NULL AND assignee_id IS NOT NULL"
            ).fetchone()[0]

        # Leads with at least one activity (followed up)
        followed = db.execute(
            f"SELECT COUNT(DISTINCT l.id) FROM leads l "
            f"WHERE l.id IN (SELECT DISTINCT lead_id FROM activities) "
            f"AND l.status NOT IN ('closed','lost') AND l.merge_target_id IS NULL {user_cond}",
            user_params
        ).fetchone()[0]

        followup_rate = round(followed / total_assigned * 100, 1) if total_assigned > 0 else 0

        # Average response time (days from assignment to first activity)
        avg_response = db.execute(
            f"SELECT AVG(julianday(a.created_at) - julianday(l.assigned_at)) as avg_days "
            f"FROM activities a JOIN leads l ON a.lead_id=l.id "
            f"WHERE l.assigned_at IS NOT NULL AND l.assigned_at != '' AND l.merge_target_id IS NULL {user_cond}",
            user_params
        ).fetchone()[0] or 0

        db.close()
        return json_resp(self, {
            "stats": {
                "today_followups": today_count,
                "overdue_followups": overdue_count,
                "total_assigned": total_assigned,
                "followed_up": followed,
                "followup_rate": followup_rate,
                "avg_response_days": round(avg_response, 1),
            }
        })

    def _handle_self_claim(self, user, lead_id):
        """Consultant self-claim a lead from the pool."""
        role = user['role']
        if role not in ('consultant', 'cs', 'academic'):
            return json_resp(self, {"error": "无权限抢单"}, 403)

        db = get_db()
        lead = db.execute("SELECT * FROM leads WHERE id=?", (lead_id,)).fetchone()
        if not lead:
            db.close()
            return json_resp(self, {"error": "线索不存在"}, 404)

        # Only allow claiming leads that are pending (in pool) or in pool_status='pool'
        if lead['pool_status'] not in ('pool',) and lead['status'] != 'pending':
            db.close()
            return json_resp(self, {"error": "该线索不可抢单"}, 400)

        if lead['assignee_id']:
            db.close()
            return json_resp(self, {"error": "该线索已有跟进人"}, 400)

        # Check daily claim limit
        cfg = _get_pool_config()
        daily_limit = int(cfg.get('daily_claim_limit', '10'))
        today = datetime.now().strftime('%Y-%m-%d')
        today_claims = db.execute(
            "SELECT COUNT(*) FROM activities WHERE user_id=? AND content LIKE '%抢单成功%' AND created_at LIKE ?",
            (user['id'], today + '%')
        ).fetchone()[0]
        if today_claims >= daily_limit:
            db.close()
            return json_resp(self, {"error": f"今日抢单已达上限 ({daily_limit}条)"}, 429)

        # Check active lead limit
        active_limit = int(cfg.get('active_lead_limit', '30'))
        active_count = db.execute(
            "SELECT COUNT(*) FROM leads WHERE assignee_id=? AND status NOT IN ('closed','lost') AND merge_target_id IS NULL",
            (user['id'],)
        ).fetchone()[0]
        if active_count >= active_limit:
            db.close()
            return json_resp(self, {"error": f"在手线索已达上限 ({active_limit}条)，请先处理现有线索"}, 429)

        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db.execute(
            "UPDATE leads SET assignee_id=?, status='assigned', pool_status='active', assigned_at=?, updated_at=? WHERE id=?",
            (user['id'], now, now, lead_id)
        )
        db.execute(
            "INSERT INTO activities (id,lead_id,user_id,type,content,created_at) VALUES (?,?,?,?,?,?)",
            (gen_id('a_'), lead_id, user['id'], 'note',
             f"系统：{user['name']} 从公海池抢单成功", now)
        )
        db.commit()
        db.close()
        _audit(user['id'], 'self_claim', 'lead', lead_id, f'{user["name"]} 抢单')
        return json_resp(self, {"status": "claimed", "assignee_name": user['name']})

    def _handle_batch_leads(self, user, data):
        """Batch operations on leads: assign, change status, return to pool."""
        action = data.get('action', '')
        lead_ids = data.get('lead_ids', [])
        if not lead_ids or not isinstance(lead_ids, list) or len(lead_ids) == 0:
            return json_resp(self, {"error": "请选择线索"}, 400)
        if user['role'] not in ('admin', 'supervisor'):
            return json_resp(self, {"error": "无权限"}, 403)

        db = get_db()
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        updated = 0

        if action == 'assign':
            assignee_id = data.get('assignee_id', '').strip()
            if not assignee_id:
                db.close()
                return json_resp(self, {"error": "请选择跟进人"}, 400)
            for lid in lead_ids:
                db.execute(
                    "UPDATE leads SET assignee_id=?, status='assigned', assigned_at=?, updated_at=? WHERE id=?",
                    (assignee_id, now, now, lid)
                )
                db.execute(
                    "INSERT INTO activities (id,lead_id,user_id,type,content,created_at) VALUES (?,?,?,?,?,?)",
                    (gen_id('a_'), lid, user['id'], 'note',
                     f"系统：批量分配给 {data.get('assignee_name','')}", now)
                )
                updated += 1
        elif action == 'status':
            new_status = data.get('status', '')
            if new_status not in ('pending', 'assigned', 'following', 'trial', 'enrolled', 'closed', 'lost'):
                db.close()
                return json_resp(self, {"error": "无效状态"}, 400)
            for lid in lead_ids:
                db.execute("UPDATE leads SET status=?, updated_at=? WHERE id=?", (new_status, now, lid))
                updated += 1
        elif action == 'pool':
            for lid in lead_ids:
                lead = db.execute("SELECT pool_return_count FROM leads WHERE id=?", (lid,)).fetchone()
                rc = (lead['pool_return_count'] or 0) + 1 if lead else 1
                ns = 'lost' if rc >= 3 else 'pending'
                ps = 'silent' if rc >= 3 else 'pool'
                db.execute(
                    "UPDATE leads SET status=?, pool_status=?, pool_return_count=?, assignee_id=NULL, updated_at=? WHERE id=?",
                    (ns, ps, rc, now, lid)
                )
                db.execute(
                    "INSERT INTO pool_return_log (id, lead_id, reason, returned_by, created_at) VALUES (?,?,?,?,?)",
                    (gen_id('pr_'), lid, '批量回池', user['id'], now)
                )
                updated += 1
        else:
            db.close()
            return json_resp(self, {"error": "未知操作"}, 400)

        db.commit()
        db.close()
        _audit(user['id'], 'batch', 'lead', '', f'批量操作: {action} ({updated}条)')
        return json_resp(self, {"status": "done", "action": action, "updated": updated})

    def _handle_import_leads(self, user, data):
        """CSV batch import of leads. Accepts CSV text or array of dicts."""
        if user['role'] not in ('admin', 'supervisor'):
            return json_resp(self, {"error": "无权限"}, 403)
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db = get_db()

        # Accept either CSV string or pre-parsed array
        rows = data.get('rows', None)
        csv_text = data.get('csv', '')

        if rows and isinstance(rows, list):
            entries = rows
        elif csv_text.strip():
            # Simple CSV parser (handles basic cases, no quoted commas)
            lines = [l for l in csv_text.strip().split('\n') if l.strip()]
            if len(lines) < 2:
                db.close()
                return json_resp(self, {"error": "CSV 至少需要表头和一行数据"}, 400)
            headers = [h.strip().lower() for h in lines[0].split(',')]
            # Map common header names to field names
            header_map = {
                'name': 'name', '姓名': 'name', '学生姓名': 'name',
                'phone': 'phone', '手机': 'phone', '手机号': 'phone', '电话': 'phone',
                'wechat': 'wechat', '微信': 'wechat', '微信号': 'wechat',
                'country': 'country', '国家': 'country', '意向国家': 'country',
                'grade': 'grade', '年级': 'grade',
                'subject': 'subject', '科目': 'subject', '辅导科目': 'subject',
                'source': 'source', '来源': 'source', '渠道': 'source',
                'notes': 'notes', '备注': 'notes',
                'utm_source': 'utm_source', 'utm来源': 'utm_source',
                'utm_campaign': 'utm_campaign', 'utm活动': 'utm_campaign',
                'utm_medium': 'utm_medium', 'utm媒介': 'utm_medium',
            }
            mapped = []
            for h in headers:
                mapped.append(header_map.get(h, h))
            entries = []
            for line in lines[1:]:
                vals = [v.strip() for v in line.split(',')]
                entry = {}
                for i, h in enumerate(mapped):
                    if i < len(vals) and vals[i]:
                        entry[h] = vals[i]
                entries.append(entry)
        else:
            db.close()
            return json_resp(self, {"error": "请提供 CSV 文本或 rows 数组"}, 400)

        # Import each entry
        success, errors = 0, []
        for idx, entry in enumerate(entries):
            name = entry.get('name', '').strip()
            if not name:
                errors.append(f"第 {idx+2} 行: 姓名为空")
                continue
            try:
                lead_id = gen_id('l_')
                phone = entry.get('phone', '').strip()
                wechat = entry.get('wechat', '').strip()
                db.execute(
                    "INSERT INTO leads (id,name,phone,wechat,country,grade,subject,source,status,created_by_id,notes,created_at,updated_at,"
                    "utm_source,utm_campaign,utm_medium,campaign,ad_source,landing_page) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (lead_id, name, phone, wechat,
                     entry.get('country', ''), entry.get('grade', ''),
                     entry.get('subject', ''), entry.get('source', ''), 'pending', user['id'],
                     entry.get('notes', ''), now, now,
                     entry.get('utm_source', ''), entry.get('utm_campaign', ''),
                     entry.get('utm_medium', ''), entry.get('campaign', ''),
                     entry.get('ad_source', ''), entry.get('landing_page', ''))
                )
                success += 1
            except Exception as e:
                errors.append(f"第 {idx+2} 行: {str(e)}")

        db.commit()
        db.close()
        _audit(user['id'], 'import', 'lead', '', f'导入线索: {success}成功, {len(errors)}失败')
        return json_resp(self, {"status": "done", "success": success, "errors": errors})

    # ═══════════════════════ SMART PARSE ═══════════════════════════════════

    def _handle_parse_lead_text(self, user, data):
        """Parse unstructured text into lead fields. Returns parsed result + duplicates."""
        text = data.get('text', '').strip()
        if not text:
            return json_resp(self, {"error": "请输入文本内容"}, 400)

        result = parse_lead_text(text)

        # Check duplicates
        if result.get('phone'):
            db = get_db()
            dup = db.execute(
                "SELECT id, name, status FROM leads WHERE phone=? AND status NOT IN ('lost','closed') AND merge_target_id IS NULL",
                (result['phone'],)
            ).fetchone()
            if dup:
                result['duplicate'] = {'id': dup['id'], 'name': dup['name'], 'status': dup['status']}
            db.close()

        if not result.get('duplicate') and result.get('wechat'):
            db = get_db()
            dup = db.execute(
                "SELECT id, name, status FROM leads WHERE wechat=? AND status NOT IN ('lost','closed') AND merge_target_id IS NULL",
                (result['wechat'],)
            ).fetchone()
            if dup:
                result['duplicate'] = {'id': dup['id'], 'name': dup['name'], 'status': dup['status']}
            db.close()

        return json_resp(self, {"result": result})

    # ─── Activities ───────────────────────────────────────────────────────

    def _handle_list_activities(self, user, lead_id):
        page, size, offset = get_page_params(self._get_params())
        db = get_db()
        rows = db.execute(
            "SELECT a.*, u.name as user_name FROM activities a "
            "LEFT JOIN users u ON a.user_id = u.id WHERE a.lead_id=? "
            "ORDER BY a.created_at DESC LIMIT ? OFFSET ?", (lead_id, size, offset)
        ).fetchall()
        db.close()
        return json_resp(self, {"activities": [dict(r) for r in rows]})

    def _get_params(self):
        parsed = urllib.parse.urlparse(self.path)
        return urllib.parse.parse_qs(parsed.query)

    def _handle_add_activity(self, user, lead_id, data):
        content = data.get('content', '').strip()
        if not content:
            return json_resp(self, {"error": "跟进内容不能为空"}, 400)
        act_id = gen_id('a_')
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db = get_db()
        lead = db.execute("SELECT status FROM leads WHERE id=?", (lead_id,)).fetchone()
        if lead and lead['status'] == 'pending':
            db.execute("UPDATE leads SET status='following', updated_at=? WHERE id=?", (now, lead_id))
        next_date = data.get('next_action_date', '').strip()
        db.execute(
            "INSERT INTO activities (id,lead_id,user_id,type,content,next_action,next_action_date,created_at) VALUES (?,?,?,?,?,?,?,?)",
            (act_id, lead_id, user['id'], data.get('type', 'note'), content, data.get('next_action', '').strip(), next_date, now)
        )
        db.commit()
        db.close()
        return json_resp(self, {"id": act_id, "status": "created"}, 201)

    # ═══════════════════ COORDINATOR DASHBOARD ═══════════════════════════════

    def _handle_coordinator_dashboard(self, user):
        db = get_db()

        pending_schedules = db.execute(
            "SELECT s.*, l.name as lead_name, u.name as tutor_name FROM schedules s "
            "JOIN leads l ON s.lead_id = l.id LEFT JOIN users u ON s.tutor_id = u.id "
            "WHERE s.status='pending' ORDER BY s.created_at DESC LIMIT 20"
        ).fetchall()

        pending_consumptions = db.execute(
            "SELECT c.*, l.name as lead_name, p.package_name FROM consumption_log c "
            "JOIN course_packages p ON c.package_id = p.id JOIN leads l ON p.lead_id = l.id "
            "WHERE c.status='pending_confirm' ORDER BY c.created_at DESC LIMIT 20"
        ).fetchall()

        today = datetime.now().strftime('%Y-%m-%d')
        today_lessons = db.execute(
            "SELECT s.*, l.name as lead_name, u.name as tutor_name FROM schedules s "
            "JOIN leads l ON s.lead_id = l.id LEFT JOIN users u ON s.tutor_id = u.id "
            "WHERE s.status='confirmed' AND date(s.start_time)=? ORDER BY s.start_time", (today,)
        ).fetchall()

        # Stats
        stats = {
            'pending_schedules': db.execute("SELECT COUNT(*) FROM schedules WHERE status='pending'").fetchone()[0],
            'pending_consumptions': db.execute("SELECT COUNT(*) FROM consumption_log WHERE status='pending_confirm'").fetchone()[0],
            'today_lessons': len(today_lessons),
            'active_packages': db.execute("SELECT COUNT(*) FROM course_packages WHERE status='active'").fetchone()[0],
            'active_tutors': db.execute("SELECT COUNT(*) FROM users WHERE is_tutor=1 AND is_active=1").fetchone()[0],
        }

        # Low hours alert
        low_hours = db.execute(
            "SELECT p.*, l.name as lead_name FROM course_packages p "
            "JOIN leads l ON p.lead_id = l.id "
            "WHERE p.status='active' AND (p.total_hours - p.used_hours) < 5 "
            "ORDER BY (p.total_hours - p.used_hours) LIMIT 10"
        ).fetchall()

        # Weekly lesson stats by timezone
        week_start = (datetime.now() - timedelta(days=datetime.now().weekday())).strftime('%Y-%m-%d')
        week_end = (datetime.now() + timedelta(days=6 - datetime.now().weekday())).strftime('%Y-%m-%d')
        weekly_lessons = db.execute(
            "SELECT COUNT(*) as total FROM schedules WHERE date(start_time) BETWEEN ? AND ?",
            (week_start, week_end)
        ).fetchone()[0]
        weekly_completed = db.execute(
            "SELECT COUNT(*) as total FROM schedules WHERE date(start_time) BETWEEN ? AND ? AND status='completed'",
            (week_start, week_end)
        ).fetchone()[0]

        db.close()
        return json_resp(self, {
            "stats": stats,
            "pending_schedules": [dict(r) for r in pending_schedules],
            "pending_consumptions": [dict(r) for r in pending_consumptions],
            "today_lessons": [dict(r) for r in today_lessons],
            "low_hours_alerts": [dict(r) for r in low_hours],
            "weekly_stats": {
                "total": weekly_lessons,
                "completed": weekly_completed,
                "completion_rate": round(weekly_completed / weekly_lessons * 100, 1) if weekly_lessons > 0 else 0,
            },
        })

    # ═══════════════════ PACKAGES ═══════════════════════════════════════════

    def _handle_list_packages(self, user, params):
        db = get_db()
        conditions, args = [], []
        lead_id = params.get('lead_id', [None])[0]
        if lead_id:
            conditions.append("p.lead_id=?")
            args.append(lead_id)
        status = params.get('status', [None])[0]
        if status and status != 'all':
            conditions.append("p.status=?")
            args.append(status)
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        page, size, offset = get_page_params(params)
        rows = db.execute(
            f"SELECT p.*, l.name as lead_name FROM course_packages p "
            f"JOIN leads l ON p.lead_id = l.id {where} "
            f"ORDER BY p.created_at DESC LIMIT ? OFFSET ?", args + [size, offset]
        ).fetchall()
        total = db.execute(f"SELECT COUNT(*) FROM course_packages p {where}", args).fetchone()[0]
        db.close()
        return json_resp(self, {"packages": [dict(r) for r in rows], "total": total})

    def _handle_get_package(self, user, pkg_id):
        db = get_db()
        pkg = db.execute(
            "SELECT p.*, l.name as lead_name FROM course_packages p "
            "JOIN leads l ON p.lead_id = l.id WHERE p.id=?", (pkg_id,)
        ).fetchone()
        if not pkg:
            db.close()
            return json_resp(self, {"error": "Not Found"}, 404)
        cons = db.execute(
            "SELECT c.*, s.topic, u.name as confirmed_by_name FROM consumption_log c "
            "LEFT JOIN schedules s ON c.schedule_id = s.id "
            "LEFT JOIN users u ON c.confirmed_by = u.id "
            "WHERE c.package_id=? ORDER BY c.created_at DESC", (pkg_id,)
        ).fetchall()
        db.close()
        return json_resp(self, {"package": dict(pkg), "consumptions": [dict(c) for c in cons]})

    def _handle_create_package(self, user, data):
        name = data.get('lead_id', '').strip()
        if not name:
            return json_resp(self, {"error": "请选择学生"}, 400)
        pkg_id = gen_id('p_')
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        total = float(data.get('total_hours', 0))
        db = get_db()
        db.execute(
            "INSERT INTO course_packages (id,lead_id,package_name,total_hours,used_hours,price,unit_price,valid_from,valid_until,status,created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (pkg_id, data['lead_id'], data.get('package_name', ''), total, 0,
             float(data.get('price', 0)), float(data.get('unit_price', 0)),
             data.get('valid_from', ''), data.get('valid_until', ''), 'active', now)
        )
        db.commit()
        db.close()
        _audit(user['id'], 'create', 'package', pkg_id, f'创建课时包: {total}h')
        return json_resp(self, {"id": pkg_id, "status": "created"}, 201)

    def _handle_update_package(self, user, pkg_id, data):
        db = get_db()
        updates = {}
        for field in ('package_name', 'total_hours', 'price', 'unit_price', 'valid_from', 'valid_until', 'status'):
            if field in data:
                updates[field] = data[field]
        if not updates:
            db.close()
            return json_resp(self, {"error": "No fields to update"}, 400)
        set_clause = ", ".join(f"{k}=?" for k in updates)
        vals = list(updates.values()) + [pkg_id]
        db.execute(f"UPDATE course_packages SET {set_clause} WHERE id=?", vals)
        db.commit()
        db.close()
        _audit(user['id'], 'update', 'package', pkg_id, '更新课时包')
        return json_resp(self, {"status": "updated"})

    # ═══════════════════ SCHEDULES ═══════════════════════════════════════════

    def _handle_list_schedules(self, user, params):
        db = get_db()
        conditions, args = [], []

        for key, col in [('status', 's.status'), ('lead_id', 's.lead_id'), ('tutor_id', 's.tutor_id')]:
            v = params.get(key, [None])[0]
            if v and v != 'all':
                conditions.append(f"{col}=?")
                args.append(v)

        date_from = params.get('from', [None])[0]
        if date_from:
            conditions.append("s.start_time >= ?")
            args.append(date_from)
        date_to = params.get('to', [None])[0]
        if date_to:
            conditions.append("s.start_time <= ?")
            args.append(date_to)

        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        page, size, offset = get_page_params(params)

        rows = db.execute(
            f"SELECT s.*, l.name as lead_name, u.name as tutor_name, p.package_name "
            f"FROM schedules s "
            f"JOIN leads l ON s.lead_id = l.id "
            f"LEFT JOIN users u ON s.tutor_id = u.id "
            f"LEFT JOIN course_packages p ON s.package_id = p.id "
            f"{where} ORDER BY s.start_time DESC LIMIT ? OFFSET ?", args + [size, offset]
        ).fetchall()
        total = db.execute(f"SELECT COUNT(*) FROM schedules s {where}", args).fetchone()[0]
        db.close()
        return json_resp(self, {"schedules": [dict(r) for r in rows], "total": total})

    def _handle_get_schedule(self, user, sch_id):
        db = get_db()
        s = db.execute(
            "SELECT s.*, l.name as lead_name, u.name as tutor_name, u2.name as coordinator_name, p.package_name "
            "FROM schedules s "
            "JOIN leads l ON s.lead_id = l.id "
            "LEFT JOIN users u ON s.tutor_id = u.id "
            "LEFT JOIN users u2 ON s.coordinator_id = u2.id "
            "LEFT JOIN course_packages p ON s.package_id = p.id "
            "WHERE s.id=?", (sch_id,)
        ).fetchone()
        if not s:
            db.close()
            return json_resp(self, {"error": "Not Found"}, 404)
        db.close()
        return json_resp(self, {"schedule": dict(s)})

    def _handle_create_schedule(self, user, data):
        lead_id = data.get('lead_id', '').strip()
        tutor_id = data.get('tutor_id', '').strip()
        start_time = data.get('start_time', '').strip()
        duration = int(data.get('duration_minutes', 60))
        if not lead_id or not start_time:
            return json_resp(self, {"error": "学生和上课时间为必填"}, 400)

        # Parse end time
        try:
            from datetime import datetime as dt
            st = dt.fromisoformat(start_time)
            et = st + timedelta(minutes=duration)
            end_time = et.isoformat()
        except:
            return json_resp(self, {"error": "时间格式错误"}, 400)

        # Conflict detection
        if tutor_id:
            db = get_db()
            conflict = db.execute(
                "SELECT s.id, l.name FROM schedules s JOIN leads l ON s.lead_id = l.id "
                "WHERE s.tutor_id=? AND s.status IN ('pending','confirmed') "
                "AND s.start_time < ? AND s.end_time > ?",
                (tutor_id, end_time, start_time)
            ).fetchone()
            if conflict:
                db.close()
                return json_resp(self, {"error": f"老师该时段已有排课: {conflict['name']}"}, 409)

        sch_id = gen_id('sch_')
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        sched_type = data.get('schedule_type', 'regular')
        if not db:
            db = get_db()
        db.execute(
            "INSERT INTO schedules (id,lead_id,tutor_id,package_id,coordinator_id,type,schedule_type,start_time,end_time,timezone,duration_minutes,status,topic,notes,created_at,updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (sch_id, lead_id, tutor_id or None, data.get('package_id', '') or None,
             user['id'], data.get('type', 'single'), sched_type, start_time, end_time,
             data.get('timezone', 'Asia/Shanghai'), duration, 'confirmed',
             data.get('topic', ''), data.get('notes', ''), now, now)
        )
        # Auto-transition lead to trial if trial schedule
        if sched_type == 'trial':
            lead = db.execute("SELECT status FROM leads WHERE id=?", (lead_id,)).fetchone()
            if lead and lead['status'] not in ('enrolled', 'closed', 'lost'):
                db.execute("UPDATE leads SET status='trial' WHERE id=?", (lead_id,))
                db.execute(
                    "INSERT INTO activities (id,lead_id,type,content,created_at) VALUES (?,?,?,?,?)",
                    (gen_id('act_'), lead_id, 'trial', '创建试听排课，线索状态自动变更为试听中', now)
                )
        db.commit()
        # Notify tutor of new schedule
        if tutor_id:
            self._create_notification_simple(db, sch_id, 'schedule_change',
                '新排课通知', f'您有一节新课安排在 {start_time[:16]}', 'schedule', tutor_id)
        # Notify assignee if trial schedule
        if sched_type == 'trial':
            ld = db.execute("SELECT assignee FROM leads WHERE id=?", (lead_id,)).fetchone()
            if ld and ld['assignee']:
                self._create_notification_simple(db, sch_id, 'lead_update',
                    '试听排课提醒', f'学生已创建试听排课，请关注跟进', 'lead', ld['assignee'])
        db.close()
        _audit(user['id'], 'create', 'schedule', sch_id, f'排课: {lead_id} + {tutor_id}')
        return json_resp(self, {"id": sch_id, "status": "created"}, 201)

    def _handle_update_schedule(self, user, sch_id, data):
        db = get_db()
        updates = {}
        for field in ('tutor_id', 'start_time', 'end_time', 'duration_minutes', 'status', 'topic', 'notes', 'timezone', 'classin_lesson_id', 'classin_room_url', 'classin_summary', 'teacher_feedback'):
            if field in data:
                updates[field] = data[field]
        if not updates:
            db.close()
            return json_resp(self, {"error": "No fields to update"}, 400)
        updates['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        set_clause = ", ".join(f"{k}=?" for k in updates)
        vals = list(updates.values()) + [sch_id]
        db.execute(f"UPDATE schedules SET {set_clause} WHERE id=?", vals)
        db.commit()
        db.close()
        _audit(user['id'], 'update', 'schedule', sch_id, '更新排课')
        return json_resp(self, {"status": "updated"})

    def _handle_confirm_schedule(self, user, sch_id, data):
        db = get_db()
        db.execute("UPDATE schedules SET status='confirmed', updated_at=? WHERE id=?",
                   (datetime.now().strftime('%Y-%m-%d %H:%M'), sch_id))
        db.commit()
        db.close()
        _audit(user['id'], 'confirm', 'schedule', sch_id, '确认排课')
        return json_resp(self, {"status": "confirmed"})

    def _handle_cancel_schedule(self, user, sch_id, data):
        db = get_db()
        reason = data.get('reason', '')
        db.execute("UPDATE schedules SET status='cancelled', notes=notes || ? , updated_at=? WHERE id=?",
                   (f"\n取消原因: {reason}" if reason else '', datetime.now().strftime('%Y-%m-%d %H:%M'), sch_id))
        # Notify tutor
        s = db.execute("SELECT tutor_id, lead_id, start_time FROM schedules WHERE id=?", (sch_id,)).fetchone()
        if s and s['tutor_id']:
            self._create_notification_simple(db, sch_id, 'schedule_change',
                '排课取消通知', f'您的排课 {s["start_time"][:16] if s["start_time"] else ""} 已被取消: {reason}',
                'schedule', s['tutor_id'])
        db.commit()
        db.close()
        _audit(user['id'], 'cancel', 'schedule', sch_id, f'取消排课: {reason}')
        return json_resp(self, {"status": "cancelled"})

    def _handle_complete_schedule(self, user, sch_id, data):
        """Mark schedule as completed, auto-create consumption_log."""
        db = get_db()
        s = db.execute(
            "SELECT s.*, l.name as lead_name, l.assignee_id FROM schedules s "
            "JOIN leads l ON s.lead_id = l.id WHERE s.id=?", (sch_id,)
        ).fetchone()
        if not s:
            db.close()
            return json_resp(self, {"error": "排课不存在"}, 404)
        if s['status'] != 'confirmed':
            db.close()
            return json_resp(self, {"error": "只能完成已确认的排课"}, 400)

        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        hours = (s['duration_minutes'] or 60) / 60.0

        # Mark schedule completed
        db.execute(
            "UPDATE schedules SET status='completed', completed_at=?, updated_at=? WHERE id=?",
            (now, now, sch_id)
        )

        # Auto-create consumption_log (pending_confirm)
        cons_id = gen_id('cons_')
        db.execute(
            "INSERT INTO consumption_log (id, package_id, schedule_id, hours_scheduled, hours_actual, hours_consumed, status, notes, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?)",
            (cons_id, s['package_id'] or None, sch_id, hours, hours, 0,
             'pending_confirm', f'由排课完成自动创建', now)
        )

        # Add activity on lead
        db.execute(
            "INSERT INTO activities (id, lead_id, user_id, type, content, created_at) VALUES (?,?,?,?,?,?)",
            (gen_id('a_'), s['lead_id'], user['id'], 'note',
             f"系统：排课已标记完成 [{s.get('topic','') or '无主题'}]", now)
        )

        db.commit()

        # Notify tutor and assignee
        if s['tutor_id']:
            self._create_notification_simple(db, sch_id, 'schedule_change',
                '排课完成确认', f'排课 [{s.get("topic","") or "无主题"}] 已完成，待确认课时',
                'schedule', s['tutor_id'])
        if s['assignee_id']:
            self._create_notification_simple(db, sch_id, 'schedule_change',
                '学生排课已完成', f'{s["lead_name"]} 的课程已完成，可查看消耗确认',
                'schedule', s['assignee_id'])

        db.close()
        _audit(user['id'], 'complete', 'schedule', sch_id, f'完成排课: {hours}h')
        return json_resp(self, {
            "status": "completed",
            "consumption_id": cons_id,
            "hours_scheduled": hours
        })

    # ═══════════════════ LESSON SERIES ══════════════════════════════════════

    def _handle_list_series(self, user, params):
        db = get_db()
        lead_id = params.get('lead_id', [None])[0]
        rows = db.execute(
            "SELECT s.*, l.name as lead_name, u.name as tutor_name, "
            "(SELECT COUNT(*) FROM schedules sch WHERE sch.lead_id=s.lead_id AND sch.type='recurring' AND sch.notes LIKE '%' || s.id || '%') as generated_count "
            "FROM lesson_series s "
            "JOIN leads l ON s.lead_id = l.id "
            "LEFT JOIN users u ON s.tutor_id = u.id "
            + ("WHERE s.lead_id=?" if lead_id else "") + " ORDER BY s.created_at DESC",
            (lead_id,) if lead_id else ()
        ).fetchall()
        db.close()
        return json_resp(self, {"series": [dict(r) for r in rows]})

    def _handle_create_series(self, user, data):
        lead_id = data.get('lead_id', '').strip()
        if not lead_id:
            return json_resp(self, {"error": "请选择学生"}, 400)
        sr_id = gen_id('sr_')
        db = get_db()
        db.execute(
            "INSERT INTO lesson_series (id,lead_id,tutor_id,coordinator_id,package_id,subject,recurrence,day_of_week,time_of_day,timezone,start_date,end_date,total_lessons,status,notes) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (sr_id, lead_id, data.get('tutor_id', ''), user['id'], data.get('package_id', ''),
             data.get('subject', ''), data.get('recurrence', 'weekly'),
             data.get('day_of_week', ''), data.get('time_of_day', ''),
             data.get('timezone', 'Asia/Shanghai'), data.get('start_date', ''),
             data.get('end_date', ''), int(data.get('total_lessons', 0)), 'active', data.get('notes', ''))
        )
        db.commit()
        db.close()
        return json_resp(self, {"id": sr_id, "status": "created"}, 201)

    def _handle_get_series(self, user, sr_id):
        db = get_db()
        r = db.execute(
            "SELECT s.*, l.name as lead_name, u.name as tutor_name, u2.name as coordinator_name, p.package_name "
            "FROM lesson_series s "
            "JOIN leads l ON s.lead_id = l.id "
            "LEFT JOIN users u ON s.tutor_id = u.id "
            "LEFT JOIN users u2 ON s.coordinator_id = u2.id "
            "LEFT JOIN course_packages p ON s.package_id = p.id "
            "WHERE s.id=?", (sr_id,)
        ).fetchone()
        db.close()
        if not r:
            return json_resp(self, {"error": "Not Found"}, 404)
        return json_resp(self, {"series": dict(r)})

    def _handle_update_series(self, user, sr_id, data):
        db = get_db()
        r = db.execute("SELECT * FROM lesson_series WHERE id=?", (sr_id,)).fetchone()
        if not r:
            db.close()
            return json_resp(self, {"error": "Not Found"}, 404)
        updates = {}
        for field in ('tutor_id', 'package_id', 'subject', 'recurrence', 'day_of_week',
                      'time_of_day', 'timezone', 'start_date', 'end_date',
                      'total_lessons', 'status', 'notes'):
            if field in data:
                updates[field] = data[field]
        if not updates:
            db.close()
            return json_resp(self, {"error": "No fields to update"}, 400)
        set_clause = ", ".join(f"{k}=?" for k in updates)
        vals = list(updates.values()) + [sr_id]
        db.execute(f"UPDATE lesson_series SET {set_clause} WHERE id=?", vals)
        db.commit()
        db.close()
        _audit(user['id'], 'update', 'series', sr_id, '更新系列排课')
        return json_resp(self, {"status": "updated"})

    def _handle_delete_series(self, user, sr_id):
        db = get_db()
        db.execute("DELETE FROM lesson_series WHERE id=?", (sr_id,))
        db.commit()
        db.close()
        _audit(user['id'], 'delete', 'series', sr_id, '删除系列排课')
        return json_resp(self, {"status": "deleted"})

    def _handle_generate_series_schedules(self, user, sr_id, data):
        """Generate individual schedule entries from a lesson_series recurrence pattern."""
        db = get_db()
        sr = db.execute("SELECT * FROM lesson_series WHERE id=?", (sr_id,)).fetchone()
        if not sr:
            db.close()
            return json_resp(self, {"error": "Series not found"}, 404)
        if sr['status'] != 'active':
            db.close()
            return json_resp(self, {"error": "Series is not active"}, 400)

        import datetime as dtmod
        from datetime import datetime, timedelta

        recurrence = sr['recurrence']  # 'weekly'
        day_of_week_str = sr['day_of_week']  # e.g. '1,4' (Mon=0, Thu=4)
        time_of_day = sr['time_of_day']  # e.g. '21:00'
        tz_str = sr['timezone'] or 'Asia/Shanghai'
        start_date_str = sr['start_date'] or datetime.now().strftime('%Y-%m-%d')
        end_date_str = sr['end_date']
        total_lessons = sr['total_lessons'] or 0
        duration = int(data.get('duration_minutes', 60))
        lead_id = sr['lead_id']
        tutor_id = sr['tutor_id'] or ''
        package_id = sr['package_id'] or ''

        # Parse days of week (support both number format "0,2" and name format "Monday,Wednesday")
        day_nums = []
        DAY_MAP = {'monday':0,'tuesday':1,'wednesday':2,'thursday':3,'friday':4,'saturday':5,'sunday':6}
        for d in day_of_week_str.split(','):
            d = d.strip()
            if not d:
                continue
            if d.isdigit():
                day_nums.append(int(d))
            else:
                day_nums.append(DAY_MAP.get(d.lower(), -1))
        day_nums = [n for n in day_nums if 0 <= n <= 6]

        if not day_nums:
            db.close()
            return json_resp(self, {"error": "未设置星期几 (day_of_week)"}, 400)

        # Parse time
        try:
            hour, minute = map(int, time_of_day.split(':'))
        except:
            hour, minute = 0, 0

        # Parse dates
        try:
            start_date = dtmod.date.fromisoformat(start_date_str)
        except:
            start_date = datetime.now().date()

        if end_date_str:
            try:
                end_date = dtmod.date.fromisoformat(end_date_str)
            except:
                end_date = start_date + timedelta(days=180)
        else:
            end_date = start_date + timedelta(days=180)

        # Limit to reasonable range
        if end_date > start_date + timedelta(days=365):
            end_date = start_date + timedelta(days=365)

        # Find already generated schedules for this series to avoid duplicates
        existing = set()
        existing_rows = db.execute(
            "SELECT start_time FROM schedules WHERE package_id=? AND lead_id=? AND start_time >= ?",
            (package_id, lead_id, start_date_str)
        ).fetchall()
        for row in existing_rows:
            existing.add(row['start_time'])

        # Generate schedule entries
        generated = []
        current = start_date
        lessons_count = 0
        max_generate = min(total_lessons, 52) if total_lessons > 0 else 52  # cap at 52 or total

        while current <= end_date and lessons_count < max_generate:
            weekday = current.weekday()  # Mon=0
            if weekday in day_nums:
                # Build datetime string in local time
                start_dt_str = f"{current.isoformat()}T{hour:02d}:{minute:02d}:00"
                try:
                    st = dtmod.datetime.fromisoformat(start_dt_str)
                    et = st + timedelta(minutes=duration)
                    end_dt_str = et.isoformat()
                except:
                    current += timedelta(days=1)
                    continue

                if start_dt_str in existing:
                    current += timedelta(days=1)
                    continue

                sch_id = gen_id('sch_')
                now = datetime.now().strftime('%Y-%m-%d %H:%M')
                db.execute(
                    "INSERT INTO schedules (id,lead_id,tutor_id,package_id,coordinator_id,type,start_time,end_time,timezone,duration_minutes,status,topic,notes,created_at,updated_at) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (sch_id, lead_id, tutor_id or None, package_id or None,
                     user['id'], 'recurring', start_dt_str, end_dt_str,
                     tz_str, duration, 'pending',
                     sr['subject'] or '', f"系列排课: {sr_id}\n{sr['notes'] or ''}", now, now)
                )
                generated.append(sch_id)
                existing.add(start_dt_str)
                lessons_count += 1
            current += timedelta(days=1)

        db.commit()
        db.close()
        _audit(user['id'], 'generate_series', 'series', sr_id,
               f'批量生成排课: {len(generated)} 节')
        return json_resp(self, {
            "generated": len(generated),
            "schedule_ids": generated,
            "message": f"已生成 {len(generated)} 节课时"
        })

    # ═══════════════════ CONSUMPTIONS ══════════════════════════════════════

    def _handle_list_consumptions(self, user, params):
        db = get_db()
        conditions, args = [], []
        status = params.get('status', [None])[0]
        if status and status != 'all':
            conditions.append("c.status=?")
            args.append(status)
        pkg_id = params.get('package_id', [None])[0]
        if pkg_id:
            conditions.append("c.package_id=?")
            args.append(pkg_id)
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        page, size, offset = get_page_params(params)
        rows = db.execute(
            f"SELECT c.*, l.name as lead_name, p.package_name, u.name as confirmed_by_name, s.topic "
            f"FROM consumption_log c "
            f"JOIN course_packages p ON c.package_id = p.id "
            f"JOIN leads l ON p.lead_id = l.id "
            f"LEFT JOIN users u ON c.confirmed_by = u.id "
            f"LEFT JOIN schedules s ON c.schedule_id = s.id "
            f"{where} ORDER BY c.created_at DESC LIMIT ? OFFSET ?", args + [size, offset]
        ).fetchall()
        total = db.execute(f"SELECT COUNT(*) FROM consumption_log c {where}", args).fetchone()[0]
        db.close()
        return json_resp(self, {"consumptions": [dict(r) for r in rows], "total": total})

    def _handle_confirm_consumption(self, user, cons_id, data):
        hours = data.get('hours_consumed')
        db = get_db()
        cons = db.execute("SELECT * FROM consumption_log WHERE id=?", (cons_id,)).fetchone()
        if not cons:
            db.close()
            return json_resp(self, {"error": "Not Found"}, 404)

        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        consumed = float(hours) if hours is not None else cons['hours_actual']
        if consumed <= 0 and hours is None:
            consumed = cons['hours_scheduled']  # fallback to scheduled

        db.execute(
            "UPDATE consumption_log SET hours_consumed=?, status='confirmed', confirmed_by=?, confirmed_at=?, notes=? WHERE id=?",
            (consumed, user['id'], now, data.get('notes', ''), cons_id)
        )

        # Update package used_hours
        old_consumed = cons['hours_consumed'] or 0
        diff = consumed - old_consumed
        db.execute("UPDATE course_packages SET used_hours = used_hours + ? WHERE id=?",
                   (diff, cons['package_id']))

        # Check if package is exhausted
        pkg = db.execute("SELECT total_hours, used_hours FROM course_packages WHERE id=?", (cons['package_id'],)).fetchone()
        if pkg and pkg['used_hours'] >= pkg['total_hours']:
            db.execute("UPDATE course_packages SET status='exhausted' WHERE id=?", (cons['package_id'],))

        db.commit()
        db.close()
        _audit(user['id'], 'confirm_consumption', 'consumption', cons_id, f'确认消耗: {consumed}h')
        return json_resp(self, {"status": "confirmed", "hours_consumed": consumed})

    # ═══════════════════ TUTOR AVAILABILITY ════════════════════════════════

    def _handle_list_tutors(self, user):
        db = get_db()
        rows = db.execute(
            "SELECT id, name, phone, role, subjects, timezone, rate, is_tutor, classin_account FROM users WHERE is_tutor=1 AND is_active=1 ORDER BY name"
        ).fetchall()

        # Attach availability and schedule count
        tutors = []
        for r in rows:
            t = dict(r)
            av = db.execute(
                "SELECT * FROM tutor_availability WHERE tutor_id=? ORDER BY weekday", (t['id'],)
            ).fetchall()
            t['availability'] = [dict(a) for a in av]

            this_week = (datetime.now() - timedelta(days=datetime.now().weekday())).strftime('%Y-%m-%d')
            sch_count = db.execute(
                "SELECT COUNT(*) FROM schedules WHERE tutor_id=? AND date(start_time) >= ?", (t['id'], this_week)
            ).fetchone()[0]
            t['week_schedule_count'] = sch_count
            tutors.append(t)

        db.close()
        return json_resp(self, {"tutors": tutors})

    def _handle_list_tutor_avail(self, user, params):
        db = get_db()
        tutor_id = params.get('tutor_id', [None])[0]
        where = "WHERE tutor_id=?" if tutor_id else ""
        args = (tutor_id,) if tutor_id else ()
        rows = db.execute(f"SELECT * FROM tutor_availability {where} ORDER BY weekday, start_slot", args).fetchall()
        db.close()
        return json_resp(self, {"availability": [dict(r) for r in rows]})

    def _handle_get_tutor_avail(self, user, avail_id):
        db = get_db()
        row = db.execute("SELECT * FROM tutor_availability WHERE id=?", (avail_id,)).fetchone()
        db.close()
        if not row:
            return json_resp(self, {"error": "Not Found"}, 404)
        return json_resp(self, {"availability": dict(row)})

    def _handle_create_tutor_avail(self, user, data):
        tutor_id = data.get('tutor_id', '').strip()
        if not tutor_id:
            return json_resp(self, {"error": "请选择老师"}, 400)
        av_id = gen_id('av_')
        db = get_db()
        db.execute(
            "INSERT INTO tutor_availability (id,tutor_id,weekday,start_slot,end_slot,timezone,effective_from,effective_until) VALUES (?,?,?,?,?,?,?,?)",
            (av_id, tutor_id, int(data['weekday']), data['start_slot'], data['end_slot'],
             data.get('timezone', 'Asia/Shanghai'), data.get('effective_from', ''), data.get('effective_until', ''))
        )
        db.commit()
        db.close()
        return json_resp(self, {"id": av_id, "status": "created"}, 201)

    def _handle_update_tutor_avail(self, user, avail_id, data):
        db = get_db()
        updates = {}
        for field in ('weekday', 'start_slot', 'end_slot', 'timezone', 'effective_from', 'effective_until'):
            if field in data:
                updates[field] = data[field]
        if not updates:
            db.close()
            return json_resp(self, {"error": "No fields to update"}, 400)
        set_clause = ", ".join(f"{k}=?" for k in updates)
        vals = list(updates.values()) + [avail_id]
        db.execute(f"UPDATE tutor_availability SET {set_clause} WHERE id=?", vals)
        db.commit()
        db.close()
        return json_resp(self, {"status": "updated"})

    def _handle_delete_tutor_avail(self, user, avail_id):
        db = get_db()
        db.execute("DELETE FROM tutor_availability WHERE id=?", (avail_id,))
        db.commit()
        db.close()
        return json_resp(self, {"status": "deleted"})

    # ═══════════════════ AI REPORTS ════════════════════════════════════════

    def _handle_ai_info(self, user):
        """Return AI module status and available report types."""
        if not _AI_AVAILABLE:
            return json_resp(self, {
                "available": False,
                "message": "AI模块未加载，请确认 .env 中已配置 DEEPSEEK_API_KEY",
            })
        types = []
        for k, v in REPORT_TYPES.items():
            types.append({"type": k, "name": v['name']})
        return json_resp(self, {
            "available": True,
            "report_types": types,
        })

    def _handle_ai_report_types(self, user):
        """Return detailed info about available report types (for frontend forms)."""
        if not _AI_AVAILABLE:
            return json_resp(self, {"error": "AI模块未加载"}, 503)
        from sm_system.ai_reports import _get_available_fields
        return json_resp(self, {"types": _get_available_fields()})

    def _handle_ai_generate(self, user, report_type, data):
        """Generate an AI report of the given type for a student.

        POST body:
          report_type: 'post_class' | 'feedback' | 'academic_report'
          lead_id: required
          schedule_id: optional (for post_class)
        """
        if not _AI_AVAILABLE:
            return json_resp(self, {"error": "AI模块未加载，请先配置 DEEPSEEK_API_KEY"}, 503)

        lead_id = data.get('lead_id', '').strip()
        if not lead_id:
            return json_resp(self, {"error": "请选择学生"}, 400)

        schedule_id = data.get('schedule_id', '').strip()
        if report_type == 'post_class' and not schedule_id:
            return json_resp(self, {"error": "课后反馈需要指定课程"}, 400)

        # Gather context from DB
        db = get_db()
        lead = db.execute("SELECT * FROM leads WHERE id=?", (lead_id,)).fetchone()
        if not lead:
            db.close()
            return json_resp(self, {"error": "学生不存在"}, 404)
        lead = dict(lead)

        # Gather activities
        activities = [dict(r) for r in db.execute(
            "SELECT a.*, u.name as user_name FROM activities a "
            "LEFT JOIN users u ON a.user_id = u.id "
            "WHERE a.lead_id=? ORDER BY a.created_at DESC LIMIT 30",
            (lead_id,)
        ).fetchall()]

        # Gather packages
        packages = [dict(r) for r in db.execute(
            "SELECT * FROM course_packages WHERE lead_id=? ORDER BY created_at DESC",
            (lead_id,)
        ).fetchall()]

        # Gather recent schedules
        recent_schedules = [dict(r) for r in db.execute(
            "SELECT s.*, u.name as tutor_name FROM schedules s "
            "LEFT JOIN users u ON s.tutor_id = u.id "
            "WHERE s.lead_id=? ORDER BY s.start_time DESC LIMIT 20",
            (lead_id,)
        ).fetchall()]

        # Gather recent consumptions
        recent_consumptions = [dict(r) for r in db.execute(
            "SELECT c.*, s.topic FROM consumption_log c "
            "LEFT JOIN schedules s ON c.schedule_id = s.id "
            "JOIN course_packages p ON c.package_id = p.id "
            "WHERE p.lead_id=? ORDER BY c.created_at DESC LIMIT 20",
            (lead_id,)
        ).fetchall()]

        db.close()

        # Build context dict
        context = {
            'student_name': lead['name'],
            'country': lead.get('country', ''),
            'grade': lead.get('grade', ''),
            'subject': lead.get('subject', ''),
            'status': lead.get('status', ''),
            'activities': activities,
            'packages': packages,
            'recent_schedules': recent_schedules,
            'recent_consumptions': recent_consumptions,
        }

        # For post_class, add schedule-specific data
        if report_type == 'post_class' and schedule_id:
            db = get_db()
            sched = db.execute(
                "SELECT s.*, u.name as tutor_name FROM schedules s "
                "LEFT JOIN users u ON s.tutor_id = u.id WHERE s.id=?",
                (schedule_id,)
            ).fetchone()
            db.close()
            if sched:
                sched = dict(sched)
                context.update({
                    'topic': sched.get('topic', ''),
                    'lesson_time': sched.get('start_time', ''),
                    'duration_minutes': sched.get('duration_minutes', 60),
                    'classin_summary': data.get('classin_summary', sched.get('classin_summary', '')),
                    'teacher_notes': data.get('teacher_notes', sched.get('teacher_feedback', '')),
                })

        # Override with any data passed in the request body
        for field in ('classin_summary', 'teacher_notes', 'topic', 'lesson_time', 'duration_minutes'):
            if field in data:
                context[field] = data[field]

        # Call AI
        result = generate_report(report_type, context)

        if 'error' in result:
            return json_resp(self, {"error": result['error']}, 500)

        # Save to DB
        report_id = gen_id('r_')
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db = get_db()
        db.execute(
            "INSERT INTO ai_reports (id,lead_id,report_type,schedule_id,title,content,raw_data,status,created_by,created_at,updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (report_id, lead_id, report_type, schedule_id or None,
             result['title'], result['content'],
             json.dumps(context, ensure_ascii=False)[:2000],
             'draft', user['id'], now, now)
        )
        db.commit()
        db.close()

        _audit(user['id'], 'ai_generate', 'ai_report', report_id, f'AI生成{report_type}报告')

        return json_resp(self, {
            "id": report_id,
            "title": result['title'],
            "content": result['content'],
            "status": "draft",
        }, 201)

    def _handle_list_ai_reports(self, user, params):
        """List AI reports, optionally filtered by lead_id."""
        db = get_db()
        lead_id = params.get('lead_id', [None])[0]
        report_type = params.get('report_type', [None])[0]
        conditions, args = [], []

        if lead_id:
            conditions.append("r.lead_id=?")
            args.append(lead_id)
        if report_type:
            conditions.append("r.report_type=?")
            args.append(report_type)

        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        rows = db.execute(
            f"SELECT r.*, l.name as lead_name, u.name as created_by_name "
            f"FROM ai_reports r "
            f"JOIN leads l ON r.lead_id = l.id "
            f"LEFT JOIN users u ON r.created_by = u.id "
            f"{where} ORDER BY r.created_at DESC LIMIT 50",
            args
        ).fetchall()
        db.close()
        return json_resp(self, {"reports": [dict(r) for r in rows]})

    def _handle_get_ai_report(self, user, report_id):
        db = get_db()
        row = db.execute(
            "SELECT r.*, l.name as lead_name, u.name as created_by_name "
            "FROM ai_reports r "
            "JOIN leads l ON r.lead_id = l.id "
            "LEFT JOIN users u ON r.created_by = u.id "
            "WHERE r.id=?", (report_id,)
        ).fetchone()
        db.close()
        if not row:
            return json_resp(self, {"error": "Not Found"}, 404)
        return json_resp(self, {"report": dict(row)})

    def _handle_update_ai_report(self, user, report_id, data):
        """Update report title/content or just save edits."""
        db = get_db()
        updates = {}
        for field in ('title', 'content', 'status'):
            if field in data:
                updates[field] = data[field]
        if not updates:
            db.close()
            return json_resp(self, {"error": "No fields to update"}, 400)
        updates['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        set_clause = ", ".join(f"{k}=?" for k in updates)
        vals = list(updates.values()) + [report_id]
        db.execute(f"UPDATE ai_reports SET {set_clause} WHERE id=?", vals)
        db.commit()
        db.close()
        _audit(user['id'], 'update', 'ai_report', report_id, '更新AI报告')
        return json_resp(self, {"status": "updated"})

    def _handle_publish_ai_report(self, user, report_id, data):
        """Publish a draft report."""
        db = get_db()
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db.execute("UPDATE ai_reports SET status='published', updated_at=? WHERE id=?",
                   (now, report_id))
        db.commit()
        db.close()
        _audit(user['id'], 'publish', 'ai_report', report_id, '发布AI报告')
        return json_resp(self, {"status": "published"})

    # ═══════════════════ Module: Financial Settlement ═══════════════════════════

    def _handle_list_settlements(self, user, params):
        """List tutor settlements with filters."""
        db = get_db()
        conditions = []
        vals = []
        tutor_id = params.get('tutor_id', [None])[0]
        status = params.get('status', [None])[0]
        period = params.get('period', [None])[0]
        if tutor_id:
            conditions.append("s.tutor_id=?")
            vals.append(tutor_id)
        if status:
            conditions.append("s.status=?")
            vals.append(status)
        if period:
            conditions.append("s.period_start >= ?")
            vals.append(period + "-01")
        w = " AND ".join(conditions) if conditions else "1=1"
        rows = db.execute(
            f"SELECT s.*, u.name as tutor_name FROM tutor_settlements s "
            "LEFT JOIN users u ON s.tutor_id = u.id "
            f"WHERE {w} ORDER BY s.created_at DESC", vals
        ).fetchall()
        db.close()
        return json_resp(self, {"settlements": [dict(r) for r in rows]})

    def _handle_get_settlement(self, user, settlement_id):
        db = get_db()
        row = db.execute(
            "SELECT s.*, u.name as tutor_name FROM tutor_settlements s "
            "LEFT JOIN users u ON s.tutor_id = u.id WHERE s.id=?", (settlement_id,)
        ).fetchone()
        db.close()
        if not row:
            return json_resp(self, {"error": "Not Found"}, 404)
        return json_resp(self, {"settlement": dict(row)})

    def _handle_calculate_settlements(self, user):
        """Auto-generate pending settlements from completed schedules."""
        db = get_db()
        now = datetime.now().strftime('%Y-%m-%d')
        # Last month period
        period_start = (datetime.now().replace(day=1) - timedelta(days=1)).replace(day=1).strftime('%Y-%m-%d')
        period_end = (datetime.now().replace(day=1) - timedelta(days=1)).strftime('%Y-%m-%d')

        completed = db.execute(
            "SELECT s.tutor_id, COUNT(*) as total_lessons, "
            "COALESCE(SUM(s.duration_minutes), 0)/60.0 as total_hours, "
            "u.rate "
            "FROM schedules s "
            "JOIN users u ON s.tutor_id = u.id "
            "WHERE s.status='completed' AND s.tutor_id IS NOT NULL "
            "AND s.start_time >= ? AND s.start_time < ? "
            "GROUP BY s.tutor_id",
            (period_start, (datetime.now().replace(day=1)).strftime('%Y-%m-%d'))
        ).fetchall()

        created = 0
        for row in completed:
            rate = row['rate'] or 0
            gross = round(row['total_hours'] * rate, 2)
            net = gross
            sid = gen_id('stl_')
            db.execute(
                "INSERT OR IGNORE INTO tutor_settlements "
                "(id, tutor_id, period_start, period_end, total_lessons, total_hours, rate, gross_amount, net_amount, created_by) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (sid, row['tutor_id'], period_start, period_end,
                 row['total_lessons'], row['total_hours'], rate, gross, net, user['id'])
            )
            created += 1
        db.commit()
        db.close()
        _audit(user['id'], 'calculate', 'settlements', '', f'生成{created}条结算单')
        return json_resp(self, {"created": created, "period_start": period_start, "period_end": period_end})

    def _handle_update_settlement(self, user, settlement_id, data):
        """Update settlement status (approve/pay)."""
        status = data.get('status', '')
        if status not in ('approved', 'paid'):
            return json_resp(self, {"error": "Invalid status"}, 400)
        db = get_db()
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        if status == 'paid':
            db.execute("UPDATE tutor_settlements SET status=?, paid_at=? WHERE id=?",
                       (status, now, settlement_id))
        else:
            db.execute("UPDATE tutor_settlements SET status=? WHERE id=?", (status, settlement_id))
        db.commit()
        # Notify the settlement tutor
        s = db.execute("SELECT tutor_id FROM tutor_settlements WHERE id=?", (settlement_id,)).fetchone()
        if s:
            self._create_notification_simple(
                db, settlement_id, 'settlement',
                f'结算单状态更新为{status}',
                f'结算单 {settlement_id} 状态已变更为 {status}',
                'settlement', s['tutor_id']
            )
        db.close()
        _audit(user['id'], 'update', 'settlement', settlement_id, f'结算单状态→{status}')
        return json_resp(self, {"status": status})

    # ─── Commissions ──────────────────────────────────────────────────────────

    def _handle_list_commissions(self, user, params):
        db = get_db()
        conditions = []
        vals = []
        user_id = params.get('user_id', [None])[0]
        contract_id = params.get('contract_id', [None])[0]
        status = params.get('status', [None])[0]
        if user_id:
            conditions.append("c.user_id=?")
            vals.append(user_id)
        if contract_id:
            conditions.append("c.contract_id=?")
            vals.append(contract_id)
        if status:
            conditions.append("c.status=?")
            vals.append(status)
        w = " AND ".join(conditions) if conditions else "1=1"
        rows = db.execute(
            f"SELECT c.*, u.name as user_name, ct.contract_no "
            "FROM commissions c "
            "LEFT JOIN users u ON c.user_id = u.id "
            "LEFT JOIN contracts ct ON c.contract_id = ct.id "
            f"WHERE {w} ORDER BY c.created_at DESC", vals
        ).fetchall()
        db.close()
        return json_resp(self, {"commissions": [dict(r) for r in rows]})

    def _handle_get_commission(self, user, commission_id):
        db = get_db()
        row = db.execute(
            "SELECT c.*, u.name as user_name, ct.contract_no "
            "FROM commissions c "
            "LEFT JOIN users u ON c.user_id = u.id "
            "LEFT JOIN contracts ct ON c.contract_id = ct.id "
            "WHERE c.id=?", (commission_id,)
        ).fetchone()
        db.close()
        if not row:
            return json_resp(self, {"error": "Not Found"}, 404)
        return json_resp(self, {"commission": dict(row)})

    def _handle_calculate_commissions(self, user):
        """Auto-generate commissions from contracts."""
        db = get_db()
        rate = 0.05  # default 5%
        created = 0
        contracts = db.execute(
            "SELECT c.*, l.assignee_id FROM contracts c "
            "JOIN leads l ON c.lead_id = l.id "
            "WHERE c.status='active' AND l.assignee_id IS NOT NULL"
        ).fetchall()
        for contract in contracts:
            if not contract['assignee_id']:
                continue
            cid = gen_id('comm_')
            amount = round(contract['total_amount'] * rate, 2)
            try:
                db.execute(
                    "INSERT OR IGNORE INTO commissions "
                    "(id, contract_id, user_id, contract_amount, commission_rate, commission_amount) "
                    "VALUES (?,?,?,?,?,?)",
                    (cid, contract['id'], contract['assignee_id'],
                     contract['total_amount'], rate, amount)
                )
                created += 1
            except Exception:
                pass
        db.commit()
        db.close()
        _audit(user['id'], 'calculate', 'commissions', '', f'生成{created}条提成')
        return json_resp(self, {"created": created, "rate": rate})

    def _handle_update_commission(self, user, commission_id, data):
        status = data.get('status', '')
        if status not in ('approved', 'paid'):
            return json_resp(self, {"error": "Invalid status"}, 400)
        db = get_db()
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        if status == 'paid':
            db.execute("UPDATE commissions SET status=?, paid_at=? WHERE id=?",
                       (status, now, commission_id))
        else:
            db.execute("UPDATE commissions SET status=? WHERE id=?", (status, commission_id))
        db.commit()
        db.close()
        _audit(user['id'], 'update', 'commission', commission_id, f'提成状态→{status}')
        return json_resp(self, {"status": status})

    # ─── Finance Report ───────────────────────────────────────────────────────

    def _handle_finance_report(self, user, params):
        """Monthly finance report: income, expenses, gross profit."""
        db = get_db()
        year = params.get('year', [datetime.now().strftime('%Y')])[0]

        # Monthly income from contracts signed
        income_rows = db.execute(
            "SELECT strftime('%m', signed_at) as month, "
            "COALESCE(SUM(total_amount),0) as income "
            "FROM contracts WHERE strftime('%Y', signed_at)=? AND status!='cancelled' "
            "GROUP BY month ORDER BY month", (year,)
        ).fetchall()

        # Monthly expenses from settlements (paid)
        expense_rows = db.execute(
            "SELECT strftime('%m', paid_at) as month, "
            "COALESCE(SUM(net_amount),0) as expenses "
            "FROM tutor_settlements WHERE strftime('%Y', paid_at)=? AND status='paid' "
            "GROUP BY month ORDER BY month", (year,)
        ).fetchall()

        # Commissions paid
        commission_rows = db.execute(
            "SELECT strftime('%m', paid_at) as month, "
            "COALESCE(SUM(commission_amount),0) as commissions "
            "FROM commissions WHERE strftime('%Y', paid_at)=? AND status='paid' "
            "GROUP BY month ORDER BY month", (year,)
        ).fetchall()

        # Build monthly map
        months = [f"{i:02d}" for i in range(1, 13)]
        income_map = {r['month']: r['income'] for r in income_rows}
        expense_map = {r['month']: r['expenses'] for r in expense_rows}
        commission_map = {r['month']: r['commissions'] for r in commission_rows}

        report = []
        for m in months:
            inc = float(income_map.get(m, 0))
            exp = float(expense_map.get(m, 0))
            comm = float(commission_map.get(m, 0))
            total_exp = exp + comm
            report.append({
                "month": m,
                "income": inc,
                "expenses": exp,
                "commissions": comm,
                "total_expenses": total_exp,
                "gross_profit": round(inc - total_exp, 2)
            })
        db.close()
        return json_resp(self, {"year": year, "report": report})

    # ═══════════════════ Module: Calendar + Reschedule ═══════════════════════════

    def _handle_calendar_view(self, user, params):
        """Return schedules grouped by date for week/month view."""
        db = get_db()
        week_start = params.get('week_start', [None])[0]
        if not week_start:
            # Default to current week starting Monday
            today = datetime.now()
            week_start = (today - timedelta(days=today.weekday())).strftime('%Y-%m-%d')
        week_end = (datetime.strptime(week_start, '%Y-%m-%d') + timedelta(days=6)).strftime('%Y-%m-%d')

        role = user['role']
        conditions = ["s.start_time >= ?", "s.start_time <= ?"]
        vals = [week_start + " 00:00", week_end + " 23:59"]
        if role in ('cs', 'consultant', 'academic'):
            conditions.append("(s.coordinator_id=? OR l.assignee_id=?)")
            vals.extend([user['id'], user['id']])

        w = " AND ".join(conditions)
        rows = db.execute(
            f"SELECT s.*, l.name as lead_name, u.name as tutor_name "
            "FROM schedules s "
            "JOIN leads l ON s.lead_id = l.id "
            "LEFT JOIN users u ON s.tutor_id = u.id "
            f"WHERE {w} ORDER BY s.start_time", vals
        ).fetchall()

        days = {}
        for i in range(7):
            d = (datetime.strptime(week_start, '%Y-%m-%d') + timedelta(days=i)).strftime('%Y-%m-%d')
            days[d] = {"date": d, "weekday": i, "schedules": []}

        for row in rows:
            d = row['start_time'][:10]
            if d in days:
                days[d]["schedules"].append({
                    "id": row['id'],
                    "lead_name": row['lead_name'],
                    "tutor_name": row['tutor_name'] or '',
                    "start_time": row['start_time'][11:16],
                    "end_time": row['end_time'][11:16],
                    "status": row['status'],
                    "timezone": row['timezone'],
                    "topic": row['topic'] or ''
                })
        db.close()
        return json_resp(self, {"week_start": week_start, "days": list(days.values())})

    def _handle_reschedule_schedule(self, user, schedule_id, data):
        """Reschedule: update time + log reason + notify."""
        new_start = data.get('new_start_time', '').strip()
        reason = data.get('reason', '').strip()
        if not new_start:
            return json_resp(self, {"error": "请选择新的上课时间"}, 400)

        db = get_db()
        schedule = db.execute("SELECT * FROM schedules WHERE id=?", (schedule_id,)).fetchone()
        if not schedule:
            db.close()
            return json_resp(self, {"error": "Not Found"}, 404)

        schedule = dict(schedule)
        orig_start = schedule['start_time']
        # Parse duration
        try:
            dur = schedule['duration_minutes'] or 60
        except:
            dur = 60
        new_end = (datetime.strptime(new_start, '%Y-%m-%d %H:%M') +
                   timedelta(minutes=dur)).strftime('%Y-%m-%d %H:%M')

        db.execute(
            "UPDATE schedules SET original_start_time=?, start_time=?, end_time=?, "
            "reschedule_reason=?, updated_at=? WHERE id=?",
            (orig_start, new_start, new_end, reason,
             datetime.now().strftime('%Y-%m-%d %H:%M'), schedule_id)
        )
        db.commit()

        # Log activity on lead
        activity_id = gen_id('act_')
        db.execute(
            "INSERT INTO activities (id, lead_id, type, content, created_by) VALUES (?,?,?,?,?)",
            (activity_id, schedule['lead_id'], 'reschedule',
             f"调课: {orig_start[:16]} → {new_start[:16]}, 原因: {reason}", user['id'])
        )
        db.commit()
        db.close()

        _audit(user['id'], 'reschedule', 'schedule', schedule_id,
               f"{orig_start[:16]}→{new_start[:16]}:{reason}")
        return json_resp(self, {"status": "rescheduled",
                                "original_start_time": orig_start,
                                "new_start_time": new_start})

    # ═══════════════════ Module: Tutor Matching ═════════════════════════════════

    def _handle_match_tutors(self, user, params):
        """Smart tutor matching by subject, timezone, availability."""
        subject = params.get('subject', [None])[0]
        timezone = params.get('timezone', [None])[0]
        weekday = params.get('weekday', [None])[0]
        start_time = params.get('start_time', [None])[0]

        db = get_db()
        tutors = db.execute(
            "SELECT * FROM users WHERE is_tutor=1 AND is_active=1"
        ).fetchall()

        matches = []
        for t in tutors:
            t = dict(t)
            score = 50  # base score
            reasons = []

            # Subject match
            try:
                subs = json.loads(t.get('subjects', '[]'))
            except:
                subs = []
            if subject and subs:
                if any(subject.lower() in s.lower() for s in subs):
                    score += 30
                    reasons.append('subject_match')

            # Timezone match
            tz = t.get('timezone', '')
            if timezone and tz:
                # Simple timezone proximity (same zone = 15 pts)
                if timezone == tz:
                    score += 15
                    reasons.append('tz_exact')
                elif timezone[:3] == tz[:3]:  # Same region prefix
                    score += 8
                    reasons.append('tz_close')

            # Availability check
            if weekday and start_time:
                try:
                    wd = int(weekday)
                    avail = db.execute(
                        "SELECT * FROM tutor_availability "
                        "WHERE tutor_id=? AND weekday=? AND start_slot <= ? AND end_slot >= ?",
                        (t['id'], wd, start_time, start_time)
                    ).fetchone()
                    if avail:
                        score += 25
                        reasons.append('avail_match')
                except:
                    pass

            # Rate score (lower rate for same quality = higher score, capped)
            rate = t.get('rate') or 0
            if rate > 0:
                score += max(0, 10 - int(rate / 50))

            matches.append({
                "tutor_id": t['id'],
                "name": t['name'],
                "subjects": t.get('subjects', '[]'),
                "timezone": t.get('timezone', ''),
                "rate": rate,
                "score": min(100, score),
                "reasons": reasons
            })
        db.close()

        # Sort by score descending
        matches.sort(key=lambda x: x['score'], reverse=True)
        return json_resp(self, {"matches": matches[:20]})

    # ═══════════════════ Module: Notifications ═══════════════════════════════════

    def _create_notification_simple(self, db, related_id, ntype, title, content, related_type, user_id):
        """Create a notification record."""
        try:
            nid = gen_id('notif_')
            db.execute(
                "INSERT INTO notifications (id, user_id, type, title, content, related_type, related_id) "
                "VALUES (?,?,?,?,?,?,?)",
                (nid, user_id, ntype, title, content, related_type, related_id)
            )
        except Exception:
            pass

    def _handle_list_notifications(self, user, params):
        db = get_db()
        is_read = params.get('is_read', [None])[0]
        limit = min(100, int(params.get('limit', ['50'])[0]))
        if is_read is not None:
            rows = db.execute(
                "SELECT * FROM notifications WHERE user_id=? AND is_read=? ORDER BY created_at DESC LIMIT ?",
                (user['id'], int(is_read), limit)
            ).fetchall()
        else:
            rows = db.execute(
                "SELECT * FROM notifications WHERE user_id=? ORDER BY created_at DESC LIMIT ?",
                (user['id'], limit)
            ).fetchall()
        # Count unread
        unread = db.execute(
            "SELECT COUNT(*) as cnt FROM notifications WHERE user_id=? AND is_read=0",
            (user['id'],)
        ).fetchone()['cnt']
        db.close()
        return json_resp(self, {"notifications": [dict(r) for r in rows], "unread": unread})

    def _handle_get_notification(self, user, notification_id):
        db = get_db()
        row = db.execute(
            "SELECT * FROM notifications WHERE id=? AND user_id=?",
            (notification_id, user['id'])
        ).fetchone()
        db.close()
        if not row:
            return json_resp(self, {"error": "Not Found"}, 404)
        return json_resp(self, {"notification": dict(row)})

    def _handle_mark_read(self, user, notification_id):
        db = get_db()
        db.execute("UPDATE notifications SET is_read=1 WHERE id=? AND user_id=?",
                   (notification_id, user['id']))
        db.commit()
        db.close()
        return json_resp(self, {"status": "read"})

    def _handle_mark_read_all(self, user):
        db = get_db()
        db.execute("UPDATE notifications SET is_read=1 WHERE user_id=? AND is_read=0",
                   (user['id'],))
        db.commit()
        db.close()
        return json_resp(self, {"status": "all_read"})

    def _handle_check_reminders(self, user):
        """Check for upcoming follow-up reminders and create notifications."""
        db = get_db()
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        today = datetime.now().strftime('%Y-%m-%d')
        created = 0

        # Find activities with upcoming next_action_date within next 2 days
        rows = db.execute(
            "SELECT a.id, a.lead_id, a.next_action, a.next_action_date, "
            "l.name as lead_name, l.assignee_id "
            "FROM activities a JOIN leads l ON a.lead_id = l.id "
            "WHERE a.next_action_date IS NOT NULL AND a.next_action_date != '' "
            "AND a.next_action_date >= ? AND a.next_action_date <= ? "
            "AND l.assignee_id IS NOT NULL "
            "ORDER BY a.next_action_date",
            (today, (datetime.now() + timedelta(days=2)).strftime('%Y-%m-%d'))
        ).fetchall()

        for r in rows:
            assignee_id = r['assignee_id']
            if not assignee_id:
                continue
            # Avoid duplicate notifications
            dup = db.execute(
                "SELECT id FROM notifications WHERE user_id=? AND related_id=? AND type='followup_reminder' AND created_at >= ?",
                (assignee_id, r['lead_id'], today)
            ).fetchone()
            if dup:
                continue
            self._create_notification_simple(db, r['lead_id'], 'followup_reminder',
                f'跟进提醒: {r["lead_name"]}',
                f'计划跟进: {r["next_action"]} (截止 {r["next_action_date"]})',
                'lead', assignee_id)
            created += 1

        db.commit()
        db.close()
        return json_resp(self, {"status": "checked", "reminders_created": created})

    def _check_reminders_silent(self, user):
        """Silent version - creates notifications without sending a response."""
        db = get_db()
        today = datetime.now().strftime('%Y-%m-%d')
        rows = db.execute(
            "SELECT a.id, a.lead_id, a.next_action, a.next_action_date, "
            "l.name as lead_name, l.assignee_id "
            "FROM activities a JOIN leads l ON a.lead_id = l.id "
            "WHERE a.next_action_date IS NOT NULL AND a.next_action_date != '' "
            "AND a.next_action_date >= ? AND a.next_action_date <= ? "
            "AND l.assignee_id IS NOT NULL",
            (today, (datetime.now() + timedelta(days=2)).strftime('%Y-%m-%d'))
        ).fetchall()
        for r in rows:
            assignee_id = r['assignee_id']
            if not assignee_id: continue
            dup = db.execute(
                "SELECT id FROM notifications WHERE user_id=? AND related_id=? AND type='followup_reminder' AND created_at >= ?",
                (assignee_id, r['lead_id'], today)
            ).fetchone()
            if dup: continue
            self._create_notification_simple(db, r['lead_id'], 'followup_reminder',
                f'跟进提醒: {r["lead_name"]}',
                f'计划跟进: {r["next_action"]} (截止 {r["next_action_date"]})',
                'lead', assignee_id)
        db.commit()
        db.close()

    def log_message(self, format, *args):
        if '/api/' in str(args[0]):
            print(f"[TMS] {args[0]} - {args[1]} {args[2]}")

    # ═══════════════════════ EXCEL EXPORT ═══════════════════════════════════

    def _handle_export_excel(self, user):
        """Export all leads as a downloadable Excel file."""
        if user['role'] not in ('admin', 'supervisor'):
            return json_resp(self, {"error": "无权限"}, 403)
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return json_resp(self, {"error": "openpyxl未安装，请先 pip3 install openpyxl"}, 500)

        db = get_db()
        rows = db.execute("""
            SELECT l.*, u.name as assignee_name, ac.name as academic_name
            FROM leads l
            LEFT JOIN users u ON l.assignee_id = u.id
            LEFT JOIN users ac ON l.academic_manager_id = ac.id
            WHERE l.merge_target_id IS NULL
            ORDER BY l.created_at DESC
        """).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "线索总表"

        # Headers matching original Excel format
        headers = ['序号', '姓名', '手机', '微信', '阶段', '国家', '科目',
                   '来源平台', '来源账号', '客户类型', '客户评级', '辅导需求',
                   '跟进顾问', '教务/学管', '线索状态', '备注', '创建时间']
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4f46e5', end_color='4f46e5', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Data rows
        lt_map = {'parent': '家长', 'student': '学生', 'agent': '留学机构'}
        for i, r in enumerate(rows, 1):
            tags = ''
            try: tags = ', '.join(json.loads(r['tags'] or '[]'))
            except: pass
            ws.cell(row=i+1, column=1, value=i).border = thin_border
            ws.cell(row=i+1, column=2, value=r['name']).border = thin_border
            ws.cell(row=i+1, column=3, value=r['phone']).border = thin_border
            ws.cell(row=i+1, column=4, value=r['wechat']).border = thin_border
            ws.cell(row=i+1, column=5, value=r['grade']).border = thin_border
            ws.cell(row=i+1, column=6, value=r['country']).border = thin_border
            ws.cell(row=i+1, column=7, value=r['subject']).border = thin_border
            ws.cell(row=i+1, column=8, value=r['source']).border = thin_border
            ws.cell(row=i+1, column=9, value=r['account_name']).border = thin_border
            ws.cell(row=i+1, column=10, value=lt_map.get(r['lead_type'], r['lead_type'])).border = thin_border
            ws.cell(row=i+1, column=11, value=r['rating']).border = thin_border
            ws.cell(row=i+1, column=12, value=tags).border = thin_border
            ws.cell(row=i+1, column=13, value=r['assignee_name']).border = thin_border
            ws.cell(row=i+1, column=14, value=r['academic_name']).border = thin_border
            status_map = {'pending':'待分配','assigned':'已分配','following':'跟进中','trial':'试听中','enrolled':'已签约','closed':'已关闭','lost':'已流失'}
            ws.cell(row=i+1, column=15, value=status_map.get(r['status'], r['status'])).border = thin_border
            ws.cell(row=i+1, column=16, value=r['notes']).border = thin_border
            ws.cell(row=i+1, column=17, value=r['created_at']).border = thin_border

        # Column widths
        widths = [6, 14, 16, 20, 10, 8, 14, 10, 26, 10, 8, 30, 10, 10, 10, 50, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        db.close()

        # Write to response
        import tempfile
        tmp_path = os.path.join(tempfile.gettempdir(), f'tms_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        wb.save(tmp_path)
        db.close()

        with open(tmp_path, 'rb') as f:
            body = f.read()
        try: os.remove(tmp_path)
        except: pass

        filename = f'tms_leads_{datetime.now().strftime("%Y%m%d")}.xlsx'
        self.send_response(200)
        self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
        self.send_header('Content-Length', str(len(body)))
        self._send_cors_headers()
        self.end_headers()
        self.wfile.write(body)

    # ═══════════════════════ ACADEMIC MANAGEMENT ════════════════════════════

    def _handle_assign_academic(self, user, lead_id, data):
        """Assign an enrolled lead to an academic manager (教务)."""
        db = get_db()
        lead = db.execute("SELECT * FROM leads WHERE id=?", (lead_id,)).fetchone()
        if not lead:
            db.close()
            return json_resp(self, {"error": "线索不存在"}, 404)
        acad_id = data.get('academic_manager_id', '').strip()
        if not acad_id:
            db.close()
            return json_resp(self, {"error": "请选择教务"}, 400)
        acad = db.execute("SELECT id, name FROM users WHERE id=? AND role='academic' AND is_active=1", (acad_id,)).fetchone()
        if not acad:
            db.close()
            return json_resp(self, {"error": "教务不存在或非活跃"}, 400)
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db.execute("UPDATE leads SET academic_manager_id=?, updated_at=? WHERE id=?", (acad_id, now, lead_id))
        # Notify
        self._create_notification_simple(db, lead_id, 'academic_assignment',
            f'学生 {lead["name"]} 已分配给您',
            f'{user["name"]} 将 {lead["name"]} 分配给您跟进',
            'lead', acad_id)
        # Audit
        db.execute(
            "INSERT INTO activities (id,lead_id,user_id,type,content,created_at) VALUES (?,?,?,?,?,?)",
            (gen_id('a_'), lead_id, user['id'], 'note',
             f'系统: 分配给教务 {acad["name"]} 跟进', now)
        )
        db.commit()
        db.close()
        _audit(user['id'], 'assign_academic', 'lead', lead_id, f'分配给教务 {acad_id}')
        return json_resp(self, {"status": "assigned", "academic_name": acad['name']})

    def _handle_academic_dashboard(self, user):
        """教务工作台 - 查看自己管理的学生"""
        if user['role'] not in ('academic', 'admin', 'supervisor'):
            return json_resp(self, {"error": "无权限"}, 403)
        db = get_db()
        if user['role'] == 'academic':
            rows = db.execute(
                "SELECT l.*, u.name as assignee_name FROM leads l "
                "LEFT JOIN users u ON l.assignee_id = u.id "
                "WHERE l.academic_manager_id=? AND l.status IN ('enrolled','following') "
                "AND l.merge_target_id IS NULL ORDER BY l.updated_at DESC",
                (user['id'],)
            ).fetchall()
        else:
            rows = db.execute(
                "SELECT l.*, u.name as assignee_name, ac.name as academic_name FROM leads l "
                "LEFT JOIN users u ON l.assignee_id = u.id "
                "LEFT JOIN users ac ON l.academic_manager_id = ac.id "
                "WHERE l.academic_manager_id IS NOT NULL AND l.academic_manager_id != '' "
                "AND l.status IN ('enrolled','following') "
                "AND l.merge_target_id IS NULL ORDER BY l.updated_at DESC"
            ).fetchall()
        # Low hours alerts
        low = db.execute(
            "SELECT l.id, l.name, p.package_name, p.total_hours, p.used_hours "
            "FROM course_packages p JOIN leads l ON p.lead_id=l.id "
            "WHERE p.status='active' AND (p.used_hours * 1.0 / p.total_hours) >= 0.7 "
            "ORDER BY (p.used_hours * 1.0 / p.total_hours) DESC LIMIT 10"
        ).fetchall()
        db.close()
        return json_resp(self, {
            "students": [dict(r) for r in rows],
            "low_hours_alerts": [dict(r) for r in low]
        })

    # ═══════════════════════ ENROLLED STUDENTS ══════════════════════════════

    def _handle_enrolled_list(self, user, params):
        """Return enrolled students with package info and academic manager."""
        db = get_db()
        rows = db.execute("""
            SELECT l.*, u.name as assignee_name, ac.name as academic_name,
                p.id as pkg_id, p.package_name, p.total_hours, p.used_hours,
                p.status as pkg_status, p.valid_until
            FROM leads l
            LEFT JOIN users u ON l.assignee_id = u.id
            LEFT JOIN users ac ON l.academic_manager_id = ac.id
            LEFT JOIN course_packages p ON p.lead_id = l.id AND p.status = 'active'
            WHERE l.status = 'enrolled' AND l.merge_target_id IS NULL
            ORDER BY l.updated_at DESC
        """).fetchall()
        # Group by lead
        enrolled = {}
        for r in rows:
            lid = r['id']
            if lid not in enrolled:
                enrolled[lid] = dict(r)
                enrolled[lid]['packages'] = []
            if r['pkg_id']:
                enrolled[lid]['packages'].append({
                    'id': r['pkg_id'],
                    'name': r['package_name'],
                    'total': r['total_hours'],
                    'used': r['used_hours'],
                    'status': r['pkg_status'],
                    'valid_until': r['valid_until'],
                })
        db.close()
        return json_resp(self, {"enrolled": list(enrolled.values())})

    # ═══════════════════════ DINGTALK WEBHOOK ═══════════════════════════════

    def _handle_dingtalk_webhook(self, data):
        """Receive DingTalk webhook, parse auto-reply messages, create leads."""
        msgtype = data.get('msgtype', '')
        if msgtype != 'text':
            return json_resp(self, {"msgtype": "text", "text": {"content": "暂不支持该消息类型"}})

        text = data.get('text', {}).get('content', '').strip()
        sender = data.get('senderNick', '')

        # Only process messages containing auto-reply format
        if '自动回复私信' not in text:
            return json_resp(self, {"msgtype": "text", "text": {"content": "收到 ✅"}})

        # Skip if already processed (dedup by msgId)
        msg_id = data.get('msgId', '')
        _last_msg_id = getattr(self, '_last_dingtalk_msg_id', '')
        if msg_id and msg_id == _last_msg_id:
            return json_resp(self, {"msgtype": "text", "text": {"content": "已处理，跳过重复"}})
        self._last_dingtalk_msg_id = msg_id

        # Parse the message
        result = parse_lead_text(text)

        if not result.get('name') and not result.get('phone') and not result.get('wechat'):
            return json_resp(self, {
                "msgtype": "text",
                "text": {"content": f"⚠️ 解析失败，未能识别有效线索信息，请手动录入。\n来源: {sender}"}
            })

        # Check duplicates
        dup_name = ''
        db = get_db()
        if result.get('phone'):
            dup = db.execute(
                "SELECT name FROM leads WHERE phone=? AND status NOT IN ('lost','closed') AND merge_target_id IS NULL",
                (result['phone'],)
            ).fetchone()
            if dup:
                dup_name = dup['name']
        if not dup_name and result.get('wechat'):
            dup = db.execute(
                "SELECT name FROM leads WHERE wechat=? AND status NOT IN ('lost','closed') AND merge_target_id IS NULL",
                (result['wechat'],)
            ).fetchone()
            if dup:
                dup_name = dup['name']

        if dup_name:
            db.close()
            return json_resp(self, {
                "msgtype": "text",
                "text": {"content": f"⚠️ 该线索已存在: {dup_name}，跳过重复创建"}
            })

        # Create lead
        lead_id = gen_id('l_')
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        source = result.get('source') or '钉钉'
        notes = result.get('notes') or '来自钉钉自动回复'
        db.execute(
            "INSERT INTO leads (id,name,phone,wechat,country,grade,subject,source,notes,created_at,updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (lead_id, result['name'], result.get('phone',''), result.get('wechat',''),
             result.get('country',''), result.get('grade',''), result.get('subject',''),
             source, notes, now, now)
        )
        db.execute(
            "INSERT INTO activities (id,lead_id,user_id,type,content,created_at) VALUES (?,?,?,?,?,?)",
            (gen_id('a_'), lead_id, 'u_admin', 'note',
             f'来自钉钉机器人自动创建 (来源: {sender})', now)
        )
        db.commit()
        db.close()

        # Build response
        contact_info = result.get('phone') or result.get('wechat') or ''
        account_info = f" [{result.get('account_name','')}]" if result.get('account_name') else ''
        return json_resp(self, {
            "msgtype": "text",
            "text": {
                "content": f"✅ 线索已创建: {result['name']} ({contact_info}){account_info}\n"
                           f"来源: {source} | 提交人: {sender}"
            }
        })

    # ═══════════════════════ BACKUP & EXPORT ═══════════════════════════════

    def _handle_backup_export(self, user):
        """Export all data as downloadable JSON."""
        if user['role'] not in ('admin', 'supervisor'):
            return json_resp(self, {"error": "无权限"}, 403)
        export = _data_export()
        body = json.dumps(export, ensure_ascii=False, indent=2).encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Disposition', f'attachment; filename="tms_backup_{datetime.now().strftime("%Y%m%d")}.json"')
        self.send_header('Content-Length', str(len(body)))
        self._send_cors_headers()
        self.end_headers()
        self.wfile.write(body)

    def _handle_backup_list(self, user):
        """List available backup files."""
        if user['role'] not in ('admin', 'supervisor'):
            return json_resp(self, {"error": "无权限"}, 403)
        import glob
        files = []
        for f in sorted(glob.glob(os.path.join(BACKUP_DIR, 'sm_*')), reverse=True)[:30]:
            stat = os.stat(f)
            size_mb = stat.st_size / (1024*1024)
            files.append({
                'name': os.path.basename(f),
                'size': f'{size_mb:.1f}MB',
                'date': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M'),
            })
        return json_resp(self, {"backups": files})


# ─── Data Safety ────────────────────────────────────────────────────────────

BACKUP_DIR = os.path.join(BASE_DIR, 'backups')

def _auto_backup():
    """Auto-backup DB before startup and daily thereafter."""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = os.path.join(BACKUP_DIR, f'sm_{ts}.db')
    try:
        # Use SQLite online backup API for safe hot backup
        src = sqlite3.connect(DB_PATH, timeout=10)
        dst = sqlite3.connect(backup_path, timeout=10)
        src.backup(dst, pages=1000)  # 1000 pages per iteration
        dst.close()
        src.close()
        # Compress backups older than 7 days
        import glob
        for f in glob.glob(os.path.join(BACKUP_DIR, 'sm_*.db')):
            age = (time.time() - os.path.getmtime(f)) / 86400
            if age > 7 and not f.endswith('.gz'):
                import gzip, shutil
                with open(f, 'rb') as fin:
                    with gzip.open(f + '.gz', 'wb') as fout:
                        shutil.copyfileobj(fin, fout)
                os.remove(f)
        # Delete backups older than 90 days
        for f in glob.glob(os.path.join(BACKUP_DIR, 'sm_*')):
            age = (time.time() - os.path.getmtime(f)) / 86400
            if age > 90:
                os.remove(f)
        print(f"  [备份] ✅ {backup_path}")
    except Exception as e:
        print(f"  [备份] ⚠️ {e}")


def _integrity_check():
    """Run integrity check on startup."""
    try:
        db = get_db()
        cur = db.execute("PRAGMA integrity_check")
        result = cur.fetchone()[0]
        db.close()
        if result == 'ok':
            print(f"  [校验] ✅ 数据库完整性正常")
        else:
            print(f"  [校验] ❌ 数据库损坏: {result}")
            print(f"  [校验] ⚠️ 建议立即从备份恢复!")
        return result == 'ok'
    except Exception as e:
        print(f"  [校验] ⚠️ {e}")
        return False


def _data_export(format='json'):
    """Export all data as JSON for portable backup."""
    import copy
    db = get_db()
    tables = ['users', 'leads', 'activities', 'course_packages', 'schedules',
              'consumption_log', 'lesson_series', 'pool_return_log',
              'contracts', 'payments', 'referrals', 'tutor_availability',
              'tutor_settlements', 'commissions', 'notifications', 'pool_config']
    export = {'version': SCHEMA_VERSION, 'exported_at': datetime.now().isoformat(), 'data': {}}
    for table in tables:
        try:
            rows = db.execute(f"SELECT * FROM {table}").fetchall()
            export['data'][table] = [dict(r) for r in rows]
        except Exception:
            pass  # table may not exist
    db.close()
    return export


# ─── Entry Point ────────────────────────────────────────────────────────────

def run_server(host='0.0.0.0', port=8766):
    print(f"\n{'='*55}")
    print(f"  TMS — 辅导管理系统 v0.2")
    print(f"  {'='*55}")
    # Step 1: Integrity check
    _integrity_check()
    # Step 2: Auto-backup
    _auto_backup()
    # Step 3: Init DB (with migration)
    init_db()
    # Step 4: Start server
    server = HTTPServer((host, port), TMSHandler)
    print(f"  访问地址: http://localhost:{port}")
    print(f"  {'='*55}")
    print(f"  演示账号:")
    print(f"    管理员:     admin / admin123")
    print(f"    教班主任:   coor01 / 123456")
    print(f"    授课老师:   tut01 / 123456")
    print(f"    客服:       cs01 / 123456")
    print(f"    顾问:       con01 / 123456")
    print(f"    学管:       mgr01 / 123456")
    print(f"    主管:       sup01 / 123456")
    print(f"  {'='*55}\n")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nShutting down...")
        server.shutdown()


if __name__ == '__main__':
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8766
    run_server(port=port)
